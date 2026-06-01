import csv
from datetime import datetime
from decimal import Decimal
from django.db.models.fields.files import FieldFile
from django.contrib import admin
from django.contrib.admin.decorators import register as admin_register
from django.contrib.auth.models import User, Group
from django import forms
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html, mark_safe
from .models import Category, Partner, Product, ProductImage, ProductColorVariant, ProductSizeVariant, Cart, CartItem, BlogPost, Contact, Order, OrderItem, Slider, HomeBanner, ProductReview, SiteLogo, PartnerBanner, PartnerSlider, NavMenu, PartnerNavMenu, SocialMediaLink, Page, SideBanner, LandingPage, ServerFee, Coupon, CouponUsage, CustomOrder, AdminCommission, PartnerWallet, WalletTransaction, WithdrawalRequest, PayoutMethod, WalletSettings, PlatformBalance, DeliveryLog, Wishlist, WishlistItem, Notification, ShippingRule, RefundRequest, ManualPaymentMethod, PosOrder, PosOrderItem, Conversation, Message, ConversationReadStatus, ProductQA, SupportTicket, TicketReply, MedicineProduct, MedicinePosOrder, MedicinePosOrderItem, MedicineSubscription, SubscriptionPackage, WalletRechargeInstruction
from .admin_site import custom_admin_site


class ExcelExportMixin:
    actions = ['export_excel']

    def export_excel(self, request, queryset):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.model._meta.verbose_name_plural.title()
        fields = [f for f in self.model._meta.fields if not f.auto_created and f.name != 'id']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='DEE2E6'),
            right=Side(style='thin', color='DEE2E6'),
            top=Side(style='thin', color='DEE2E6'),
            bottom=Side(style='thin', color='DEE2E6'),
        )
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field.verbose_name.title() if hasattr(field, 'verbose_name') else field.name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        for row_idx, obj in enumerate(queryset, 2):
            for col_idx, field in enumerate(fields, 1):
                val = getattr(obj, field.name)
                if callable(val):
                    val = val()
                if isinstance(val, Decimal):
                    val = float(val)
                if isinstance(val, bool):
                    val = 'Yes' if val else 'No'
                if val is None:
                    val = ''
                if isinstance(val, FieldFile):
                    val = str(val) if val else ''
                if hasattr(field, 'remote_field') and field.remote_field:
                    val = str(val)
                from datetime import timezone as dt_timezone
                if isinstance(val, datetime) and val.tzinfo is not None:
                    val = val.astimezone(dt_timezone.utc).replace(tzinfo=None)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        ws.column_dimensions['A'].width = 5
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.model._meta.model_name}s.xlsx"'
        wb.save(response)
        return response
    export_excel.short_description = 'Export selected to Excel'


class AdminPartnerAddressWidget(forms.Widget):
    template_name = 'admin/widgets/partner_address.html'

    def get_context(self, name, value, attrs=None):
        return super().get_context(name, value, attrs)


class PartnerAdminForm(forms.ModelForm):
    class Meta:
        model = Partner
        fields = '__all__'


# ─── Inline Classes ───

class PartnerSliderInline(admin.TabularInline):
    model = PartnerSlider
    extra = 1


class PartnerBannerInline(admin.TabularInline):
    model = PartnerBanner
    extra = 1


class PartnerNavMenuInline(admin.TabularInline):
    model = PartnerNavMenu
    extra = 1
    fk_name = 'partner'


class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1


class ProductColorVariantInline(admin.TabularInline):
    model = ProductColorVariant
    extra = 1
    fields = ['color_name', 'color_code', 'price_adjustment', 'stock', 'image', 'order']


class ProductSizeVariantInline(admin.TabularInline):
    model = ProductSizeVariant
    extra = 1
    fields = ['size_name', 'price_adjustment', 'stock', 'order']


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ['product', 'product_name', 'price', 'quantity']

    def has_add_permission(self, request, obj=None):
        return False


# ─── Auth: User & Group ───

class StaffUserCreationForm(forms.ModelForm):
    password1 = forms.CharField(label='Password', widget=forms.PasswordInput, required=False,
                                help_text='Leave blank to auto-generate a random password.')
    password2 = forms.CharField(label='Confirm Password', widget=forms.PasswordInput, required=False)
    role = forms.ChoiceField(choices=[
        ('', 'No preset'),
        ('order_manager', 'Order Manager — manage orders, delivery, refunds'),
        ('content_manager', 'Content Manager — manage products, blogs, pages, banners'),
        ('finance_manager', 'Finance Manager — manage commissions, wallets, withdrawals'),
        ('support_agent', 'Support Agent — view orders, manage contacts, reviews'),
    ], required=False, label='Role Preset')

    class Meta:
        model = User
        fields = ['username', 'email', 'first_name', 'last_name', 'is_active']

    def save(self, commit=True):
        user = super().save(commit=False)
        password = self.cleaned_data.get('password1')
        if password:
            user.set_password(password)
        else:
            random_pw = User.objects.make_random_password(12)
            user.set_password(random_pw)
            user._generated_password = random_pw
        if commit:
            user.save()
            self.save_m2m()
            role = self.cleaned_data.get('role')
            if role:
                group_map = {
                    'order_manager': 'Order Manager',
                    'content_manager': 'Content Manager',
                    'finance_manager': 'Finance Manager',
                    'support_agent': 'Support Agent',
                }
                group_name = group_map.get(role)
                if group_name:
                    group, _ = Group.objects.get_or_create(name=group_name)
                    user.groups.add(group)
        return user


@admin_register(User, site=custom_admin_site)
class UserAdmin(admin.ModelAdmin):
    add_form = StaffUserCreationForm
    list_display = ['username', 'email', 'first_name', 'last_name', 'is_staff', 'is_active', 'role_badges', 'date_joined']
    list_filter = ['is_staff', 'is_active', 'is_superuser', 'groups']
    search_fields = ['username', 'email', 'first_name', 'last_name']
    date_hierarchy = 'date_joined'
    ordering = ['-date_joined']
    fieldsets = [
        (None, {'fields': ['username', 'email', 'password']}),
        ('Personal Info', {'fields': ['first_name', 'last_name']}),
        ('Permissions', {'fields': ['is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions']}),
        ('Important Dates', {'fields': ['last_login', 'date_joined']}),
    ]
    readonly_fields = ['password', 'last_login', 'date_joined']
    filter_horizontal = ['groups', 'user_permissions']
    actions = ['activate_users', 'deactivate_users', 'add_to_group']

    def get_form(self, request, obj=None, **kwargs):
        defaults = {}
        if obj is None:
            defaults['form'] = self.add_form
            defaults['fieldsets'] = [
                (None, {'fields': ['username', 'email', 'password1', 'password2']}),
                ('Personal Info', {'fields': ['first_name', 'last_name']}),
                ('Role Preset', {'fields': ['role'], 'description': 'Select a role to auto-assign permissions and groups.'}),
                ('Status', {'fields': ['is_active']}),
            ]
        return super().get_form(request, obj, **defaults)

    def get_readonly_fields(self, request, obj=None):
        if not obj:
            return []
        return self.readonly_fields

    def role_badges(self, obj):
        badges = ''
        for group in obj.groups.all()[:3]:
            colors = {
                'Order Manager': '#17a2b8',
                'Content Manager': '#28a745',
                'Finance Manager': '#6f42c1',
                'Support Agent': '#ffc107',
            }
            c = colors.get(group.name, '#6c757d')
            badges += f'<span style="background:{c};color:#fff;padding:1px 8px;border-radius:8px;font-size:10px;margin:1px;display:inline-block;">{group.name}</span> '
        if obj.is_superuser:
            badges += '<span style="background:#dc3545;color:#fff;padding:1px 8px;border-radius:8px;font-size:10px;">Superuser</span>'
        return mark_safe(badges or '<span style="color:#999;">—</span>')
    role_badges.short_description = 'Roles'

    def activate_users(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'{updated} user(s) activated.')
    activate_users.short_description = 'Activate selected users'

    def deactivate_users(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'{updated} user(s) deactivated.')
    deactivate_users.short_description = 'Deactivate selected users'

    def add_to_group(self, request, queryset):
        group_name = request.POST.get('group_name', '')
        if group_name:
            group, _ = Group.objects.get_or_create(name=group_name)
            for user in queryset:
                user.groups.add(group)
            self.message_user(request, f'{queryset.count()} user(s) added to "{group_name}".')
        else:
            self.message_user(request, 'Please provide a group name.', level='ERROR')
    add_to_group.short_description = 'Add selected to group...'

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        if not change and hasattr(form, '_generated_password'):
            from django.core.mail import send_mail
            from django.conf import settings
            try:
                send_mail(
                    'Your Admin Account Credentials',
                    f'Hi {form.instance.username},\n\nYour admin account has been created.\nUsername: {form.instance.username}\nPassword: {form._generated_password}\n\nPlease login and change your password.',
                    settings.DEFAULT_FROM_EMAIL,
                    [form.instance.email],
                )
            except Exception:
                pass


@admin_register(Group, site=custom_admin_site)
class GroupAdmin(admin.ModelAdmin):
    list_display = ['name', 'user_count', 'permission_count']
    search_fields = ['name']
    filter_horizontal = ['permissions']

    def user_count(self, obj):
        return obj.user_set.count()
    user_count.short_description = 'Users'

    def permission_count(self, obj):
        return obj.permissions.count()
    permission_count.short_description = 'Permissions'


# ─── Activity Log ───

from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION
from django.contrib.contenttypes.models import ContentType

@admin_register(LogEntry, site=custom_admin_site)
class ActivityLogAdmin(admin.ModelAdmin):
    list_display = ['action_time', 'user', 'action_flag_display', 'content_type', 'object_repr', 'change_message_short']
    list_filter = ['action_time', 'user', 'action_flag', 'content_type']
    search_fields = ['user__username', 'object_repr', 'change_message']
    date_hierarchy = 'action_time'
    readonly_fields = [f.name for f in LogEntry._meta.fields]
    list_select_related = ['user', 'content_type']
    ordering = ['-action_time']

    def action_flag_display(self, obj):
        labels = {ADDITION: 'Created', CHANGE: 'Updated', DELETION: 'Deleted'}
        colors = {ADDITION: '#28a745', CHANGE: '#17a2b8', DELETION: '#dc3545'}
        label = labels.get(obj.action_flag, str(obj.action_flag))
        c = colors.get(obj.action_flag, '#6c757d')
        return format_html('<span style="background:{};color:#fff;padding:2px 10px;border-radius:8px;font-size:10px;">{}</span>', c, label)
    action_flag_display.short_description = 'Action'

    def change_message_short(self, obj):
        msg = obj.change_message or ''
        return msg[:80] + '...' if len(msg) > 80 else msg
    change_message_short.short_description = 'Message'

    def has_add_permission(self, request): return False
    def has_change_permission(self, request, obj=None): return False
    def has_delete_permission(self, request, obj=None): return False


# ─── Category ───

@admin_register(Category, site=custom_admin_site)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'slug', 'product_count', 'parent']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name']
    list_filter = ['parent']
    change_list_template = 'admin/store/category/change_list.html'

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('products', 'children').select_related('parent')

    def product_count(self, obj):
        return obj.products.count()
    product_count.short_description = 'Products'

    def changelist_view(self, request, extra_context=None):
        qs = self.get_queryset(request)
        parent_categories = [c for c in qs if c.children.exists()]
        orphan_categories = [c for c in qs if c.parent is None and not c.children.exists()]
        extra_context = extra_context or {}
        extra_context['parent_categories'] = parent_categories
        extra_context['orphan_categories'] = orphan_categories
        return super().changelist_view(request, extra_context=extra_context)


# ─── Partner ───

@admin_register(Partner, site=custom_admin_site)
class PartnerAdmin(ExcelExportMixin, admin.ModelAdmin):
    form = PartnerAdminForm
    change_list_template = 'admin/store/partner/change_list.html'
    list_display = ['name', 'type_badges', 'phone', 'address_division', 'blocked_badge', 'logo_preview', 'profile_preview']
    list_filter = ['is_dealer', 'is_seller', 'is_union_agent', 'shop_style', 'address_division', 'address_district', 'blocked']
    search_fields = ['name', 'phone', 'address', 'description']
    inlines = [PartnerBannerInline, PartnerNavMenuInline]
    fieldsets = [
        (None, {'fields': ['user', 'name', 'slug', 'description', 'phone']}),
        ('Address (Bangladesh)', {'fields': ['address_street', 'address_division', 'address_district', 'address_upazila'], 'classes': ['bd-address-picker']}),
        ('Verification', {'fields': ['voter_id']}),
        ('Branding', {'fields': ['logo', 'profile_image', 'banner']}),
        ('Settings', {'fields': ['is_dealer', 'is_seller', 'is_union_agent', 'show_products', 'shop_style', 'has_medicine_access', 'medicine_pos_enabled']}),
        ('Admin', {'fields': ['blocked'], 'classes': ['collapse'], 'description': 'Block deletion to prevent accidental removal of this partner and its related data.'}),
    ]

    def type_badges(self, obj):
        badges = ''
        if obj.is_seller:
            badges += '<span style="background:#17a2b8;color:#fff;padding:2px 8px;border-radius:10px;font-size:10px;margin-right:4px;">Seller</span> '
        if obj.is_dealer:
            badges += '<span style="background:#6f42c1;color:#fff;padding:2px 8px;border-radius:10px;font-size:10px;margin-right:4px;">Dealer</span> '
        if obj.is_union_agent:
            badges += '<span style="background:#28a745;color:#fff;padding:2px 8px;border-radius:10px;font-size:10px;margin-right:4px;">Union Agent</span> '
        return mark_safe(badges or '<span style="color:#999;font-size:11px;">-</span>')
    type_badges.short_description = 'Type'

    def logo_preview(self, obj):
        if obj.logo:
            return format_html('<img src="{}" height="36" style="border-radius:4px;object-fit:contain;" />', obj.logo.url)
        return ''
    logo_preview.short_description = 'Logo'

    def profile_preview(self, obj):
        if obj.profile_image:
            return format_html('<img src="{}" height="36" style="border-radius:50%;object-fit:cover;" />', obj.profile_image.url)
        return ''
    profile_preview.short_description = 'Photo'

    def blocked_badge(self, obj):
        if obj.blocked:
            return mark_safe('<span style="background:#dc3545;color:#fff;padding:2px 8px;border-radius:10px;font-size:10px;">Blocked</span>')
        return mark_safe('<span style="background:#28a745;color:#fff;padding:2px 8px;border-radius:10px;font-size:10px;">Unblocked</span>')
    blocked_badge.short_description = 'Delete'
    blocked_badge.admin_order_field = 'blocked'

    def delete_queryset(self, request, queryset):
        """Bypass admin cascade-permission checks by using ORM-level delete.
        Skips blocked partners.
        """
        from django.db.models import ProtectedError
        blocked = queryset.filter(blocked=True)
        unblocked = queryset.filter(blocked=False)
        blocked_count = blocked.count()
        try:
            deleted, _ = unblocked.delete()
            if blocked_count:
                self.message_user(request, f'{deleted} partner(s) deleted. {blocked_count} blocked partner(s) skipped.', level='WARNING')
            else:
                self.message_user(request, f'{deleted} partner(s) deleted.')
        except ProtectedError as e:
            self.message_user(request, f'Cannot delete: protected objects prevent deletion ({e}).', level='ERROR')

    def has_delete_permission(self, request, obj=None):
        if obj is not None and obj.blocked:
            return False
        return super().has_delete_permission(request, obj)

    actions = ['block_selected', 'unblock_selected']

    def block_selected(self, request, queryset):
        updated = queryset.update(blocked=True)
        self.message_user(request, f'{updated} partner(s) blocked.')
    block_selected.short_description = 'Block deletion for selected partners'

    def unblock_selected(self, request, queryset):
        updated = queryset.update(blocked=False)
        self.message_user(request, f'{updated} partner(s) unblocked.')
    unblock_selected.short_description = 'Unblock deletion for selected partners'

    # ─── 30-Column Export (round-trip with bulk import) ───

    def export_excel(self, request, queryset):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        HEADERS = [
            'User', 'Name', 'Slug', 'Logo', 'Profile Image', 'Banner',
            'Description', 'Phone', 'Address', 'Street / Village / House / Flat',
            'Division', 'District', 'Upazila', 'Voter Id Card',
            'Is Dealer', 'Is Seller', 'Is Union Agent', 'Is Active',
            'Dealer (Parent)', 'Meta Title', 'Meta Description',
            'Show Products On Website',
            'Show "All Over Available Product" Section On Store Page',
            'Show "Random Product List" Section On Store Page',
            'Show Slider On Store Page', 'Shop Style',
            'Medicine Catalog Access', 'Medicine Pos Enabled',
            'Created', 'Updated',
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Partners'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='DEE2E6'),
            right=Side(style='thin', color='DEE2E6'),
            top=Side(style='thin', color='DEE2E6'),
            bottom=Side(style='thin', color='DEE2E6'),
        )

        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        def _img(val):
            return val.url if val else ''

        def _bool(val):
            return 'Yes' if val else 'No'

        def _dt(val):
            if val:
                return val.strftime('%Y-%m-%d %H:%M:%S.%f')[:23]
            return ''

        for row_idx, obj in enumerate(queryset, 2):
            values = [
                '',
                obj.name,
                obj.slug,
                _img(obj.logo),
                _img(obj.profile_image),
                _img(obj.banner),
                obj.description,
                obj.phone,
                obj.address,
                obj.address_street,
                obj.address_division,
                obj.address_district,
                obj.address_upazila,
                _img(obj.voter_id),
                _bool(obj.is_dealer),
                _bool(obj.is_seller),
                _bool(obj.is_union_agent),
                _bool(obj.is_active),
                obj.parent.name if obj.parent else '',
                obj.meta_title,
                obj.meta_description,
                _bool(obj.show_products),
                _bool(obj.show_all_products_section),
                _bool(obj.show_random_products_section),
                _bool(obj.show_slider),
                obj.shop_style,
                _bool(obj.has_medicine_access),
                _bool(obj.medicine_pos_enabled),
                _dt(obj.created),
                _dt(obj.updated),
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="partners.xlsx"'
        wb.save(response)
        return response
    export_excel.short_description = 'Export selected to Excel (30-column import format)'


# ─── Product ───

@admin_register(Product, site=custom_admin_site)
class ProductAdmin(ExcelExportMixin, admin.ModelAdmin):
    list_display = ['name', 'sku', 'barcode', 'price_display', 'cost_price_display', 'profit_display', 'stock_badge', 'is_published_badge', 'category', 'partner', 'available', 'label_badge', 'image_preview']
    list_per_page = 50
    list_filter = ['is_published', 'available', 'label', 'category', 'partner', 'created']
    list_editable = ['available']
    search_fields = ['name', 'sku', 'barcode', 'description', 'category__name', 'partner__name']
    prepopulated_fields = {'slug': ('name',)}
    date_hierarchy = 'created'
    inlines = [ProductImageInline, ProductColorVariantInline, ProductSizeVariantInline]
    actions = ['mark_available', 'mark_unavailable', 'approve_products', 'reject_products']

    def price_display(self, obj):
        html = f'<strong>৳{obj.price}</strong>'
        if obj.old_price:
            html += f' <span style="color:#999;text-decoration:line-through;font-size:11px;">৳{obj.old_price}</span>'
        return mark_safe(html)
    price_display.short_description = 'Price'

    def stock_badge(self, obj):
        if obj.stock == 0:
            return mark_safe('<span style="background:#dc3545;color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;">Out</span>')
        elif obj.stock <= 5:
            return format_html('<span style="background:#ffc107;color:#333;padding:2px 10px;border-radius:10px;font-size:11px;">{}</span>', obj.stock)
        return format_html('<span style="color:#28a745;font-weight:600;">{}</span>', obj.stock)
    stock_badge.short_description = 'Stock'

    def label_badge(self, obj):
        if not obj.label:
            return ''
        colors = {'new': '#17a2b8', 'hot': '#dc3545', 'discounted': '#28a745'}
        return format_html('<span style="background:{};color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;">{}</span>', colors[obj.label], obj.get_label_display())
    label_badge.short_description = 'Label'

    def available_status(self, obj):
        if obj.available:
            return mark_safe('<span style="color:#28a745;font-weight:600;">✓ Active</span>')
        return mark_safe('<span style="color:#dc3545;">✗ Hidden</span>')
    available_status.short_description = 'Status'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" height="36" style="border-radius:4px;object-fit:cover;" />', obj.image.url)
        return ''
    image_preview.short_description = 'Image'

    def mark_available(self, request, queryset):
        updated = queryset.update(available=True)
        self.message_user(request, f'{updated} product(s) marked as available.')
    mark_available.short_description = 'Mark selected as available'

    def mark_unavailable(self, request, queryset):
        updated = queryset.update(available=False)
        self.message_user(request, f'{updated} product(s) marked as unavailable.')
    mark_unavailable.short_description = 'Mark selected as unavailable'

    def is_published_badge(self, obj):
        if obj.is_published:
            return mark_safe('<span style="color:#28a745;font-weight:600;">Published</span>')
        return mark_safe('<span style="background:#ffc107;color:#333;padding:2px 10px;border-radius:10px;font-size:11px;font-weight:600;">Pending</span>')
    is_published_badge.short_description = 'Status'
    is_published_badge.admin_order_field = 'is_published'

    def cost_price_display(self, obj):
        if obj.cost_price is not None:
            return f'৳{obj.cost_price}'
        return '—'
    cost_price_display.short_description = 'Cost Price'

    def profit_display(self, obj):
        p = obj.profit
        if p is not None:
            color = '#28a745' if p >= 0 else '#dc3545'
            return format_html('<span style="color:{};font-weight:600;">৳{}</span>', color, p)
        return '—'
    profit_display.short_description = 'Profit'

    def approve_products(self, request, queryset):
        count = queryset.update(is_published=True)
        for p in queryset.filter(is_published=True):
            if p.partner and p.partner.user:
                Notification.objects.create(
                    user=p.partner.user,
                    partner=p.partner,
                    type='order_status',
                    title=f'Product Approved: {p.name}',
                    message=f'Your product "{p.name}" has been approved and is now live.',
                    link='/dashboard/products/',
                )
        self.message_user(request, f'{count} product(s) approved.')
    approve_products.short_description = 'Approve selected products (publish)'

    def reject_products(self, request, queryset):
        count = queryset.update(is_published=False)
        self.message_user(request, f'{count} product(s) marked as pending.')
    reject_products.short_description = 'Mark selected as pending (unpublish)'


# ─── Cart ───

@admin_register(Cart, site=custom_admin_site)
class CartAdmin(admin.ModelAdmin):
    list_display = ['id', 'user_info', 'items_count', 'total', 'is_active_badge', 'created']
    list_filter = ['is_active', 'created']
    search_fields = ['user__username', 'session_id']
    date_hierarchy = 'created'

    def user_info(self, obj):
        if obj.user:
            return format_html('<a href="{}">{}</a>', obj.user.pk, obj.user.username)
        return format_html('<span style="color:#999;">Guest ({})</span>', obj.session_id[:12] if obj.session_id else '-')
    user_info.short_description = 'User'

    def items_count(self, obj): return obj.total_items()
    items_count.short_description = 'Items'

    def total(self, obj): return format_html('<strong>৳{}</strong>', obj.total_amount())
    total.short_description = 'Total'

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;">Active</span>' if obj.is_active else '<span style="color:#999;">Completed</span>')
    is_active_badge.short_description = 'Status'


# ─── CartItem ───

@admin_register(CartItem, site=custom_admin_site)
class CartItemAdmin(admin.ModelAdmin):
    list_display = ['cart_link', 'product', 'quantity', 'subtotal']
    search_fields = ['product__name', 'cart__id']

    def cart_link(self, obj): return format_html('<a href="{}">Cart #{}</a>', obj.cart.pk, obj.cart.id)
    cart_link.short_description = 'Cart'

    def subtotal(self, obj): return format_html('৳{}', obj.subtotal())
    subtotal.short_description = 'Subtotal'


# ─── Blog ───

@admin_register(BlogPost, site=custom_admin_site)
class BlogPostAdmin(ExcelExportMixin, admin.ModelAdmin):
    list_display = ['title', 'author', 'views', 'comments_count', 'created_date', 'image_preview']
    list_filter = ['created', 'author']
    search_fields = ['title', 'content', 'author']
    prepopulated_fields = {'slug': ('title',)}
    date_hierarchy = 'created'

    def comments_count(self, obj): return obj.comments
    comments_count.short_description = 'Comments'

    def created_date(self, obj): return obj.created.strftime('%d %b %Y')
    created_date.short_description = 'Published'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" height="36" style="border-radius:4px;object-fit:cover;" />', obj.image.url)
        return ''
    image_preview.short_description = 'Image'


# ─── Contact ───

@admin_register(Contact, site=custom_admin_site)
class ContactAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'subject_short', 'created', 'message_preview']
    list_filter = ['created']
    search_fields = ['name', 'email', 'subject', 'message']
    date_hierarchy = 'created'
    readonly_fields = ['name', 'email', 'phone', 'subject', 'message', 'created']

    def subject_short(self, obj): return (obj.subject[:30] + '...') if len(obj.subject) > 30 else obj.subject
    subject_short.short_description = 'Subject'

    def message_preview(self, obj): return (obj.message[:60] + '...') if len(obj.message) > 60 else obj.message
    message_preview.short_description = 'Message'

    def has_add_permission(self, request): return False


# ─── Order ───

class DeliveryLogInline(admin.TabularInline):
    model = DeliveryLog
    extra = 0
    readonly_fields = ['status', 'note', 'created_by', 'created']
    can_delete = False
    max_num = 0
    verbose_name = 'Delivery Status Update'
    verbose_name_plural = 'Delivery Timeline'

    def has_add_permission(self, request, obj=None): return False


@admin_register(Order, site=custom_admin_site)
class OrderAdmin(ExcelExportMixin, admin.ModelAdmin):
    list_display = ['id', 'name', 'phone', 'total_display', 'payment_status_badge', 'payment_method', 'status_badge', 'delivery_badge', 'courier_badge', 'tracking_id', 'created_date', 'action_buttons']
    list_filter = ['status', 'delivery_status', 'payment_status', 'payment_method', 'courier_name', 'created']
    search_fields = ['name', 'phone', 'address', 'id', 'tracking_id', 'admin_notes']
    date_hierarchy = 'created'
    inlines = [OrderItemInline, DeliveryLogInline]
    actions = ['export_csv', 'mark_processing', 'mark_completed', 'mark_cancelled', 'mark_paid', 'mark_refunded', 'mark_picked_up', 'mark_in_transit', 'mark_out_for_delivery', 'mark_delivered', 'mark_failed']
    change_form_template = 'admin/store/order/change_form.html'

    fieldsets = [
        (None, {'fields': ['name', 'phone', 'address', 'shipping_address', 'delivery_notes', 'delivery_partner']}),
        ('Pricing', {'fields': ['subtotal', 'discount_amount', 'fees_total', 'total']}),
        ('Payment', {'fields': ['payment_method', 'payment_status']}),
        ('Status & Delivery', {'fields': ['status', 'delivery_status', 'coupon', 'tracking_id', 'courier_name', 'estimated_delivery']}),
        ('Internal Notes', {'fields': ['admin_notes', 'applied_fees']}),
    ]
    readonly_fields = ['subtotal', 'discount_amount', 'fees_total', 'total', 'coupon', 'applied_fees']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-csv/<int:order_id>/', self.admin_site.admin_view(self.export_single_csv), name='export-single-order-csv'),
            path('print-invoice/<int:order_id>/', self.admin_site.admin_view(self.print_invoice), name='print-order-invoice'),
        ]
        return custom_urls + urls

    def export_single_csv(self, request, order_id):
        order = get_object_or_404(Order, id=order_id)
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="order_{order.id}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(['Field', 'Value'])
        writer.writerow(['Order ID', order.id])
        writer.writerow(['Date', order.created.strftime('%d %b %Y, %I:%M %p')])
        writer.writerow(['Customer Name', order.name])
        writer.writerow(['Phone', order.phone])
        writer.writerow(['Address', order.address])
        writer.writerow(['Payment Method', order.payment_method])
        writer.writerow(['Payment Status', order.get_payment_status_display()])
        writer.writerow(['Order Status', order.get_status_display()])
        writer.writerow(['Subtotal', f'৳{order.subtotal}'])
        writer.writerow(['Discount', f'৳{order.discount_amount}'])
        writer.writerow(['Fees', f'৳{order.fees_total}'])
        writer.writerow(['Total', f'৳{order.total}'])
        writer.writerow(['Tracking ID', order.tracking_id or 'N/A'])
        writer.writerow([])
        writer.writerow(['Item', 'Qty', 'Price', 'Total'])
        for item in order.items.all():
            writer.writerow([item.product_name, item.quantity, f'৳{item.price}', f'৳{item.price * item.quantity}'])
        return response

    def print_invoice(self, request, order_id):
        order = get_object_or_404(Order, id=order_id)
        return render(request, 'admin/store/print_invoice.html', {'order': order})

    def action_buttons(self, obj):
        csv_url = reverse('admin:export-single-order-csv', args=[obj.id])
        print_url = reverse('admin:print-order-invoice', args=[obj.id])
        return mark_safe(
            f'<a href="{csv_url}" class="button" style="background:#17a2b8;color:#fff;padding:3px 8px;border-radius:4px;font-size:11px;text-decoration:none;margin-right:4px;">CSV</a>'
            f'<a href="{print_url}" target="_blank" class="button" style="background:#28a745;color:#fff;padding:3px 8px;border-radius:4px;font-size:11px;text-decoration:none;">Print</a>'
        )
    action_buttons.short_description = 'Actions'

    def total_display(self, obj):
        html = f'<strong>৳{obj.total}</strong>'
        if obj.discount_amount:
            html += f'<br><span style="color:#999;font-size:10px;">(disc: ৳{obj.discount_amount})</span>'
        return mark_safe(html)
    total_display.short_description = 'Total'

    def payment_status_badge(self, obj):
        colors = {'unpaid': ('#dc3545', '#fff'), 'paid': ('#28a745', '#fff'), 'refunded': ('#ffc107', '#333')}
        bg, fg = colors.get(obj.payment_status, ('#6c757d', '#fff'))
        return format_html('<span style="background:{};color:{};padding:2px 8px;border-radius:8px;font-size:10px;font-weight:600;">{}</span>', bg, fg, obj.get_payment_status_display())
    payment_status_badge.short_description = 'Payment'

    def created_date(self, obj): return obj.created.strftime('%d %b %Y, %I:%M %p')
    created_date.short_description = 'Date'

    def status_badge(self, obj):
        colors = {'pending': ('#ffc107', '#333'), 'processing': ('#17a2b8', '#fff'), 'completed': ('#28a745', '#fff'), 'cancelled': ('#dc3545', '#fff')}
        bg, fg = colors.get(obj.status, ('#6c757d', '#fff'))
        return format_html('<span style="background:{};color:{};padding:3px 12px;border-radius:12px;font-size:11px;font-weight:600;">{}</span>', bg, fg, obj.get_status_display())
    status_badge.short_description = 'Status'

    def delivery_badge(self, obj):
        colors = {
            'pending': '#ffc107', 'picked_up': '#17a2b8', 'in_transit': '#2196F3',
            'out_for_delivery': '#FF9800', 'delivered': '#28a745', 'failed': '#dc3545',
        }
        bg = colors.get(obj.delivery_status, '#6c757d')
        return format_html('<span style="background:{};color:#fff;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:600;">{}</span>', bg, obj.get_delivery_status_display())
    delivery_badge.short_description = 'Delivery'

    def courier_badge(self, obj):
        if not obj.courier_name:
            return '—'
        colors = {
            'steadfast': '#2196F3', 'pathao': '#E91E63', 'sundarban': '#4CAF50',
            'ecourier': '#FF9800', 'redx': '#9C27B0', 'sa_paribahan': '#00BCD4',
            'other': '#607D8B',
        }
        bg = colors.get(obj.courier_name, '#607D8B')
        return format_html('<span style="background:{};color:#fff;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:600;">{}</span>', bg, obj.get_courier_name_display())
    courier_badge.short_description = 'Courier'

    def export_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="orders_export.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(['Order ID', 'Date', 'Customer', 'Phone', 'Address', 'Payment Method', 'Payment Status', 'Order Status', 'Subtotal', 'Discount', 'Fees', 'Total', 'Items', 'Tracking ID'])
        for obj in queryset:
            items_str = ' | '.join([f'{i.product_name} x{i.quantity} (৳{i.price})' for i in obj.items.all()])
            writer.writerow([
                obj.id, obj.created.strftime('%d/%m/%Y %H:%M'), obj.name, obj.phone, obj.address,
                obj.payment_method, obj.get_payment_status_display(), obj.get_status_display(),
                f'৳{obj.subtotal}', f'৳{obj.discount_amount}', f'৳{obj.fees_total}', f'৳{obj.total}',
                items_str, obj.tracking_id or ''
            ])
        self.message_user(request, f'{queryset.count()} order(s) exported.')
        return response
    export_csv.short_description = 'Export selected to CSV'

    def mark_processing(self, request, queryset):
        updated = queryset.update(status='processing')
        self.message_user(request, f'{updated} order(s) marked as processing.')
    mark_processing.short_description = 'Mark selected as processing'

    def mark_completed(self, request, queryset):
        for order in queryset:
            order.status = 'completed'
            order.save(update_fields=['status'])
            _credit_partner_wallets(order)
        updated = queryset.count()
        self.message_user(request, f'{updated} order(s) marked as completed and partner wallets credited.')
    mark_completed.short_description = 'Mark selected as completed and credit wallets'

    def mark_cancelled(self, request, queryset):
        updated = queryset.update(status='cancelled')
        self.message_user(request, f'{updated} order(s) marked as cancelled.')
    mark_cancelled.short_description = 'Mark selected as cancelled'

    def _update_delivery_status(self, request, queryset, status, label):
        now = timezone.now()
        for obj in queryset:
            obj.delivery_status = status
            obj.save(update_fields=['delivery_status'])
            DeliveryLog.objects.create(order=obj, status=status, created_by=request.user)
        self.message_user(request, f'{queryset.count()} order(s) marked as {label}.')

    def mark_picked_up(self, request, queryset):
        self._update_delivery_status(request, queryset, 'picked_up', 'Picked Up')
    mark_picked_up.short_description = 'Mark delivery as Picked Up'

    def mark_in_transit(self, request, queryset):
        self._update_delivery_status(request, queryset, 'in_transit', 'In Transit')
    mark_in_transit.short_description = 'Mark delivery as In Transit'

    def mark_out_for_delivery(self, request, queryset):
        self._update_delivery_status(request, queryset, 'out_for_delivery', 'Out for Delivery')
    mark_out_for_delivery.short_description = 'Mark delivery as Out for Delivery'

    def mark_delivered(self, request, queryset):
        self._update_delivery_status(request, queryset, 'delivered', 'Delivered')
    mark_delivered.short_description = 'Mark delivery as Delivered'

    def mark_failed(self, request, queryset):
        self._update_delivery_status(request, queryset, 'failed', 'Delivery Failed')
    mark_failed.short_description = 'Mark delivery as Failed'

    def mark_paid(self, request, queryset):
        updated = queryset.update(payment_status='paid')
        self.message_user(request, f'{updated} order(s) marked as paid.')
    mark_paid.short_description = 'Mark selected as Paid'

    def mark_refunded(self, request, queryset):
        updated = queryset.update(payment_status='refunded')
        self.message_user(request, f'{updated} order(s) marked as refunded.')
    mark_refunded.short_description = 'Mark selected as Refunded'

    def has_add_permission(self, request): return False


# ─── Home Banner ───

@admin_register(HomeBanner, site=custom_admin_site)
class HomeBannerAdmin(admin.ModelAdmin):
    list_display = ['title', 'order', 'is_active_badge', 'image_preview']
    list_editable = ['order']
    list_filter = ['is_active']
    search_fields = ['title']

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" height="50" style="border-radius:4px;object-fit:cover;" />', obj.image.url)
        return ''

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;font-weight:600;">✓ Active</span>' if obj.is_active else '<span style="color:#dc3545;">✗ Inactive</span>')
    is_active_badge.short_description = 'Status'


# ─── Product Review ───

@admin_register(ProductReview, site=custom_admin_site)
class ProductReviewAdmin(admin.ModelAdmin):
    list_display = ['product', 'user', 'rating_stars', 'comment_short', 'created']
    list_filter = ['rating', 'created']
    search_fields = ['product__name', 'user__username', 'comment']
    date_hierarchy = 'created'
    readonly_fields = ['product', 'user', 'rating', 'comment', 'created']
    change_list_template = 'admin/store/productreview/change_list.html'

    def rating_stars(self, obj):
        stars = ''.join(['★' if i <= obj.rating else '☆' for i in range(1, 6)])
        return format_html('<span style="color:#ffc107;letter-spacing:2px;">{}</span>', stars)
    rating_stars.short_description = 'Rating'

    def comment_short(self, obj): return (obj.comment[:50] + '...') if len(obj.comment) > 50 else obj.comment
    comment_short.short_description = 'Comment'

    def has_add_permission(self, request): return False

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('product', 'user')

    def changelist_view(self, request, extra_context=None):
        from collections import OrderedDict
        from django.db.models import Avg
        qs = self.get_queryset(request)
        grouped = OrderedDict()
        product_ratings = {}
        for review in qs:
            product = review.product
            if product not in grouped:
                grouped[product] = []
                product_ratings[product] = []
            grouped[product].append(review)
            product_ratings[product].append(review.rating)
        for product in grouped:
            ratings = product_ratings[product]
            product.avg_rating = sum(ratings) / len(ratings) if ratings else 0
        extra_context = extra_context or {}
        extra_context['grouped_reviews'] = grouped
        return super().changelist_view(request, extra_context=extra_context)


# ─── Slider ───

@admin_register(Slider, site=custom_admin_site)
class SliderAdmin(admin.ModelAdmin):
    list_display = ['title', 'subtitle_short', 'order', 'is_active_badge', 'image_preview']
    list_editable = ['order']
    list_filter = ['is_active']
    search_fields = ['title', 'subtitle']

    def subtitle_short(self, obj): return (obj.subtitle[:30] + '...') if len(obj.subtitle) > 30 else obj.subtitle
    subtitle_short.short_description = 'Subtitle'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" height="50" style="border-radius:4px;object-fit:cover;" />', obj.image.url)
        return ''

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;font-weight:600;">✓ Active</span>' if obj.is_active else '<span style="color:#dc3545;">✗ Inactive</span>')
    is_active_badge.short_description = 'Status'


# ─── Site Settings ───

@admin_register(SiteLogo, site=custom_admin_site)
class SiteLogoAdmin(admin.ModelAdmin):
    list_display = ['site_name', 'site_tagline', 'logo_preview', 'updated']
    search_fields = ['site_name', 'site_tagline']

    def logo_preview(self, obj):
        if obj.logo:
            return format_html('<img src="{}" height="50" style="border-radius:4px;object-fit:contain;" />', obj.logo.url)
        return ''
    logo_preview.short_description = 'Logo'

    def has_add_permission(self, request): return not SiteLogo.objects.exists()

    def get_fieldsets(self, request, obj=None):
        if obj is None:
            return super().get_fieldsets(request, obj)
        return [
            ('Site Identity', {'fields': ['site_name', 'site_tagline']}),
            ('Branding', {'fields': ['logo']}),
            ('Checkout Settings', {'fields': ['partner_delivery_enabled']}),
        ]


# ─── Partner Nav Menu ───

@admin_register(PartnerNavMenu, site=custom_admin_site)
class PartnerNavMenuAdmin(admin.ModelAdmin):
    list_display = ['title', 'partner', 'parent', 'order', 'is_active_badge']
    list_filter = ['partner', 'is_active']
    search_fields = ['title', 'partner__name']
    list_select_related = ['partner', 'parent']
    change_list_template = 'admin/store/partnermenu/change_list.html'

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;">Active</span>' if obj.is_active else '<span style="color:#999;">Inactive</span>')
    is_active_badge.short_description = 'Status'

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('partner', 'parent').prefetch_related('children')

    def changelist_view(self, request, extra_context=None):
        from collections import OrderedDict
        qs = self.get_queryset(request)
        grouped = OrderedDict()
        for item in qs:
            partner = item.partner
            if partner not in grouped:
                grouped[partner] = []
            if item.parent is None:
                grouped[partner].append(item)
        for partner in grouped:
            items = grouped[partner]
            items.sort(key=lambda x: x.order)
        extra_context = extra_context or {}
        extra_context['grouped_menus'] = grouped
        return super().changelist_view(request, extra_context=extra_context)


# ─── Social Media ───

@admin_register(SocialMediaLink, site=custom_admin_site)
class SocialMediaLinkAdmin(admin.ModelAdmin):
    list_display = ['name', 'url', 'icon_class', 'order', 'is_active_badge']
    list_editable = ['order']
    list_filter = ['is_active']
    search_fields = ['name', 'url']

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;">Active</span>' if obj.is_active else '<span style="color:#999;">Inactive</span>')
    is_active_badge.short_description = 'Status'


# ─── Side Banner ───

@admin_register(SideBanner, site=custom_admin_site)
class SideBannerAdmin(admin.ModelAdmin):
    list_display = ['title', 'partner', 'order', 'is_active_badge', 'image_preview']
    list_editable = ['order']
    list_filter = ['is_active', 'partner']
    search_fields = ['title']
    change_list_template = 'admin/store/sidebanner/change_list.html'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" height="50" style="border-radius:4px;object-fit:cover;" />', obj.image.url)
        return ''

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;font-weight:600;">✓ Active</span>' if obj.is_active else '<span style="color:#dc3545;">✗ Inactive</span>')
    is_active_badge.short_description = 'Status'

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('partner')

    def changelist_view(self, request, extra_context=None):
        from collections import OrderedDict
        qs = self.get_queryset(request)
        grouped = OrderedDict()
        for banner in qs:
            partner = banner.partner
            if partner not in grouped:
                grouped[partner] = []
            grouped[partner].append(banner)
        extra_context = extra_context or {}
        extra_context['grouped_banners'] = grouped
        return super().changelist_view(request, extra_context=extra_context)


# ─── Page ───

@admin_register(Page, site=custom_admin_site)
class PageAdmin(admin.ModelAdmin):
    list_display = ['title', 'slug', 'template_badge', 'show_in_nav', 'order', 'is_active_badge', 'updated']
    list_editable = ['order']
    list_filter = ['is_active', 'template_name', 'show_in_header', 'show_in_footer']
    search_fields = ['title', 'content', 'meta_title', 'meta_description', 'meta_keywords']
    prepopulated_fields = {'slug': ('title',)}
    date_hierarchy = 'created'
    readonly_fields = ['created', 'updated']

    fieldsets = [
        (None, {'fields': ['title', 'slug', 'content']}),
        ('SEO Settings', {'fields': ['meta_title', 'meta_description', 'meta_keywords'], 'classes': ['collapse']}),
        ('Appearance', {'fields': ['featured_image', 'template_name', 'custom_css', 'custom_js'], 'classes': ['collapse']}),
        ('Navigation', {'fields': ['show_in_header', 'show_in_footer', 'order']}),
        ('Status', {'fields': ['is_active', 'created', 'updated']}),
    ]

    def template_badge(self, obj):
        colors = {'default': '#6c757d', 'full_width': '#17a2b8', 'landing': '#28a745'}
        c = colors.get(obj.template_name, '#6c757d')
        return mark_safe(f'<span style="background:{c};color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;">{obj.get_template_name_display()}</span>')
    template_badge.short_description = 'Template'

    def show_in_nav(self, obj):
        badges = []
        if obj.show_in_header:
            badges.append('<span style="background:#007bff;color:#fff;padding:1px 6px;border-radius:8px;font-size:10px;">Header</span>')
        if obj.show_in_footer:
            badges.append('<span style="background:#28a745;color:#fff;padding:1px 6px;border-radius:8px;font-size:10px;">Footer</span>')
        return mark_safe(' '.join(badges) or '<span style="color:#999;">-</span>')
    show_in_nav.short_description = 'Nav'

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;font-weight:600;">✓ Active</span>' if obj.is_active else '<span style="color:#dc3545;">✗ Inactive</span>')
    is_active_badge.short_description = 'Status'

    class Media:
        css = {'all': ['https://cdn.jsdelivr.net/npm/codemirror@5/lib/codemirror.css']}
        js = ['https://cdn.jsdelivr.net/npm/codemirror@5/lib/codemirror.js',
              'https://cdn.jsdelivr.net/npm/codemirror@5/mode/xml/xml.js',
              'https://cdn.jsdelivr.net/npm/codemirror@5/mode/css/css.js',
              'https://cdn.jsdelivr.net/npm/codemirror@5/mode/javascript/javascript.js',
              'https://cdn.jsdelivr.net/npm/codemirror@5/mode/htmlmixed/htmlmixed.js']


# ─── Landing Page ───

@admin_register(LandingPage, site=custom_admin_site)
class LandingPageAdmin(admin.ModelAdmin):
    list_display = ['title', 'slug', 'order', 'is_active_badge', 'updated']
    list_editable = ['order']
    list_filter = ['is_active']
    search_fields = ['title', 'meta_title', 'meta_description']
    prepopulated_fields = {'slug': ('title',)}
    date_hierarchy = 'created'
    readonly_fields = ['created', 'updated']
    filter_horizontal = ['featured_products']

    fieldsets = [
        (None, {'fields': ['title', 'slug']}),
        ('Hero Section', {'fields': ['hero_title', 'hero_subtitle', 'hero_background', 'hero_overlay_color', 'hero_cta_text', 'hero_cta_link']}),
        ('About Section', {'fields': ['about_heading', 'about_content', 'about_image']}),
        ('Featured Products', {'fields': ['featured_products_heading', 'featured_products']}),
        ('Partners', {'fields': ['show_partners', 'partners_heading']}),
        ('Statistics Counters', {'fields': ['show_stats', 'stat_1_label', 'stat_1_value', 'stat_2_label', 'stat_2_value', 'stat_3_label', 'stat_3_value', 'stat_4_label', 'stat_4_value'], 'classes': ['collapse']}),
        ('Bottom CTA', {'fields': ['bottom_cta_heading', 'bottom_cta_text', 'bottom_cta_link']}),
        ('Theme', {'fields': ['primary_color', 'secondary_color', 'custom_css']}),
        ('SEO', {'fields': ['meta_title', 'meta_description', 'meta_keywords'], 'classes': ['collapse']}),
        ('Settings', {'fields': ['is_active', 'order', 'created', 'updated']}),
    ]

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;font-weight:600;">✓ Active</span>' if obj.is_active else '<span style="color:#dc3545;">✗ Inactive</span>')
    is_active_badge.short_description = 'Status'


# ─── Nav Menu ───

@admin_register(NavMenu, site=custom_admin_site)
class NavMenuAdmin(admin.ModelAdmin):
    list_display = ['title', 'url', 'url_type', 'order', 'is_active_badge', 'login_required', 'logout_required']
    list_editable = ['order']
    list_filter = ['is_active', 'url_type', 'login_required', 'logout_required']
    search_fields = ['title', 'url']

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;font-weight:600;">✓ Active</span>' if obj.is_active else '<span style="color:#dc3545;">✗ Inactive</span>')
    is_active_badge.short_description = 'Status'


# ─── Server Fee ───

@admin_register(ServerFee, site=custom_admin_site)
class ServerFeeAdmin(admin.ModelAdmin):
    list_display = ['name', 'fee_type_badge', 'value_display', 'min_order_display', 'is_active_badge', 'order', 'has_link']
    list_editable = ['order']
    list_filter = ['fee_type', 'is_active']
    search_fields = ['name', 'description', 'link_text']
    prepopulated_fields = {'slug': ('name',)}
    fieldsets = [
        (None, {'fields': ['name', 'slug', 'description']}),
        ('Fee Configuration', {'fields': ['fee_type', 'value', 'min_order_amount']}),
        ('Link / Info Page', {'fields': ['link_url', 'link_text'], 'classes': ['collapse'], 'description': 'Optionally link to a policy or information page about this fee.'}),
        ('Settings', {'fields': ['is_active', 'order']}),
    ]

    def fee_type_badge(self, obj):
        colors = {'fixed': '#17a2b8', 'percentage': '#6f42c1'}
        c = colors.get(obj.fee_type, '#6c757d')
        return mark_safe(f'<span style="background:{c};color:#fff;padding:2px 8px;border-radius:8px;font-size:10px;">{obj.get_fee_type_display()}</span>')
    fee_type_badge.short_description = 'Type'

    def value_display(self, obj):
        if obj.fee_type == 'percentage':
            return f'{obj.value}%'
        return f'৳{obj.value}'
    value_display.short_description = 'Value'

    def min_order_display(self, obj):
        if obj.min_order_amount:
            return f'≥ ৳{obj.min_order_amount}'
        return mark_safe('<span style="color:#999;">-</span>')
    min_order_display.short_description = 'Min. Order'

    def has_link(self, obj):
        if obj.link_url:
            return mark_safe(f'<a href="{obj.link_url}" target="_blank" style="color:#17a2b8;">{obj.link_text or "Link"}</a>')
        return mark_safe('<span style="color:#999;">-</span>')
    has_link.short_description = 'Link'

    def is_active_badge(self, obj):
        return mark_safe('<span style="color:#28a745;font-weight:600;">✓ Active</span>' if obj.is_active else '<span style="color:#dc3545;">✗ Inactive</span>')
    is_active_badge.short_description = 'Status'


# ─── Coupon ───

class CouponUsageInline(admin.TabularInline):
    model = CouponUsage
    extra = 0
    readonly_fields = ['order', 'user', 'discount_amount', 'created']
    can_delete = False

    def has_add_permission(self, request, obj=None):
        return False


@admin_register(Coupon, site=custom_admin_site)
class CouponAdmin(admin.ModelAdmin):
    list_display = ['code', 'discount_display', 'scope_badge', 'min_order_display', 'usage_display', 'valid_period', 'is_active_badge']
    list_filter = ['is_active', 'discount_type', 'scope', 'valid_from', 'valid_to']
    search_fields = ['code']
    filter_horizontal = ['products']
    inlines = [CouponUsageInline]
    actions = ['reset_usage']
    fieldsets = [
        (None, {'fields': ['code']}),
        ('Discount', {'fields': ['discount_type', 'discount_value', 'max_discount']}),
        ('Conditions', {'fields': ['min_order_amount', 'scope', 'products', 'max_uses']}),
        ('Validity', {'fields': ['valid_from', 'valid_to', 'is_active']}),
    ]

    def discount_display(self, obj):
        if obj.discount_type == 'percentage':
            txt = f'{obj.discount_value}%'
            if obj.max_discount:
                txt += f' (max ৳{obj.max_discount})'
            return txt
        return f'৳{obj.discount_value}'
    discount_display.short_description = 'Discount'

    def scope_badge(self, obj):
        if obj.scope == 'all':
            return mark_safe('<span style="background:#28a745;color:#fff;padding:2px 8px;border-radius:8px;font-size:10px;">All Products</span>')
        return mark_safe(f'<span style="background:#6f42c1;color:#fff;padding:2px 8px;border-radius:8px;font-size:10px;">{obj.products.count()} product(s)</span>')
    scope_badge.short_description = 'Scope'

    def min_order_display(self, obj):
        if obj.min_order_amount:
            return f'≥ ৳{obj.min_order_amount}'
        return mark_safe('<span style="color:#999;">-</span>')
    min_order_display.short_description = 'Min. Order'

    def usage_display(self, obj):
        if obj.max_uses > 0:
            return f'{obj.used_count} / {obj.max_uses}'
        return f'{obj.used_count} (unlimited)'
    usage_display.short_description = 'Used'

    def valid_period(self, obj):
        return f'{obj.valid_from.strftime("%d/%m/%y")} → {obj.valid_to.strftime("%d/%m/%y")}'
    valid_period.short_description = 'Valid Period'

    def is_active_badge(self, obj):
        from django.utils import timezone
        now = timezone.now()
        if not obj.is_active:
            return mark_safe('<span style="color:#dc3545;">Disabled</span>')
        if now < obj.valid_from:
            return mark_safe('<span style="color:#ffc107;color:#333;">Scheduled</span>')
        if now > obj.valid_to:
            return mark_safe('<span style="color:#dc3545;">Expired</span>')
        return mark_safe('<span style="color:#28a745;font-weight:600;">Active</span>')
    is_active_badge.short_description = 'Status'

    def reset_usage(self, request, queryset):
        updated = queryset.update(used_count=0)
        self.message_user(request, f'Usage count reset for {updated} coupon(s).')
    reset_usage.short_description = 'Reset usage count'


# ─── Custom Order ───

@admin_register(CustomOrder, site=custom_admin_site)
class CustomOrderAdmin(admin.ModelAdmin):
    list_display = ['id', 'product_title', 'partner_link', 'buyer_name', 'price_display', 'status_badge', 'created_date']
    list_filter = ['status', 'created']
    search_fields = ['product_title', 'buyer_name', 'buyer_phone', 'partner__name']
    date_hierarchy = 'created'
    actions = ['mark_approved', 'mark_pending', 'mark_rejected']
    readonly_fields = ['partner', 'seller_name', 'seller_phone', 'created', 'updated', 'created_order_link']

    fieldsets = [
        ('Product Info', {'fields': ['product_title', 'product_description', 'price', 'quantity']}),
        ('Buyer Info', {'fields': ['buyer_name', 'buyer_phone', 'buyer_address', 'buyer_email']}),
        ('Seller Info', {'fields': ['partner', 'seller_name', 'seller_phone']}),
        ('Status', {'fields': ['status', 'admin_notes', 'created_order_link', 'created', 'updated']}),
    ]

    def partner_link(self, obj):
        return format_html('<a href="{}">{}</a>', reverse('admin:store_partner_change', args=[obj.partner.id]), obj.partner.name)
    partner_link.short_description = 'Partner'

    def price_display(self, obj):
        return format_html('৳{}', obj.price)
    price_display.short_description = 'Price'

    def created_date(self, obj):
        return obj.created.strftime('%d %b %Y, %I:%M %p')
    created_date.short_description = 'Created'

    def created_order_link(self, obj):
        if obj.created_order:
            return format_html('<a href="{}">Order #{}</a>', reverse('admin:store_order_change', args=[obj.created_order.id]), obj.created_order.id)
        return mark_safe('<span style="color:#999;">Not yet approved</span>')
    created_order_link.short_description = 'Created Order'

    def status_badge(self, obj):
        colors = {'pending': ('#ffc107', '#333'), 'approved': ('#28a745', '#fff'), 'rejected': ('#dc3545', '#fff')}
        bg, fg = colors.get(obj.status, ('#6c757d', '#fff'))
        return format_html('<span style="background:{};color:{};padding:3px 12px;border-radius:12px;font-size:11px;font-weight:600;">{}</span>', bg, fg, obj.get_status_display())
    status_badge.short_description = 'Status'

    def mark_approved(self, request, queryset):
        for obj in queryset:
            if obj.status != 'approved':
                from django.utils import timezone
                order = _create_order_from_custom_order(obj)
                obj.status = 'approved'
                obj.created_order = order
                obj.save()
        self.message_user(request, f'{queryset.count()} custom order(s) approved and orders created.')
    mark_approved.short_description = 'Approve selected (creates Order)'

    def mark_pending(self, request, queryset):
        updated = queryset.update(status='pending')
        self.message_user(request, f'{updated} custom order(s) marked as pending.')
    mark_pending.short_description = 'Mark selected as Pending'

    def mark_rejected(self, request, queryset):
        updated = queryset.update(status='rejected')
        self.message_user(request, f'{updated} custom order(s) marked as rejected.')
    mark_rejected.short_description = 'Mark selected as Rejected'


def _create_order_from_custom_order(custom_order):
    from django.contrib.auth.models import User
    user = custom_order.partner.user
    if not user:
        username = f'custom_{custom_order.id}_{custom_order.buyer_phone}'
        user, _ = User.objects.get_or_create(username=username, defaults={'first_name': custom_order.buyer_name})
    order = Order.objects.create(
        user=user,
        name=custom_order.buyer_name,
        phone=custom_order.buyer_phone,
        address=custom_order.buyer_address or 'Custom Order',
        subtotal=custom_order.price * custom_order.quantity,
        total=custom_order.price * custom_order.quantity,
        payment_method='Custom Order',
        admin_notes=f'Auto-created from Custom Order #{custom_order.id}: {custom_order.product_title}',
    )
    OrderItem.objects.create(
        order=order,
        product=None,
        product_name=f'[Custom] {custom_order.product_title}',
        price=custom_order.price,
        quantity=custom_order.quantity,
    )
    return order


@admin_register(AdminCommission, site=custom_admin_site)
class AdminCommissionAdmin(admin.ModelAdmin):
    list_display = ['commission_percentage', 'commission_on', 'is_active', 'updated']
    list_editable = ['is_active']
    fieldsets = [
        (None, {
            'fields': ['commission_percentage', 'commission_on', 'is_active'],
            'description': 'Set the platform commission charged from partners. Only one record exists — changing it updates the existing value.',
        }),
    ]

    def has_add_permission(self, request):
        return not AdminCommission.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


def _credit_partner_wallets(order):
    for item in order.items.all():
        if not item.product or not item.product.partner:
            continue
        partner = item.product.partner
        wallet, _ = PartnerWallet.objects.get_or_create(partner=partner)
        gross_profit = (item.price - (item.cost_price or 0)) * item.quantity
        sale_total = item.price * item.quantity
        ac = AdminCommission.objects.first()
        commission = ac.calculate_commission(gross_profit, sale_total) if ac else 0
        net_credit = gross_profit - commission
        if net_credit <= 0:
            continue
        wallet.balance += net_credit
        wallet.total_earned += net_credit
        wallet.total_admin_commission += commission

        # Check if locked bonus should be released (cumulative commission >= 1000)
        prev_locked = wallet.locked_balance
        if prev_locked > 0 and wallet.total_admin_commission >= 1000:
            wallet.locked_balance = 0

        wallet.save()

        WalletTransaction.objects.create(
            wallet=wallet,
            amount=net_credit,
            balance_before=wallet.balance - net_credit,
            balance_after=wallet.balance,
            type='order_credit',
            description=f'Order #{order.id} - {item.product_name} x {item.quantity}',
            order=order,
        )

        if prev_locked > 0 and wallet.total_admin_commission >= 1000:
            WalletTransaction.objects.create(
                wallet=wallet,
                amount=prev_locked,
                balance_before=wallet.balance,
                balance_after=wallet.balance,
                type='bonus_release',
                description='Signup bonus unlocked (admin commission threshold reached)',
                order=order,
            )

    # Update platform balance with total commission from this order
    pb = PlatformBalance.objects.first()
    if pb:
        total_commission = Decimal(0)
        for item in order.items.all():
            if not item.product or not item.product.partner:
                continue
            gross_profit = (item.price - (item.cost_price or 0)) * item.quantity
            sale_total = item.price * item.quantity
            ac = AdminCommission.objects.first()
            commission = ac.calculate_commission(gross_profit, sale_total) if ac else 0
            total_commission += commission
        if total_commission > 0:
            pb.balance += total_commission
            pb.total_commission_collected += total_commission
            pb.save()


@admin_register(PartnerWallet, site=custom_admin_site)
class PartnerWalletAdmin(admin.ModelAdmin):
    list_display = ['partner', 'balance_display', 'available_display', 'locked_display', 'total_admin_commission_display', 'total_earned_display', 'total_withdrawn_display', 'is_frozen', 'credit_link', 'debit_link']
    list_filter = ['is_frozen']
    search_fields = ['partner__name', 'partner__phone']
    list_editable = ['is_frozen']
    readonly_fields = ['balance', 'locked_balance', 'total_admin_commission', 'total_earned', 'total_withdrawn', 'total_fees_paid', 'created', 'updated']

    def balance_display(self, obj):
        return f'৳{obj.balance}'
    balance_display.short_description = 'Total'

    def available_display(self, obj):
        return f'৳{obj.available_balance}'
    available_display.short_description = 'Available'

    def locked_display(self, obj):
        return f'৳{obj.locked_balance}'
    locked_display.short_description = 'Locked'

    def total_admin_commission_display(self, obj):
        return f'৳{obj.total_admin_commission}'
    total_admin_commission_display.short_description = 'Commission'

    def total_earned_display(self, obj):
        return f'৳{obj.total_earned}'
    total_earned_display.short_description = 'Earned'

    def total_withdrawn_display(self, obj):
        return f'৳{obj.total_withdrawn}'
    total_withdrawn_display.short_description = 'Withdrawn'

    def credit_link(self, obj):
        url = reverse('admin:credit-wallet', args=[obj.pk])
        return format_html('<a class="button" href="{}" style="background:#28a745;color:#fff;padding:3px 10px;border-radius:3px;text-decoration:none;font-size:11px;">+ Credit</a>', url)
    credit_link.short_description = ''

    def debit_link(self, obj):
        url = reverse('admin:debit-wallet', args=[obj.pk])
        return format_html('<a class="button" href="{}" style="background:#dc3545;color:#fff;padding:3px 10px;border-radius:3px;text-decoration:none;font-size:11px;">- Debit</a>', url)
    debit_link.short_description = ''

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('credit/<int:wallet_id>/', self.admin_site.admin_view(self.credit_wallet_view), name='credit-wallet'),
            path('debit/<int:wallet_id>/', self.admin_site.admin_view(self.debit_wallet_view), name='debit-wallet'),
        ]
        return custom_urls + urls

    def credit_wallet_view(self, request, wallet_id):
        wallet = get_object_or_404(PartnerWallet, id=wallet_id)
        if request.method == 'POST':
            amount = request.POST.get('amount', '').strip()
            reason = request.POST.get('reason', '').strip()
            if not amount or not amount.replace('.', '', 1).isdigit() or Decimal(amount) <= 0:
                self.message_user(request, 'Invalid amount.', level='ERROR')
            elif not reason:
                self.message_user(request, 'Reason is required.', level='ERROR')
            else:
                amount = Decimal(amount)
                pb = PlatformBalance.objects.first()
                if pb:
                    pb.balance -= amount
                    pb.total_manual_credits_given += amount
                    pb.save()
                wallet.balance += amount
                wallet.total_earned += amount
                wallet.save()
                WalletTransaction.objects.create(
                    wallet=wallet,
                    amount=amount,
                    balance_before=wallet.balance - amount,
                    balance_after=wallet.balance,
                    type='manual_credit',
                    description=f'Admin credit: {reason}',
                )
                self.message_user(request, f'Credited ৳{amount} to {wallet.partner.name}.')
                return redirect('admin:store_partnerwallet_changelist')
        return render(request, 'admin/wallet_credit.html', {
            'wallet': wallet,
            'action': 'Credit',
            'action_color': '#28a745',
        })

    def debit_wallet_view(self, request, wallet_id):
        wallet = get_object_or_404(PartnerWallet, id=wallet_id)
        if request.method == 'POST':
            amount = request.POST.get('amount', '').strip()
            reason = request.POST.get('reason', '').strip()
            if not amount or not amount.replace('.', '', 1).isdigit() or Decimal(amount) <= 0:
                self.message_user(request, 'Invalid amount.', level='ERROR')
            elif not reason:
                self.message_user(request, 'Reason is required.', level='ERROR')
            elif Decimal(amount) > wallet.balance:
                self.message_user(request, 'Amount exceeds wallet balance.', level='ERROR')
            else:
                amount = Decimal(amount)
                wallet.balance -= amount
                wallet.save()
                WalletTransaction.objects.create(
                    wallet=wallet,
                    amount=-amount,
                    balance_before=wallet.balance + amount,
                    balance_after=wallet.balance,
                    type='manual_debit',
                    description=f'Admin debit: {reason}',
                )
                self.message_user(request, f'Debited ৳{amount} from {wallet.partner.name}.')
                return redirect('admin:store_partnerwallet_changelist')
        return render(request, 'admin/wallet_credit.html', {
            'wallet': wallet,
            'action': 'Debit',
            'action_color': '#dc3545',
        })


@admin_register(WalletTransaction, site=custom_admin_site)
class WalletTransactionAdmin(admin.ModelAdmin):
    list_display = ['reference_id', 'wallet', 'amount_display', 'type', 'description_short', 'order_link', 'created']
    list_filter = ['type', 'created']
    search_fields = ['reference_id', 'wallet__partner__name', 'description']
    readonly_fields = [f.name for f in WalletTransaction._meta.fields]
    date_hierarchy = 'created'

    def amount_display(self, obj):
        color = '#28a745' if obj.amount >= 0 else '#dc3545'
        return format_html('<span style="color:{};font-weight:600;">{}</span>', color, f'৳{obj.amount}')
    amount_display.short_description = 'Amount'

    def description_short(self, obj):
        return obj.description[:60] + '...' if len(obj.description) > 60 else obj.description
    description_short.short_description = 'Description'

    def order_link(self, obj):
        if obj.order_id:
            url = reverse('admin:store_order_change', args=[obj.order_id])
            return format_html('<a href="{}">Order #{}</a>', url, obj.order_id)
        return '—'
    order_link.short_description = 'Order'

    def has_add_permission(self, request): return False
    def has_change_permission(self, request, obj=None): return False
    def has_delete_permission(self, request, obj=None):
        return False


# ─── Feature 1: Wishlist Admin ───

@admin_register(Wishlist, site=custom_admin_site)
class WishlistAdmin(admin.ModelAdmin):
    list_display = ['user', 'session_id', 'total_items', 'created']
    search_fields = ['user__username', 'session_id']


@admin_register(WishlistItem, site=custom_admin_site)
class WishlistItemAdmin(admin.ModelAdmin):
    list_display = ['wishlist', 'product', 'added']
    search_fields = ['product__name', 'wishlist__user__username']


# ─── Feature 4: Notification Admin ───

@admin_register(Notification, site=custom_admin_site)
class NotificationAdmin(admin.ModelAdmin):
    list_display = ['title', 'type', 'user', 'partner', 'is_read', 'created']
    list_filter = ['type', 'is_read', 'created']
    search_fields = ['title', 'user__username', 'partner__name']
    readonly_fields = [f.name for f in Notification._meta.fields]
    date_hierarchy = 'created'

    def has_add_permission(self, request): return False
    def has_change_permission(self, request, obj=None): return False


# ─── Feature 5: Shipping Rule Admin ───

@admin_register(ShippingRule, site=custom_admin_site)
class ShippingRuleAdmin(admin.ModelAdmin):
    list_display = ['name', 'courier', 'based_on', 'cost', 'free_above', 'is_active']
    list_filter = ['based_on', 'courier', 'is_active']
    search_fields = ['name']


# ─── Feature 6: Refund Request Admin ───

@admin_register(RefundRequest, site=custom_admin_site)
class RefundRequestAdmin(admin.ModelAdmin):
    list_display = ['id', 'order', 'user', 'partner', 'amount_display', 'status_badge', 'created']
    list_filter = ['status', 'created']
    search_fields = ['order__id', 'user__username', 'partner__name']
    readonly_fields = ['order', 'user', 'partner', 'amount', 'reason', 'created']
    actions = ['approve_refund', 'reject_refund']

    def amount_display(self, obj):
        return f'৳{obj.amount}'
    amount_display.short_description = 'Amount'

    def status_badge(self, obj):
        colors = {'pending': ('#ffc107', '#333'), 'approved': ('#28a745', '#fff'), 'rejected': ('#dc3545', '#fff')}
        bg, fg = colors.get(obj.status, ('#6c757d', '#fff'))
        return format_html('<span style="background:{};color:{};padding:2px 10px;border-radius:10px;font-size:11px;font-weight:600;">{}</span>', bg, fg, obj.get_status_display())
    status_badge.short_description = 'Status'

    def approve_refund(self, request, queryset):
        for r in queryset.filter(status='pending'):
            wallet = PartnerWallet.objects.filter(partner=r.partner).first()
            if wallet and r.amount <= wallet.balance:
                wallet.balance -= r.amount
                wallet.save()
                WalletTransaction.objects.create(
                    wallet=wallet,
                    amount=-r.amount,
                    balance_before=wallet.balance + r.amount,
                    balance_after=wallet.balance,
                    type='manual_debit',
                    description=f'Refund for Order #{r.order.id} - {r.reason[:50]}',
                    order=r.order,
                )
            r.status = 'approved'
            r.processed_by = request.user
            r.processed_at = timezone.now()
            r.save()
        self.message_user(request, f'{queryset.count()} refund(s) approved.')
    approve_refund.short_description = 'Approve selected refunds'

    def reject_refund(self, request, queryset):
        updated = queryset.filter(status='pending').update(status='rejected', processed_by=request.user, processed_at=timezone.now())
        self.message_user(request, f'{updated} refund(s) rejected.')
    reject_refund.short_description = 'Reject selected refunds'


@admin_register(WithdrawalRequest, site=custom_admin_site)
class WithdrawalRequestAdmin(admin.ModelAdmin):
    list_display = ['id', 'partner', 'amount_display', 'method_display', 'status_badge', 'requested_at', 'processed_at']
    list_filter = ['status', 'method', 'requested_at']
    search_fields = ['partner__name', 'partner__phone', 'account_number']
    readonly_fields = ['amount', 'fee', 'net_amount', 'method', 'account_number', 'account_holder', 'bank_name', 'branch', 'routing_number', 'note', 'partner', 'requested_at']
    actions = ['approve_withdrawal', 'reject_withdrawal', 'mark_completed']

    def amount_display(self, obj):
        return f'৳{obj.amount} (net: ৳{obj.net_amount})'
    amount_display.short_description = 'Amount'

    def method_display(self, obj):
        return obj.get_method_display()
    method_display.short_description = 'Method'

    def status_badge(self, obj):
        colors = {'pending': '#ffc107', 'approved': '#17a2b8', 'rejected': '#dc3545', 'completed': '#28a745'}
        return format_html('<span style="background:{};color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;">{}</span>', colors[obj.status], obj.get_status_display())
    status_badge.short_description = 'Status'

    def approve_withdrawal(self, request, queryset):
        for w in queryset.filter(status='pending'):
            wallet = PartnerWallet.objects.filter(partner=w.partner).first()
            if not wallet or wallet.balance < w.amount:
                self.message_user(request, f'Withdrawal #{w.id}: insufficient balance.', level='ERROR')
                continue
            if wallet.is_frozen:
                self.message_user(request, f'Withdrawal #{w.id}: wallet is frozen.', level='ERROR')
                continue
            wallet.balance = wallet.balance - w.net_amount - w.fee
            wallet.total_withdrawn += w.net_amount
            wallet.total_fees_paid += w.fee
            wallet.save()
            WalletTransaction.objects.create(
                wallet=wallet,
                amount=-w.net_amount,
                balance_before=wallet.balance + w.net_amount + w.fee,
                balance_after=wallet.balance,
                type='withdrawal',
                description=f'Withdrawal #{w.id} - {w.get_method_display()} {w.account_number}',
                withdrawal=w,
            )
            if w.fee > 0:
                WalletTransaction.objects.create(
                    wallet=wallet,
                    amount=-w.fee,
                    balance_before=wallet.balance + w.fee,
                    balance_after=wallet.balance,
                    type='withdrawal_fee',
                    description=f'Withdrawal fee for #{w.id}',
                    withdrawal=w,
                )
            w.status = 'approved'
            w.processed_by = request.user
            w.processed_at = timezone.now()
            w.save()
        self.message_user(request, f'{queryset.count()} withdrawal(s) approved.')
    approve_withdrawal.short_description = 'Approve selected withdrawals'

    def reject_withdrawal(self, request, queryset):
        updated = queryset.filter(status='pending').update(status='rejected', processed_by=request.user, processed_at=timezone.now())
        self.message_user(request, f'{updated} withdrawal(s) rejected.')
    reject_withdrawal.short_description = 'Reject selected withdrawals'

    def mark_completed(self, request, queryset):
        updated = queryset.filter(status='approved').update(status='completed')
        self.message_user(request, f'{updated} withdrawal(s) marked as completed.')
    mark_completed.short_description = 'Mark selected as Completed'


@admin_register(PayoutMethod, site=custom_admin_site)
class PayoutMethodAdmin(admin.ModelAdmin):
    list_display = ['partner', 'method_type', 'account_number', 'is_default', 'is_active']
    list_filter = ['method_type', 'is_default', 'is_active']
    search_fields = ['partner__name', 'account_number']


@admin_register(WalletSettings, site=custom_admin_site)
class WalletSettingsAdmin(admin.ModelAdmin):
    list_display = ['min_withdrawal', 'max_withdrawal', 'withdrawal_fee_value', 'withdrawal_fee_type', 'auto_approve_withdrawals', 'max_pending_withdrawals']
    fieldsets = [
        (None, {
            'fields': ['min_withdrawal', 'max_withdrawal', 'max_pending_withdrawals'],
            'description': 'Control withdrawal limits per partner.',
        }),
        ('Withdrawal Fee', {
            'fields': ['withdrawal_fee_type', 'withdrawal_fee_value'],
            'description': 'Fee charged to the partner on each withdrawal. Fixed amount (e.g. 10 = ৳10) or percentage (e.g. 1 = 1%).',
        }),
        ('Auto-Approval', {
            'fields': ['auto_approve_withdrawals'],
            'description': 'When enabled, withdrawals under the minimum threshold are auto-approved.',
        }),
    ]

    def has_add_permission(self, request):
        return not WalletSettings.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin_register(PlatformBalance, site=custom_admin_site)
class PlatformBalanceAdmin(admin.ModelAdmin):
    list_display = ['balance_display', 'total_commission_display', 'total_credits_display', 'updated']
    readonly_fields = ['balance', 'total_commission_collected', 'total_manual_credits_given', 'updated']

    def balance_display(self, obj):
        return f'৳{obj.balance}'
    balance_display.short_description = 'Platform Balance'

    def total_commission_display(self, obj):
        return f'৳{obj.total_commission_collected}'
    total_commission_display.short_description = 'Total Commission Collected'

    def total_credits_display(self, obj):
        return f'৳{obj.total_manual_credits_given}'
    total_credits_display.short_description = 'Total Manual Credits Given'

    def has_add_permission(self, request):
        return not PlatformBalance.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


@admin_register(ManualPaymentMethod, site=custom_admin_site)
class ManualPaymentMethodAdmin(admin.ModelAdmin):
    list_display = ['name', 'account_number', 'account_type', 'is_active', 'order']
    list_editable = ['is_active', 'order']
    list_filter = ['is_active', 'name']
    search_fields = ['name', 'account_number']
    fieldsets = [
        (None, {'fields': ['name', 'account_number', 'account_type']}),
        ('Display', {'fields': ['is_active', 'order']}),
        ('Instructions', {'fields': ['instructions']}),
    ]


# ─── POS Orders (default admin, read-only) ───

@admin.register(PosOrder)
class PosOrderAdmin(admin.ModelAdmin):
    list_display = ['invoice_number', 'partner', 'customer_name', 'total', 'payment_method', 'status', 'created']
    list_filter = ['status', 'payment_method', 'created']
    search_fields = ['invoice_number', 'customer_name', 'customer_phone', 'partner__name']
    readonly_fields = ['invoice_number', 'partner', 'customer_name', 'customer_phone', 'subtotal', 'discount', 'tax', 'total', 'payment_method', 'status', 'notes', 'created', 'updated']
    date_hierarchy = 'created'

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(PosOrderItem)
class PosOrderItemAdmin(admin.ModelAdmin):
    list_display = ['pos_order', 'product_name', 'price', 'quantity', 'total_price']
    search_fields = ['product_name', 'pos_order__invoice_number']
    readonly_fields = ['pos_order', 'product', 'product_name', 'price', 'quantity', 'cost_price', 'discount']

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


# ─── Messaging System Admin ───

@admin_register(Conversation, site=custom_admin_site)
class ConversationAdmin(admin.ModelAdmin):
    list_display = ['id', 'subject', 'partner_link', 'product_link', 'is_broadcast', 'participant_count', 'message_count', 'created', 'updated']
    list_filter = ['is_broadcast', 'created']
    search_fields = ['subject', 'partner__name']
    readonly_fields = ['created', 'updated']
    filter_horizontal = ['participants']

    def partner_link(self, obj):
        if obj.partner:
            return obj.partner.name
        return '—'
    partner_link.short_description = 'Partner'

    def product_link(self, obj):
        if obj.product:
            return obj.product.name[:40]
        return '—'
    product_link.short_description = 'Product'

    def participant_count(self, obj):
        return obj.participants.count()
    participant_count.short_description = 'Participants'

    def message_count(self, obj):
        return obj.messages.count()
    message_count.short_description = 'Messages'


@admin_register(Message, site=custom_admin_site)
class MessageAdmin(admin.ModelAdmin):
    list_display = ['id', 'conversation_id', 'sender_name', 'preview', 'created']
    list_filter = ['created']
    search_fields = ['content', 'sender__username', 'conversation__subject']
    readonly_fields = ['created']

    def sender_name(self, obj):
        return obj.sender.username
    sender_name.short_description = 'Sender'

    def preview(self, obj):
        return obj.content[:80] + '...' if len(obj.content) > 80 else obj.content
    preview.short_description = 'Message'


@admin_register(ConversationReadStatus, site=custom_admin_site)
class ConversationReadStatusAdmin(admin.ModelAdmin):
    list_display = ['conversation_id', 'user', 'last_read']
    search_fields = ['user__username']


@admin_register(ProductQA, site=custom_admin_site)
class ProductQAAdmin(admin.ModelAdmin):
    list_display = ['product_name', 'question_preview', 'user', 'has_answer', 'created']
    list_filter = ['product', 'created']
    search_fields = ['question', 'product__name', 'user__username']

    def product_name(self, obj):
        return obj.product.name[:40]
    product_name.short_description = 'Product'

    def question_preview(self, obj):
        return obj.question[:60] + '...' if len(obj.question) > 60 else obj.question
    question_preview.short_description = 'Question'

    def has_answer(self, obj):
        return bool(obj.answer)
    has_answer.boolean = True


# ─── Support Ticket System Admin ───

@admin_register(SupportTicket, site=custom_admin_site)
class SupportTicketAdmin(admin.ModelAdmin):
    list_display = ['id', 'partner_link', 'subject_short', 'priority_badge', 'status', 'status_badge', 'replies_count', 'created']
    list_filter = ['status', 'priority', 'created']
    list_editable = ['status']
    search_fields = ['subject', 'description', 'partner__name', 'partner__phone']
    date_hierarchy = 'created'
    readonly_fields = ['partner', 'subject', 'description', 'image', 'priority', 'created', 'updated']
    actions = ['mark_in_progress', 'mark_resolved', 'mark_closed']
    fieldsets = [
        (None, {'fields': ['partner', 'subject', 'description', 'image']}),
        ('Status', {'fields': ['status', 'priority']}),
        ('Timestamps', {'fields': ['created', 'updated']}),
    ]

    def partner_link(self, obj):
        url = reverse('admin:store_partner_change', args=[obj.partner.id])
        return format_html('<a href="{}">{}</a>', url, obj.partner.name)
    partner_link.short_description = 'Partner'

    def subject_short(self, obj):
        return obj.subject[:50] + '...' if len(obj.subject) > 50 else obj.subject
    subject_short.short_description = 'Subject'

    def replies_count(self, obj):
        return obj.replies.count()
    replies_count.short_description = 'Replies'

    def priority_badge(self, obj):
        colors = {'low': '#6c757d', 'medium': '#17a2b8', 'high': '#ffc107', 'urgent': '#dc3545'}
        c = colors.get(obj.priority, '#6c757d')
        return format_html('<span style="background:{};color:#fff;padding:2px 8px;border-radius:8px;font-size:10px;">{}</span>', c, obj.get_priority_display())
    priority_badge.short_description = 'Priority'

    def status_badge(self, obj):
        colors = {'open': '#dc3545', 'in_progress': '#17a2b8', 'resolved': '#28a745', 'closed': '#6c757d'}
        c = colors.get(obj.status, '#6c757d')
        return format_html('<span style="background:{};color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;font-weight:600;">{}</span>', c, obj.get_status_display())
    status_badge.short_description = 'Status'

    def mark_in_progress(self, request, queryset):
        updated = queryset.filter(status='open').update(status='in_progress')
        self.message_user(request, f'{updated} ticket(s) marked as in progress.')
    mark_in_progress.short_description = 'Mark selected as In Progress'

    def mark_resolved(self, request, queryset):
        updated = queryset.exclude(status='closed').update(status='resolved')
        self.message_user(request, f'{updated} ticket(s) marked as resolved.')
    mark_resolved.short_description = 'Mark selected as Resolved'

    def mark_closed(self, request, queryset):
        updated = queryset.exclude(status='closed').update(status='closed')
        self.message_user(request, f'{updated} ticket(s) marked as closed.')
    mark_closed.short_description = 'Mark selected as Closed'

    def has_add_permission(self, request):
        return False


@admin_register(TicketReply, site=custom_admin_site)
class TicketReplyAdmin(admin.ModelAdmin):
    list_display = ['ticket', 'user', 'message_short', 'has_attachment', 'created']
    list_filter = ['created']
    search_fields = ['message', 'user__username', 'ticket__subject']
    readonly_fields = ['ticket', 'user', 'message', 'attachment', 'created']

    def message_short(self, obj):
        return obj.message[:60] + '...' if len(obj.message) > 60 else obj.message
    message_short.short_description = 'Message'

    def has_attachment(self, obj):
        return bool(obj.attachment)
    has_attachment.boolean = True

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False


# ─── Medicine System Admin ───

@admin_register(MedicineProduct, site=custom_admin_site)
class MedicineProductAdmin(admin.ModelAdmin):
    list_display = ['sku', 'brand_name', 'generic_name_short', 'strength', 'dosage_form', 'price_display', 'stock_badge', 'is_approved', 'is_approved_badge', 'requested_by_info', 'created']
    list_filter = ['is_approved', 'dosage_form', 'created']
    list_editable = ['is_approved']
    search_fields = ['brand_name', 'generic_name', 'sku', 'strength', 'dosage_form', 'requested_by__name']
    date_hierarchy = 'created'
    fieldsets = [
        (None, {'fields': ['brand_name', 'generic_name', 'strength', 'dosage_form', 'sku']}),
        ('Pricing & Stock', {'fields': ['price', 'stock']}),
        ('Media', {'fields': ['image', 'description']}),
        ('Approval', {'fields': ['is_approved', 'requested_by']}),
    ]
    readonly_fields = ['sku', 'requested_by']
    actions = ['approve_selected', 'reject_selected']

    def generic_name_short(self, obj):
        return obj.generic_name[:50] + '...' if len(obj.generic_name) > 50 else obj.generic_name
    generic_name_short.short_description = 'Generic Name'

    def price_display(self, obj):
        return format_html('<strong>৳{}</strong>', obj.price)
    price_display.short_description = 'Price'

    def stock_badge(self, obj):
        if obj.stock <= 0:
            return mark_safe('<span style="background:#dc3545;color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;">Out</span>')
        elif obj.stock <= 5:
            return format_html('<span style="background:#ffc107;color:#333;padding:2px 10px;border-radius:10px;font-size:11px;">{}</span>', obj.stock)
        return format_html('<span style="color:#28a745;font-weight:600;">{}</span>', obj.stock)
    stock_badge.short_description = 'Stock'

    def is_approved_badge(self, obj):
        if obj.is_approved:
            return mark_safe('<span style="background:#28a745;color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;font-weight:600;">✓ Approved</span>')
        return mark_safe('<span style="background:#ffc107;color:#333;padding:2px 10px;border-radius:10px;font-size:11px;">⏳ Pending</span>')
    is_approved_badge.short_description = 'Status'

    def requested_by_info(self, obj):
        if obj.requested_by:
            return format_html('<a href="{}">{}</a>', reverse('custom_admin:store_partner_change', args=[obj.requested_by.id]), obj.requested_by.name)
        return mark_safe('<span style="color:#999;">Admin</span>')
    requested_by_info.short_description = 'Requested By'

    def approve_selected(self, request, queryset):
        updated = queryset.update(is_approved=True)
        self.message_user(request, f'{updated} medicine product(s) approved.')
    approve_selected.short_description = 'Approve selected'

    def reject_selected(self, request, queryset):
        updated = queryset.filter(is_approved=False).delete()[0]
        self.message_user(request, f'{updated} pending medicine product(s) deleted.')
    reject_selected.short_description = 'Delete selected pending (reject)'


@admin_register(MedicinePosOrder, site=custom_admin_site)
class MedicinePosOrderAdmin(admin.ModelAdmin):
    list_display = ['invoice_number', 'partner', 'customer_name', 'total_display', 'payment_method', 'status_badge', 'created']
    list_filter = ['status', 'payment_method', 'created']
    search_fields = ['invoice_number', 'customer_name', 'customer_phone', 'partner__name']
    readonly_fields = ['invoice_number', 'partner', 'customer_name', 'customer_phone', 'subtotal', 'discount', 'total', 'payment_method', 'status', 'notes', 'created', 'updated']
    date_hierarchy = 'created'

    def total_display(self, obj):
        return f'৳{obj.total}'
    total_display.short_description = 'Total'

    def status_badge(self, obj):
        colors = {'completed': '#28a745', 'refunded': '#ffc107', 'cancelled': '#dc3545'}
        c = colors.get(obj.status, '#6c757d')
        return format_html('<span style="background:{};color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;font-weight:600;">{}</span>', c, obj.get_status_display())
    status_badge.short_description = 'Status'

    def has_add_permission(self, request, obj=None): return False
    def has_delete_permission(self, request, obj=None):
        if obj is None:
            return False
        return not obj.partner.blocked


@admin_register(MedicinePosOrderItem, site=custom_admin_site)
class MedicinePosOrderItemAdmin(admin.ModelAdmin):
    list_display = ['pos_order', 'product_name', 'price', 'quantity', 'total_price']
    search_fields = ['product_name', 'pos_order__invoice_number']
    readonly_fields = ['pos_order', 'product', 'product_name', 'price', 'quantity', 'discount']

    def has_add_permission(self, request, obj=None): return False
    def has_delete_permission(self, request, obj=None):
        if obj is None:
            return False
        return not obj.pos_order.partner.blocked


@admin_register(MedicineSubscription, site=custom_admin_site)
class MedicineSubscriptionAdmin(admin.ModelAdmin):
    list_display = ['partner_link', 'status_badge', 'trial_period', 'current_period', 'last_payment_at', 'next_payment_due', 'wallet_link', 'activation_buttons', 'updated']
    list_filter = ['status']
    search_fields = ['partner__name']
    readonly_fields = ['partner', 'status', 'trial_started_at', 'trial_ends_at', 'current_period_start', 'current_period_end', 'last_payment_at', 'next_payment_due', 'created', 'updated']
    change_form_template = 'admin/store/medicinesubscription/change_form.html'

    def partner_link(self, obj):
        return format_html(
            '<a href="{}" style="font-weight:600;">{}</a>'
            '<br><small><a href="{}" style="color:#6f42c1;">Wallet</a> &middot; '
            '<a href="{}" style="color:#17a2b8;">Products</a> &middot; '
            '<a href="{}" style="color:#28a745;">Orders</a></small>',
            reverse('admin:store_partner_change', args=[obj.partner.id]),
            obj.partner.name,
            reverse('admin:store_partnerwallet_changelist') + f'?partner__id__exact={obj.partner.id}',
            reverse('admin:store_medicineproduct_changelist') + f'?requested_by__id__exact={obj.partner.id}',
            reverse('admin:store_medicineposorder_changelist') + f'?partner__id__exact={obj.partner.id}',
        )
    partner_link.short_description = 'Partner'
    partner_link.admin_order_field = 'partner__name'

    def wallet_link(self, obj):
        wallet = PartnerWallet.objects.filter(partner=obj.partner).first()
        if wallet:
            return format_html(
                '<a href="{}" style="color:#6f42c1;font-weight:600;">৳{}</a>',
                reverse('admin:store_partnerwallet_change', args=[wallet.id]),
                wallet.balance,
            )
        return '—'
    wallet_link.short_description = 'Wallet'

    def status_badge(self, obj):
        colors = {'inactive': '#6c757d', 'trial': '#17a2b8', 'active': '#28a745', 'expired': '#dc3545', 'cancelled': '#ffc107'}
        c = colors.get(obj.status, '#6c757d')
        return format_html('<span style="background:{};color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;font-weight:600;">{}</span>', c, obj.get_status_display())
    status_badge.short_description = 'Status'

    def trial_period(self, obj):
        if obj.trial_started_at and obj.trial_ends_at:
            return f'{obj.trial_started_at.strftime("%d %b")} → {obj.trial_ends_at.strftime("%d %b")}'
        return '—'
    trial_period.short_description = 'Trial Period'

    def current_period(self, obj):
        if obj.current_period_start and obj.current_period_end:
            return f'{obj.current_period_start.strftime("%d %b")} → {obj.current_period_end.strftime("%d %b")}'
        return '—'
    current_period.short_description = 'Current Period'

    def activation_buttons(self, obj):
        if obj.status in ('expired', 'inactive', 'cancelled', 'trial'):
            return format_html(
                '<a class="button" style="background:#28a745;color:#fff;padding:3px 10px;border-radius:4px;text-decoration:none;font-size:11px;margin-right:4px;" href="{}?days=30">+1 Month</a>'
                '<a class="button" style="background:#17a2b8;color:#fff;padding:3px 10px;border-radius:4px;text-decoration:none;font-size:11px;margin-right:4px;" href="{}?days=60">+2 Months</a>'
                '<a class="button" style="background:#6f42c1;color:#fff;padding:3px 10px;border-radius:4px;text-decoration:none;font-size:11px;" href="{}">Custom</a>',
                reverse('admin:activate-subscription', args=[obj.id]),
                reverse('admin:activate-subscription', args=[obj.id]),
                reverse('admin:store_medicinesubscription_change', args=[obj.id]),
            )
        return format_html(
            '<span style="color:#28a745;font-weight:600;">Active</span>'
            ' &nbsp; <a class="button" style="background:#dc3545;color:#fff;padding:3px 10px;border-radius:4px;text-decoration:none;font-size:11px;" href="{}">Deactivate</a>',
            reverse('admin:deactivate-subscription', args=[obj.id]),
        )
    activation_buttons.short_description = 'Activate / Deactivate'

    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('<int:sub_id>/activate/',
                 self.admin_site.admin_view(self.activate_subscription_view),
                 name='activate-subscription'),
            path('<int:sub_id>/deactivate/',
                 self.admin_site.admin_view(self.deactivate_subscription_view),
                 name='deactivate-subscription'),
        ]
        return custom_urls + urls

    def activate_subscription_view(self, request, sub_id):
        from django.shortcuts import get_object_or_404, redirect
        from django.contrib import messages
        sub = get_object_or_404(MedicineSubscription, id=sub_id)
        days = request.GET.get('days', '30')
        try:
            days = int(days)
            if days < 1:
                days = 30
        except (ValueError, TypeError):
            days = 30

        now = timezone.now()
        sub.status = 'active'
        sub.current_period_start = now
        sub.current_period_end = now + timedelta(days=days)
        sub.last_payment_at = now
        sub.next_payment_due = sub.current_period_end
        if not sub.trial_started_at:
            sub.trial_started_at = now
            sub.trial_ends_at = now
        sub.save()
        messages.success(request, f'Subscription activated for {sub.partner.name} — {days} days.')
        return redirect('admin:store_medicinesubscription_changelist')

    def deactivate_subscription_view(self, request, sub_id):
        from django.shortcuts import get_object_or_404, redirect
        from django.contrib import messages
        sub = get_object_or_404(MedicineSubscription, id=sub_id)
        sub.status = 'inactive'
        sub.save()
        messages.success(request, f'Subscription deactivated for {sub.partner.name}.')
        return redirect('admin:store_medicinesubscription_changelist')

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        sub = get_object_or_404(MedicineSubscription, id=object_id)
        extra_context['wallet'] = PartnerWallet.objects.filter(partner=sub.partner).first()
        extra_context['medicine_products'] = MedicineProduct.objects.filter(requested_by=sub.partner).count()
        extra_context['medicine_orders'] = MedicinePosOrder.objects.filter(partner=sub.partner).count()
        extra_context['partner_edit_url'] = reverse('admin:store_partner_change', args=[sub.partner.id])
        extra_context['wallet_url'] = reverse('admin:store_partnerwallet_changelist') + f'?partner__id__exact={sub.partner.id}'
        extra_context['products_url'] = reverse('admin:store_medicineproduct_changelist') + f'?requested_by__id__exact={sub.partner.id}'
        extra_context['orders_url'] = reverse('admin:store_medicineposorder_changelist') + f'?partner__id__exact={sub.partner.id}'
        extra_context['medicine_access_url'] = reverse('custom_admin:medicine-partner-access')
        extra_context['pending_requests_url'] = reverse('custom_admin:medicine-pending-requests')
        extra_context['recharge_instructions_url'] = reverse('admin:store_walletrechargeinstruction_changelist')
        extra_context['subscription_list_url'] = reverse('admin:store_medicinesubscription_changelist')
        return super().change_view(request, object_id, form_url, extra_context)

    def has_add_permission(self, request): return False
    def has_delete_permission(self, request, obj=None):
        if obj is None:
            return False
        return not obj.partner.blocked


@admin_register(SubscriptionPackage, site=custom_admin_site)
class SubscriptionPackageAdmin(admin.ModelAdmin):
    list_display = ['name', 'price_display', 'duration_days', 'sort_order', 'is_active', 'created']
    list_editable = ['sort_order', 'is_active']
    list_filter = ['is_active']
    search_fields = ['name', 'description']
    fieldsets = [
        (None, {'fields': ['name', 'price', 'duration_days', 'description']}),
        ('Display', {'fields': ['is_active', 'sort_order']}),
    ]

    def price_display(self, obj):
        return f'৳{int(obj.price)}'
    price_display.short_description = 'Price'
    price_display.admin_order_field = 'price'


@admin_register(WalletRechargeInstruction, site=custom_admin_site)
class WalletRechargeInstructionAdmin(admin.ModelAdmin):
    list_display = ['title', 'is_active', 'order', 'created']
    list_editable = ['is_active', 'order']
    list_filter = ['is_active']
    search_fields = ['title', 'instruction_text']
    fieldsets = [
        (None, {'fields': ['title', 'instruction_text']}),
        ('Display', {'fields': ['is_active', 'order']}),
    ]
