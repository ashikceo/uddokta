import csv
import io
from django.contrib.admin import AdminSite
from django.contrib import messages
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncDate
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import path
from django.utils import timezone
from django.utils.text import slugify
from datetime import timedelta
from . import bd_data
from .models import Product, Order, OrderItem, Contact, ProductReview, Partner, RefundRequest, WithdrawalRequest, CustomOrder, SiteLogo, Conversation, SupportTicket, MedicineProduct, MedicineSubscription


MODEL_GROUPS = {
    'Shop': {
        'models': ['Product', 'Category', 'Order', 'OrderItem', 'Coupon', 'CouponUsage', 'CustomOrder'],
        'icon': 'shopping-bag',
    },
    'Partners': {
        'models': ['Partner', 'PartnerWallet', 'WalletTransaction', 'WithdrawalRequest', 'PayoutMethod'],
        'icon': 'users',
    },
    'Content': {
        'models': ['BlogPost', 'Page', 'LandingPage', 'Slider', 'HomeBanner', 'SideBanner', 'NavMenu', 'PartnerNavMenu'],
        'icon': 'file-text',
    },
    'Finance': {
        'models': ['AdminCommission', 'PlatformBalance', 'ShippingRule', 'RefundRequest', 'ServerFee', 'WalletSettings', 'ManualPaymentMethod'],
        'icon': 'money',
    },
    'System': {
        'models': ['User', 'Group', 'Contact', 'SiteLogo', 'SocialMediaLink'],
        'icon': 'cog',
    },
    'Features': {
        'models': ['Wishlist', 'WishlistItem', 'Notification', 'ProductReview'],
        'icon': 'star',
    },
    'Support': {
        'models': ['SupportTicket', 'TicketReply'],
        'icon': 'life-ring',
    },
    'Medicine': {
        'models': ['MedicineProduct', 'MedicinePosOrder', 'MedicinePosOrderItem', 'MedicineSubscription', 'SubscriptionPackage', 'WalletRechargeInstruction', 'Partner'],
        'icon': 'medkit',
    },
}

class CustomAdminSite(AdminSite):
    site_header = 'Uddoktar Dokan Administration'
    site_title = 'Uddoktar Dokan Admin'
    index_title = 'Dashboard'

    def each_context(self, request):
        ctx = super().each_context(request)
        logo = SiteLogo.objects.first()
        ctx['site_logo'] = logo
        ctx['site_name'] = logo.site_name if logo and logo.site_name else 'Uddoktar Dokan'
        ctx['site_tagline'] = logo.site_tagline if logo else ''
        from django.conf import settings
        ctx['admin_banner_message'] = getattr(settings, 'ADMIN_BANNER_MESSAGE', '')
        return ctx

    def get_app_list(self, request, app_label=None):
        app_list = super().get_app_list(request)
        if not app_list:
            return app_list

        store_app = None
        for app in app_list:
            if app.get('app_label') == 'store':
                store_app = app
                break

        if not store_app:
            return app_list

        models = store_app.get('models', [])

        model_map = {}
        for m in models:
            name = m['object_name']
            model_map[name] = m

        grouped = []
        for group_name, group_info in MODEL_GROUPS.items():
            group_models = []
            for model_name in group_info['models']:
                if model_name in model_map:
                    group_models.append(model_map[model_name])
            if group_models:
                grouped.append({
                    'name': group_name,
                    'app_label': 'store',
                    'app_url': '#',
                    'has_module_perms': True,
                    'models': group_models,
                })

        auth_app = None
        for app in app_list:
            if app.get('app_label') in ('auth',):
                auth_app = app
                break
        if auth_app:
            system_group = None
            for g in grouped:
                if g['name'] == 'System':
                    system_group = g
                    break
            if not system_group:
                system_group = {'name': 'System', 'app_label': 'store', 'app_url': '#', 'has_module_perms': True, 'models': []}
                grouped.append(system_group)
            for m in auth_app.get('models', []):
                if m['object_name'] in ('User', 'Group'):
                    system_group['models'].append(m)

        return grouped

    def index(self, request, extra_context=None):
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_ago = now - timedelta(days=7)

        # General stats
        context = {
            'total_products': Product.objects.filter(trashed=False).count(),
            'active_products': Product.objects.filter(available=True, trashed=False).count(),
            'total_orders': Order.objects.count(),
            'pending_orders': Order.objects.filter(status='pending').count(),
            'processing_orders': Order.objects.filter(status='processing').count(),
            'completed_orders': Order.objects.filter(status='completed').count(),
            'cancelled_orders': Order.objects.filter(status='cancelled').count(),
            'total_revenue': Order.objects.filter(status__in=['completed', 'processing']).aggregate(t=Sum('total'))['t'] or 0,
            'total_partners': Partner.objects.count(),
            'total_dealers': Partner.objects.filter(is_dealer=True).count(),
            'total_sellers': Partner.objects.filter(is_seller=True).count(),
            'total_reviews': ProductReview.objects.count(),
            'avg_rating': ProductReview.objects.aggregate(a=Avg('rating'))['a'] or 0,
            'total_messages': Contact.objects.count(),
            'unread_messages_count': Contact.objects.filter(created__gte=timezone.now() - timedelta(days=7)).count(),
            'recent_orders': Order.objects.select_related('user').order_by('-created')[:10],
            'low_stock_products': Product.objects.filter(available=True, trashed=False).annotate(
                threshold=F('low_stock_threshold'),
            ).filter(
                Q(stock__gt=0, stock__lte=F('threshold')) |
                Q(color_variants__stock__gt=0, color_variants__stock__lte=F('threshold')) |
                Q(size_variants__stock__gt=0, size_variants__stock__lte=F('threshold'))
            ).distinct().order_by('stock')[:10],
            'recent_reviews': ProductReview.objects.select_related('product', 'user').order_by('-created')[:5],
            'today_orders': Order.objects.filter(created__date=today_start.date()).count(),
            'today_revenue': Order.objects.filter(
                created__date=today_start.date(),
                status__in=['completed', 'processing']
            ).aggregate(t=Sum('total'))['t'] or 0,
        }

        # Pending items needing attention
        pending_refunds = RefundRequest.objects.filter(status='pending').select_related('order', 'user')[:5]
        pending_withdrawals = WithdrawalRequest.objects.filter(status='pending').select_related('partner')[:5]
        pending_custom_orders = CustomOrder.objects.filter(status='pending').select_related('partner')[:5]
        unread_messages = Contact.objects.order_by('-created')[:5]
        recent_conversations = Conversation.objects.prefetch_related('participants', 'partner').annotate(
            msg_count=Count('messages'),
        ).order_by('-updated')[:8]
        context['pending_refunds'] = pending_refunds
        context['pending_withdrawals'] = pending_withdrawals
        context['pending_custom_orders'] = pending_custom_orders
        context['unread_messages'] = unread_messages
        context['recent_conversations'] = recent_conversations

        # Open tickets needing attention
        open_tickets = SupportTicket.objects.filter(status__in=['open', 'in_progress']).select_related('partner').annotate(
            reply_count=Count('replies'),
        ).order_by('-priority', '-created')[:8]
        context['open_tickets'] = open_tickets
        context['open_tickets_count'] = SupportTicket.objects.filter(status__in=['open', 'in_progress']).count()

        # 7-day sales data for chart
        daily_sales = OrderItem.objects.filter(
            order__created__gte=week_ago,
            order__status__in=['completed', 'processing'],
        ).annotate(
            date=TruncDate('order__created')
        ).values('date').annotate(
            revenue=Sum(F('price') * F('quantity')),
        ).order_by('date')

        chart_labels = []
        chart_data = []
        for entry in daily_sales:
            chart_labels.append(entry['date'].strftime('%d %b'))
            chart_data.append(float(entry['revenue'] or 0))

        context['chart_labels'] = chart_labels
        context['chart_data'] = chart_data

        # Status counts for doughnut
        status_labels = []
        status_counts = []
        order_counts = Order.objects.values('status').annotate(cnt=Count('id'))
        for oc in order_counts:
            status_labels.append(dict(Order.STATUS_CHOICES).get(oc['status'], oc['status']))
            status_counts.append(oc['cnt'])
        context['status_labels'] = status_labels
        context['status_counts'] = status_counts

        # Medicine-specific stats
        medicine_products = MedicineProduct.objects.filter(is_approved=True)
        medicine_partners = Partner.objects.filter(shop_style='medicine')
        medicine_authorized = medicine_partners.filter(has_medicine_access=True)
        pending_requests = MedicineProduct.objects.filter(is_approved=False).count()
        context['medicine_product_count'] = medicine_products.count()
        context['medicine_partner_count'] = medicine_partners.count()
        context['medicine_authorized_count'] = medicine_authorized.count()
        context['pending_medicine_requests'] = pending_requests
        context['recent_medicine_products'] = medicine_products.order_by('-created')[:10]

        return super().index(request, extra_context={**context, **(extra_context or {})})

    # ─── Medicine Admin Views ───

    def medicine_pending_requests(self, request):
        products = MedicineProduct.objects.filter(is_approved=False).select_related('requested_by').order_by('-created')

        return render(request, 'admin/store/medicine_pending_requests.html', {
            'products': products,
            'title': 'Pending Medicine Product Requests',
            'site_header': self.site_header,
            'site_title': self.site_title,
            'has_permission': self.has_permission(request),
        })

    def medicine_partner_access(self, request):
        partners = Partner.objects.filter(shop_style='medicine').annotate(
            medicine_product_count=Count('medicine_requests', filter=Q(
                medicine_requests__is_approved=True,
            ))
        ).order_by('name')

        if request.method == 'POST':
            partner_id = request.POST.get('partner_id')
            action = request.POST.get('action')
            partner = get_object_or_404(Partner, id=partner_id)

            if action == 'grant':
                partner.has_medicine_access = True
                partner.save(update_fields=['has_medicine_access'])
                MedicineSubscription.objects.get_or_create(partner=partner)
                messages.success(request, f'Medicine catalog access granted to {partner.name}')

            elif action == 'revoke':
                partner.has_medicine_access = False
                partner.save(update_fields=['has_medicine_access'])
                messages.warning(request, f'Medicine catalog access revoked for {partner.name}')

            elif action == 'enable_pos':
                partner.medicine_pos_enabled = True
                partner.save(update_fields=['medicine_pos_enabled'])
                sub, _ = MedicineSubscription.objects.get_or_create(partner=partner)
                if sub.status == 'inactive':
                    now = timezone.now()
                    sub.status = 'trial'
                    sub.trial_started_at = now
                    sub.trial_ends_at = now + timedelta(days=60)
                    sub.current_period_start = now
                    sub.current_period_end = now + timedelta(days=60)
                    sub.save()
                    messages.success(request, f'Medicine POS enabled for {partner.name} (60-day trial started)')
                else:
                    messages.warning(request, f'Medicine POS already active for {partner.name} (status: {sub.get_status_display()})')

            elif action == 'disable_pos':
                partner.medicine_pos_enabled = False
                partner.save(update_fields=['medicine_pos_enabled'])
                MedicineSubscription.objects.filter(partner=partner).update(status='cancelled')
                messages.warning(request, f'Medicine POS disabled for {partner.name}')

            return redirect('custom_admin:medicine-partner-access')

        partners_authorized = sum(1 for p in partners if p.has_medicine_access)
        partners_revoked = partners.count() - partners_authorized
        partners_pos_enabled = sum(1 for p in partners if p.medicine_pos_enabled)

        return render(request, 'admin/store/medicine_partner_access.html', {
            'partners': partners,
            'partners_authorized': partners_authorized,
            'partners_revoked': partners_revoked,
            'partners_pos_enabled': partners_pos_enabled,
            'title': 'Partner Medicine Access',
            'site_header': self.site_header,
            'site_title': self.site_title,
            'has_permission': self.has_permission(request),
        })

    # ─── Partner Bulk Import ───

    def _parse_rows(self, file):
        """Read rows from CSV or XLSX file, return list of dicts."""
        ext = file.name.rsplit('.', 1)[-1].lower()
        if ext == 'csv':
            data = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(data))
            return list(reader)
        elif ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append(dict(zip(headers, row)))
            return rows
        return None

    def _parse_row_data(self, raw_row, row_index):
        """Map verbose 30-column headers to model fields + auto-generate blanks."""
        row = {}
        for csv_key, val in raw_row.items():
            field = bd_data.resolve_header(csv_key)
            if val is not None:
                if isinstance(val, bool):
                    row[field] = 'Yes' if val else 'No'
                else:
                    row[field] = str(val).strip()
            else:
                row[field] = ''

        # Extract geographic values
        raw_upazila = bd_data.parse_upazila(row.get('address_upazila', ''))
        raw_district = bd_data.parse_district(row.get('address_district', ''))
        raw_division = row.get('address_division', '')

        # Auto-lookup division if blank
        if not raw_division and raw_district:
            raw_division = bd_data.lookup_division(raw_district)

        # Auto-fill Name from Upazila if blank
        name = (row.get('name') or '').strip()
        if not name and raw_upazila:
            name = raw_upazila

        # Generate slug
        slug = (row.get('slug') or '').strip()
        if not slug and raw_upazila and raw_district:
            slug = bd_data.generate_slug(raw_upazila, raw_district)
        elif not slug:
            slug_text = name or raw_upazila or 'partner'
            slug = slugify(slug_text)

        # Generate description (Bengali) if blank
        description = row.get('description', '').strip()
        if not description and raw_upazila and raw_district:
            description = bd_data.generate_description(raw_upazila, raw_district)

        # Sequential assets (logo, profile_image) — left empty (ImageField cannot accept string paths)
        logo = ''
        profile_image = ''

        # Default phone
        phone = (row.get('phone') or '').strip()
        if not phone:
            phone = '1000000000'

        # Boolean fields with Yes/No support
        show_products = str(row.get('show_products', 'Yes'))
        show_all = str(row.get('show_all_products_section', 'Yes'))
        show_random = str(row.get('show_random_products_section', 'Yes'))
        show_slider = str(row.get('show_slider', 'Yes'))
        has_medicine = str(row.get('has_medicine_access', 'No'))
        medicine_pos = str(row.get('medicine_pos_enabled', 'No'))

        return {
            'user': None,
            'name': name,
            'slug': slug,
            'logo': logo,
            'profile_image': profile_image,
            'banner': None,
            'description': description,
            'phone': phone,
            'address': (row.get('address') or '').strip(),
            'address_street': (row.get('address_street') or '').strip(),
            'address_division': raw_division,
            'address_district': raw_district,
            'address_upazila': raw_upazila,
            'voter_id': None,
            'is_dealer': bd_data.parse_bool(row.get('is_dealer', 'No')),
            'is_seller': bd_data.parse_bool(row.get('is_seller', 'Yes')),
            'is_union_agent': bd_data.parse_bool(row.get('is_union_agent', 'No')),
            'is_active': bd_data.parse_bool(row.get('is_active', 'Yes')),
            'parent_name': (row.get('parent') or '').strip(),
            'parent': None,
            'meta_title': (row.get('meta_title') or '').strip(),
            'meta_description': (row.get('meta_description') or '').strip(),
            'show_products': bd_data.parse_bool(show_products),
            'show_all_products_section': bd_data.parse_bool(show_all),
            'show_random_products_section': bd_data.parse_bool(show_random),
            'show_slider': bd_data.parse_bool(show_slider),
            'shop_style': bd_data.normalize_shop_style(row.get('shop_style', 'general')),
            'has_medicine_access': bd_data.parse_bool(has_medicine),
            'medicine_pos_enabled': bd_data.parse_bool(medicine_pos),
            'created': row.get('created', ''),
            'updated': row.get('updated', ''),
        }

    def _create_partner_row(self, data, row_index):
        """Create or update partner from parsed data dict."""
        messages_list = []
        name = data['name']

        if not name:
            return None, f'Name could not be determined (row {row_index})', 'error'

        # Resolve parent
        parent = data.get('parent')
        parent_name = data.get('parent_name', '')
        if parent_name:
            parent = Partner.objects.filter(name__iexact=parent_name).first()
            if not parent:
                messages_list.append(f'Parent "{parent_name}" not found')
            data['parent'] = parent

        defaults = {
            'slug': data['slug'],
            'description': data['description'],
            'phone': data['phone'],
            'address': data['address'],
            'address_street': data['address_street'],
            'address_division': data['address_division'],
            'address_district': data['address_district'],
            'address_upazila': data['address_upazila'],
            'is_dealer': data['is_dealer'],
            'is_seller': data['is_seller'],
            'is_union_agent': data['is_union_agent'],
            'is_active': data['is_active'],
            'parent': data['parent'],
            'meta_title': data['meta_title'],
            'meta_description': data['meta_description'],
            'show_products': data['show_products'],
            'show_all_products_section': data['show_all_products_section'],
            'show_random_products_section': data['show_random_products_section'],
            'show_slider': data['show_slider'],
            'shop_style': data['shop_style'],
            'has_medicine_access': data['has_medicine_access'],
            'medicine_pos_enabled': data['medicine_pos_enabled'],
        }

        # Always create a new record, even if name already exists
        slug = data['slug'] or slugify(name)
        if Partner.objects.filter(slug=slug).exists():
            from .models import _unique_slug
            slug = _unique_slug(Partner, 'slug', slug)
        defaults['slug'] = slug
        partner = Partner.objects.create(name=name, **defaults)
        created = True

        if created and data['created']:
            try:
                from django.utils.dateparse import parse_datetime
                dt = parse_datetime(data['created'])
                if dt:
                    Partner.objects.filter(pk=partner.pk).update(created=dt)
            except Exception:
                pass

        status = 'created'
        msg = ''
        if messages_list:
            msg = '; '.join(messages_list)
            if status == 'created':
                status = 'warning'

        return partner, msg, status

    def partner_bulk_import(self, request):
        results = None

        if request.method == 'POST' and request.FILES.get('file'):
            file = request.FILES['file']
            results = {'created': 0, 'errors': [], 'total': 0, 'rows': []}
            created_ids = []

            try:
                rows = self._parse_rows(file)
                if rows is None:
                    results['errors'].append('Unsupported file format. Use CSV (.csv) or Excel (.xlsx).')
                    rows = []

                for i, raw_row in enumerate(rows, 1):
                    try:
                        # Skip completely empty rows
                        if not any(v is not None and str(v).strip() for v in raw_row.values()):
                            results['rows'].append({'row': i, 'name': '', 'status': 'warning', 'message': 'Empty row skipped'})
                            continue
                        data = self._parse_row_data(raw_row, i)
                        partner, msg, status = self._create_partner_row(data, i)
                        name = data['name'] or ''

                        if status == 'error':
                            results['errors'].append(f'Row {i}: {msg}')
                            results['rows'].append({'row': i, 'name': name, 'status': 'error', 'message': msg})
                        else:
                            results['created'] += 1
                            created_ids.append(partner.pk)
                            results['rows'].append({'row': i, 'name': name, 'status': 'created', 'message': msg})

                    except Exception as e:
                        name_raw = raw_row.get('name', '') if isinstance(raw_row, dict) else ''
                        results['errors'].append(f'Row {i} ({name_raw}): {e}')
                        results['rows'].append({'row': i, 'name': str(name_raw)[:50], 'status': 'error', 'message': str(e)})

                results['total'] = len(rows)
                if created_ids:
                    request.session['undo_bulk_import_ids'] = created_ids
                    request.session['undo_bulk_import_count'] = len(created_ids)

            except Exception as e:
                results['errors'].append(f'File error: {e}')

        COLUMNS = [
            {'name': 'User', 'required': False, 'auto': False},
            {'name': 'Name', 'required': False, 'auto': True},
            {'name': 'Slug', 'required': False, 'auto': True},
            {'name': 'Logo', 'required': False, 'auto': True},
            {'name': 'Profile Image', 'required': False, 'auto': True},
            {'name': 'Banner', 'required': False, 'auto': False},
            {'name': 'Description', 'required': False, 'auto': True},
            {'name': 'Phone', 'required': False, 'auto': True},
            {'name': 'Address', 'required': False, 'auto': False},
            {'name': 'Street / Village / House / Flat', 'required': False, 'auto': False},
            {'name': 'Division', 'required': False, 'auto': True},
            {'name': 'District', 'required': False, 'auto': False},
            {'name': 'Upazila', 'required': True, 'auto': False},
            {'name': 'Voter Id Card', 'required': False, 'auto': False},
            {'name': 'Is Dealer', 'required': False, 'auto': False},
            {'name': 'Is Seller', 'required': False, 'auto': False},
            {'name': 'Is Union Agent', 'required': False, 'auto': False},
            {'name': 'Is Active', 'required': False, 'auto': False},
            {'name': 'Dealer (Parent)', 'required': False, 'auto': False},
            {'name': 'Meta Title', 'required': False, 'auto': False},
            {'name': 'Meta Description', 'required': False, 'auto': False},
            {'name': 'Show Products On Website', 'required': False, 'auto': False},
            {'name': 'Show "All Over Available Product" Section On Store Page', 'required': False, 'auto': False},
            {'name': 'Show "Random Product List" Section On Store Page', 'required': False, 'auto': False},
            {'name': 'Show Slider On Store Page', 'required': False, 'auto': False},
            {'name': 'Shop Style', 'required': False, 'auto': False},
            {'name': 'Medicine Catalog Access', 'required': False, 'auto': False},
            {'name': 'Medicine Pos Enabled', 'required': False, 'auto': False},
            {'name': 'Created', 'required': False, 'auto': True},
            {'name': 'Updated', 'required': False, 'auto': True},
        ]
        return render(request, 'admin/store/partner/bulk_import.html', {
            'title': 'Bulk Import Partners',
            'site_header': self.site_header,
            'site_title': self.site_title,
            'has_permission': self.has_permission(request),
            'results': results,
            'columns': COLUMNS,
            'can_undo': request.session.get('undo_bulk_import_ids') is not None,
            'undo_count': request.session.get('undo_bulk_import_count', 0),
        })

    def partner_bulk_import_undo(self, request):
        ids = request.session.pop('undo_bulk_import_ids', None)
        request.session.pop('undo_bulk_import_count', None)
        from django.contrib import messages
        if not ids:
            messages.warning(request, 'Nothing to undo.')
            from django.shortcuts import redirect
            return redirect('custom_admin:partner-bulk-import')

        from .models import Partner
        from django.db.models import ProtectedError

        try:
            # QuerySet.delete() cascades via ORM, bypassing admin permission checks
            deleted, _ = Partner.objects.filter(pk__in=ids).delete()
            messages.success(request, f'Undo successful — {deleted} partner(s) deleted.')
        except ProtectedError:
            messages.error(
                request,
                'Cannot undo: some partners have protected related objects. '
                'Delete those manually first.'
            )
        except Exception as e:
            messages.error(request, f'Undo failed: {e}')

        from django.shortcuts import redirect
        return redirect('custom_admin:partner-bulk-import')

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('medicine/pending/', self.admin_view(self.medicine_pending_requests), name='medicine-pending-requests'),
            path('medicine/partners/', self.admin_view(self.medicine_partner_access), name='medicine-partner-access'),
            path('partner/bulk-import/', self.admin_view(self.partner_bulk_import), name='partner-bulk-import'),
            path('partner/bulk-import-undo/', self.admin_view(self.partner_bulk_import_undo), name='partner-bulk-import-undo'),
        ]
        return custom_urls + urls


custom_admin_site = CustomAdminSite(name='custom_admin')
