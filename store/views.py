import csv
import json
import logging
import re
import secrets
import uuid
from datetime import timedelta
from decimal import Decimal
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Avg, Count, Sum, Max, F, ExpressionWrapper, DecimalField, OuterRef, Subquery
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django_ratelimit.decorators import ratelimit

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'svg', 'bmp', 'ico'}
MAX_UPLOAD_SIZE = 5 * 1024 * 1024  # 5 MB


def validate_upload(file_obj):
    ext = file_obj.name.rsplit('.', 1)[-1].lower() if '.' in file_obj.name else ''
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f'Invalid file type: .{ext}. Allowed: {", ".join(sorted(ALLOWED_EXTENSIONS))}')
    if file_obj.size > MAX_UPLOAD_SIZE:
        raise ValueError(f'File too large ({file_obj.size // 1024} KB). Maximum 5 MB.')
    return True
from .models import Product, Category, Partner, Cart, CartItem, BlogPost, Contact, Order, OrderItem, Slider, HomeBanner, ProductReview, PartnerBanner, PartnerNavMenu, Page, SideBanner, ProductColorVariant, ProductSizeVariant, ProductImage, LandingPage, ServerFee, Coupon, CouponUsage, CustomOrder, AdminCommission, PartnerWallet, WalletTransaction, WithdrawalRequest, PayoutMethod, WalletSettings, PlatformBalance, Wishlist, WishlistItem, Notification, ShippingRule, RefundRequest, SiteLogo, ManualPaymentMethod, PosOrder, PosOrderItem, PartnerSlider, Conversation, Message, ConversationReadStatus, ProductQA, SupportTicket, TicketReply, PRESET_COLORS, PRESET_SIZES, MedicineProduct, MedicinePosOrder, MedicinePosOrderItem, MedicineSubscription, WalletRechargeInstruction, SubscriptionPackage, Address, THEME_CHOICES


def home(request):
    base_qs = Product.objects.filter(available=True, trashed=False, is_published=True).filter(Q(partner__isnull=True) | Q(partner__show_products=True)).annotate(review_count=Count('reviews'), review_avg=Avg('reviews__rating'))
    products = base_qs[:16]
    partners = Partner.objects.all()[:20]
    latest_products = base_qs[:8]
    blog_posts = BlogPost.objects.all()[:3]
    new_products = base_qs.filter(label='new')[:12]
    hot_products = base_qs.filter(label='hot')[:12]
    discounted_products = base_qs.filter(label='discounted')[:12]
    partner_products = base_qs.filter(partner__isnull=False)[:12]
    best_selling_products = base_qs.order_by('?')[:12]
    categories = Category.objects.annotate(product_count=Count('products')).filter(product_count__gt=0)
    sliders = Slider.objects.filter(is_active=True).order_by('order')
    home_banners = HomeBanner.objects.filter(is_active=True).order_by('order')
    side_banners = SideBanner.objects.filter(partner__isnull=True, is_active=True)[:2]
    context = {
        'products': products,
        'partners': partners,
        'latest_products': latest_products,
        'blog_posts': blog_posts,
        'new_products': new_products,
        'hot_products': hot_products,
        'discounted_products': discounted_products,
        'partner_products': partner_products,
        'best_selling_products': best_selling_products,
        'categories': categories,
        'sliders': sliders,
        'home_banners': home_banners,
        'side_banners': side_banners,
    }
    return render(request, 'store/home.html', context)


def shop_grid(request):
    products = Product.objects.filter(available=True, trashed=False, is_published=True).filter(Q(partner__isnull=True) | Q(partner__show_products=True)).annotate(review_count=Count('reviews'), review_avg=Avg('reviews__rating'))
    categories = Category.objects.annotate(product_count=Count('products')).filter(product_count__gt=0)
    category_slug = request.GET.get('category')
    if category_slug:
        products = products.filter(category__slug=category_slug)
    search_query = request.GET.get('q')
    if search_query:
        products = products.filter(Q(name__icontains=search_query) | Q(description__icontains=search_query))
    label_filter = request.GET.get('label')
    if label_filter:
        products = products.filter(label=label_filter)
    min_price = request.GET.get('min_price')
    if min_price:
        try:
            products = products.filter(price__gte=Decimal(min_price))
        except Exception:
            pass
    max_price = request.GET.get('max_price')
    if max_price:
        try:
            products = products.filter(price__lte=Decimal(max_price))
        except Exception:
            pass

    sort = request.GET.get('sort', '')
    sort_map = {
        'price_asc': 'price',
        'price_desc': '-price',
        'name_asc': 'name',
        'name_desc': '-name',
        'newest': '-created',
    }
    if sort in sort_map:
        products = products.order_by(sort_map[sort])
    else:
        products = products.order_by('-created')

    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    if request.GET.get('ajax'):
        html = render_to_string('store/product_cards_partial.html', {'products': page_obj.object_list}, request)
        return JsonResponse({
            'html': html,
            'has_next': page_obj.has_next(),
            'total_pages': paginator.num_pages,
        })
    context = {
        'page_obj': page_obj,
        'products': page_obj.object_list,
        'categories': categories,
        'current_sort': sort,
        'current_category': category_slug or '',
        'current_label': label_filter or '',
        'search_query': search_query or '',
        'min_price': min_price or '',
        'max_price': max_price or '',
    }
    return render(request, 'store/shop_grid.html', context)


def quick_view(request, product_id):
    product = get_object_or_404(Product.objects.filter(trashed=False).annotate(review_count=Count('reviews'), review_avg=Avg('reviews__rating')), id=product_id)
    html = render_to_string('store/quick_view_partial.html', {'product': product}, request)
    return JsonResponse({'html': html})


def product_detail(request, slug):
    product = get_object_or_404(Product.objects.filter(trashed=False), slug=slug)
    related_products = Product.objects.filter(category=product.category, available=True, trashed=False, is_published=True).filter(Q(partner__isnull=True) | Q(partner__show_products=True)).exclude(id=product.id)[:4]
    can_review = False
    review_avg = 0
    reviews = product.reviews.all()
    if reviews:
        review_avg = sum(r.rating for r in reviews) / len(reviews)
    if request.user.is_authenticated:
        has_purchased = Order.objects.filter(
            user=request.user,
            items__product=product,
            status__in=['pending', 'processing', 'completed']
        ).exists()
        can_review = has_purchased or request.user.is_superuser
    recently = request.session.get('recently_viewed', [])
    if product.id not in recently:
        recently.insert(0, product.id)
        request.session['recently_viewed'] = recently[:10]
    return render(request, 'store/product_detail.html', {
        'product': product,
        'related_products': related_products,
        'can_review': can_review,
        'review_avg': review_avg,
        'reviews': reviews,
        'qa_list': product.qa_pairs.select_related('user', 'answered_by').all(),
    })


def partner_list(request):
    partners = Partner.objects.all()
    q = request.GET.get('q', '').strip()
    if q:
        partners = partners.filter(
            Q(address_street__icontains=q) |
            Q(address_division__icontains=q) |
            Q(address_district__icontains=q) |
            Q(address_upazila__icontains=q) |
            Q(address__icontains=q)
        )
    return render(request, 'store/partner_list.html', {'partners': partners, 'q': q})


def dealer_list(request):
    dealers = Partner.objects.filter(is_dealer=True)
    return render(request, 'store/dealer_list.html', {'dealers': dealers})


def seller_list(request):
    sellers = Partner.objects.filter(is_seller=True)
    return render(request, 'store/seller_list.html', {'sellers': sellers})


def partner_detail(request, slug):
    partner = get_object_or_404(Partner, slug=slug)
    partner_products = Product.objects.filter(partner=partner, available=True, trashed=False, is_published=True, partner__show_products=True).annotate(review_count=Count('reviews'), review_avg=Avg('reviews__rating'))
    base_qs = Product.objects.filter(available=True, trashed=False, is_published=True).filter(Q(partner__isnull=True) | Q(partner__show_products=True)).annotate(review_count=Count('reviews'), review_avg=Avg('reviews__rating'))
    partner_nav = PartnerNavMenu.objects.filter(partner=partner, is_active=True, parent=None).prefetch_related('children')
    side_banners = SideBanner.objects.filter(partner=partner, is_active=True)[:2]
    partner_sliders = PartnerSlider.objects.filter(partner=partner, is_active=True).order_by('order')
    page_number = request.GET.get('page', 1)
    paginator = Paginator(partner_products, 999999)
    page_obj = paginator.get_page(page_number)
    context = {
        'partner': partner,
        'products': page_obj,
        'all_products': base_qs[:16],
        'new_products': base_qs.filter(label='new')[:12],
        'hot_products': base_qs.filter(label='hot')[:12],
        'discounted_products': base_qs.filter(label='discounted')[:12],
        'latest_products': base_qs[:8],
        'extra_banners': partner.extra_banners.all(),
        'partner_nav_items': partner_nav,
        'has_partner_nav': partner_nav.exists(),
        'side_banners': side_banners,
        'sellers': partner.sellers.all() if (partner.is_dealer or partner.is_union_agent) else [],
        'union_agents': Partner.objects.filter(parent=partner, is_union_agent=True) if partner.is_dealer else [],
        'show_all_products_section': partner.show_all_products_section,
        'show_random_products_section': partner.show_random_products_section,
        'show_slider': partner.show_slider,
        'partner_sliders': partner_sliders,
    }
    return render(request, 'store/partner_store.html', context)


def page_detail(request, slug):
    page = get_object_or_404(Page, slug=slug, is_active=True)
    template_map = {
        'default': 'store/page_detail.html',
        'full_width': 'store/page_full_width.html',
        'landing': 'store/page_landing.html',
    }
    template = template_map.get(page.template_name, 'store/page_detail.html')
    import bleach
    allowed_tags = ['p', 'br', 'b', 'i', 'u', 'em', 'strong', 'a', 'ul', 'ol', 'li', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'blockquote', 'pre', 'code', 'span', 'div', 'img', 'hr', 'table', 'thead', 'tbody', 'tr', 'th', 'td', 'figure', 'figcaption', 'video', 'source', 'iframe']
    allowed_attrs = {'a': ['href', 'title', 'target', 'rel'], 'img': ['src', 'alt', 'width', 'height', 'class'], 'video': ['src', 'controls', 'width', 'height'], 'source': ['src', 'type'], 'iframe': ['src', 'width', 'height', 'frameborder', 'allowfullscreen'], '*': ['class', 'style', 'id']}
    page.content = bleach.clean(page.content, tags=allowed_tags, attributes=allowed_attrs, strip=True)
    return render(request, template, {'page': page})


def landing_page(request, slug):
    landing = get_object_or_404(LandingPage, slug=slug, is_active=True)
    all_partners = Partner.objects.all()
    return render(request, 'store/landing_page.html', {'landing': landing, 'all_partners': all_partners})


def blog_list(request):
    posts = BlogPost.objects.all()
    paginator = Paginator(posts, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'store/blog_list.html', {'posts': page_obj.object_list, 'page_obj': page_obj})


def blog_detail(request, slug):
    post = get_object_or_404(BlogPost, slug=slug)
    return render(request, 'store/blog_detail.html', {'post': post})


def contact(request):
    if request.method == 'POST':
        Contact.objects.create(
            name=request.POST.get('name'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone', ''),
            subject=request.POST.get('subject', ''),
            message=request.POST.get('message'),
        )
        return redirect('contact')
    return render(request, 'store/contact.html')


BACKEND = 'django.contrib.auth.backends.ModelBackend'


@ratelimit(key='ip', rate='10/m', method='POST')
def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user, backend=BACKEND)
            next_url = request.GET.get('next', '/')
            if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                next_url = '/'
            return redirect(next_url)
        logger.warning('Failed login attempt for %s from %s', request.POST.get('username', ''), request.META.get('REMOTE_ADDR', ''))
    else:
        form = AuthenticationForm()
    return render(request, 'store/login.html', {'form': form})


@ratelimit(key='ip', rate='5/h', method='POST')
def register_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user, backend=BACKEND)
            return redirect('/')
    else:
        form = UserCreationForm()
    return render(request, 'store/register.html', {'form': form})


def add_to_cart(request, product_id):
    product = get_object_or_404(Product.objects.filter(trashed=False), id=product_id)
    if request.method == 'POST':
        if request.user.is_authenticated:
            cart, _ = Cart.objects.get_or_create(user=request.user, is_active=True)
        else:
            if not request.session.session_key:
                request.session.save()
            session_id = request.session.session_key
            cart, _ = Cart.objects.get_or_create(session_id=session_id, is_active=True)
        qty = request.POST.get('quantity', '1')
        try:
            qty = int(qty)
            if qty < 1:
                qty = 1
        except (ValueError, TypeError):
            qty = 1

        color_id = request.POST.get('color_variant_id', '')
        size_id = request.POST.get('size_variant_id', '')
        color_variant = ProductColorVariant.objects.filter(id=color_id, product=product).first() if color_id else None
        size_variant = ProductSizeVariant.objects.filter(id=size_id, product=product).first() if size_id else None

        # Check variant stock
        available_stock = product.stock
        if color_variant and color_variant.stock > 0:
            available_stock = color_variant.stock
        if size_variant and size_variant.stock > 0:
            available_stock = size_variant.stock
        if qty > available_stock:
            messages.error(request, f'Only {available_stock} available for the selected variant.')
            return redirect('product_detail', slug=product.slug)

        cart_item, created = CartItem.objects.get_or_create(
            cart=cart,
            product=product,
            color_variant=color_variant,
            size_variant=size_variant,
        )
        if not created:
            new_qty = cart_item.quantity + qty
            if new_qty > available_stock:
                messages.error(request, f'Cannot add more — only {available_stock} available.')
                return redirect('cart')
            cart_item.quantity = new_qty
        else:
            cart_item.quantity = qty
        cart_item.save()
        cache.delete(f'cart_{request.user.id}') if request.user.is_authenticated else None
        return redirect('cart')
    return redirect('product_detail', slug=product.slug)


@ratelimit(key='user_or_ip', rate='5/m', block=True)
def buy_now(request, product_id):
    product = get_object_or_404(Product.objects.filter(trashed=False), id=product_id)
    if request.user.is_authenticated:
        cart, _ = Cart.objects.get_or_create(user=request.user, is_active=True)
    else:
        if not request.session.session_key:
            request.session.save()
        session_id = request.session.session_key
        cart, _ = Cart.objects.get_or_create(session_id=session_id, is_active=True)
    qty = request.POST.get('quantity', '1')
    try:
        qty = int(qty)
        if qty < 1:
            qty = 1
    except (ValueError, TypeError):
        qty = 1

    color_id = request.POST.get('color_variant_id', '')
    size_id = request.POST.get('size_variant_id', '')
    color_variant = ProductColorVariant.objects.filter(id=color_id, product=product).first() if color_id else None
    size_variant = ProductSizeVariant.objects.filter(id=size_id, product=product).first() if size_id else None

    # Check variant stock
    available_stock = product.stock
    if color_variant and color_variant.stock > 0:
        available_stock = color_variant.stock
    if size_variant and size_variant.stock > 0:
        available_stock = size_variant.stock
    if qty > available_stock:
        messages.error(request, f'Only {available_stock} available for the selected variant.')
        return redirect('product_detail', slug=product.slug)

    cart_item, created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        color_variant=color_variant,
        size_variant=size_variant,
    )
    if not created:
        new_qty = cart_item.quantity + qty
        if new_qty > available_stock:
            messages.error(request, f'Cannot add more — only {available_stock} available.')
            return redirect('cart')
        cart_item.quantity = new_qty
    else:
        cart_item.quantity = qty
    cart_item.save()
    return redirect('checkout')


def cart_view(request):
    cart = None
    if request.user.is_authenticated:
        cart = Cart.objects.filter(user=request.user, is_active=True).first()
    else:
        session_id = request.session.session_key
        if session_id:
            cart = Cart.objects.filter(session_id=session_id, is_active=True).first()
    return render(request, 'store/cart.html', {'cart': cart})


def remove_from_cart(request, item_id):
    if request.user.is_authenticated:
        cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    else:
        session_id = request.session.session_key
        cart_item = get_object_or_404(CartItem, id=item_id, cart__session_id=session_id, cart__user__isnull=True)
    cart_item.delete()
    return redirect('cart')


def update_cart(request, item_id):
    if request.user.is_authenticated:
        cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    else:
        session_id = request.session.session_key
        cart_item = get_object_or_404(CartItem, id=item_id, cart__session_id=session_id, cart__user__isnull=True)
    quantity = request.POST.get('quantity', 1)
    try:
        cart_item.quantity = int(quantity)
        if cart_item.quantity < 1:
            cart_item.delete()
        else:
            cart_item.save()
    except ValueError:
        pass
    return redirect('cart')


def logout_view(request):
    logout(request)
    return render(request, 'store/logout.html')


def get_or_create_active_cart(request):
    if request.user.is_authenticated:
        cart, _ = Cart.objects.get_or_create(user=request.user, is_active=True)
    else:
        if not request.session.session_key:
            request.session.save()
        cart, _ = Cart.objects.get_or_create(session_id=request.session.session_key, is_active=True)
    return cart


def checkout(request):
    cart = get_or_create_active_cart(request)
    if not cart.items.exists():
        return redirect('cart')

    # Check if partner delivery is enabled in site settings
    site_settings = SiteLogo.objects.first()
    partner_delivery_enabled = site_settings.partner_delivery_enabled if site_settings else True

    # Fetch partners with addresses for selection
    partners = []
    if partner_delivery_enabled:
        partners = Partner.objects.exclude(
            address_street='', address=''
        ).distinct()[:50]

    name = phone = address = ''
    selected_partner = None
    partner_street = partner_division = partner_district = partner_upazila = ''
    saved_addresses = []

    if request.user.is_authenticated:
        name = request.user.first_name
        saved_addresses = list(request.user.addresses.all())
        default_addr = next((a for a in saved_addresses if a.is_default), saved_addresses[0] if saved_addresses else None)
        if default_addr:
            addr_parts = [default_addr.street_address, default_addr.area, default_addr.district, default_addr.division]
            address = ', '.join(p for p in addr_parts if p)
            phone = default_addr.phone or ''
            name = default_addr.name or request.user.first_name
        else:
            last_order = Order.objects.filter(user=request.user).first()
            if last_order:
                phone = last_order.phone
                address = last_order.address

    # Calculate subtotal
    subtotal = cart.total_amount()

    # Get active server fees
    active_fees = ServerFee.objects.filter(is_active=True).order_by('order')

    # Check for applied coupon from session
    coupon_code = request.session.get('applied_coupon', None)
    coupon = None
    discount_amount = 0
    if coupon_code:
        try:
            coupon = Coupon.objects.get(code=coupon_code, is_active=True)
            valid, msg = coupon.is_valid(request.user if request.user.is_authenticated else None,
                                         subtotal, [item.product for item in cart.items.all()])
            if valid:
                discount_amount = coupon.calculate_discount(subtotal, [item.product for item in cart.items.all()])
            else:
                coupon = None
                request.session.pop('applied_coupon', None)
        except Coupon.DoesNotExist:
            coupon = None
            request.session.pop('applied_coupon', None)

    # Calculate applicable fees
    applicable_fees = []
    for fee in active_fees:
        if fee.min_order_amount and subtotal < fee.min_order_amount:
            continue
        fee_amount = fee.calculate(subtotal)
        if fee_amount > 0:
            applicable_fees.append({
                'name': fee.name,
                'value': fee_amount,
                'link_url': fee.link_url,
                'link_text': fee.link_text,
            })
    fees_total = sum(f['value'] for f in applicable_fees)

    # Calculate final total
    total_after_discount = subtotal - discount_amount
    grand_total = total_after_discount + fees_total

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        delivery_option = request.POST.get('delivery_option', 'manual')
        delivery_partner_id = request.POST.get('delivery_partner', '')

        partner_street = request.POST.get('partner_street', '').strip()
        partner_division = request.POST.get('partner_division', '').strip()
        partner_district = request.POST.get('partner_district', '').strip()
        partner_upazila = request.POST.get('partner_upazila', '').strip()
        address = request.POST.get('address', '').strip()

        # Build address from structured fields
        if delivery_option == 'partner' and delivery_partner_id:
            try:
                selected_partner = Partner.objects.get(id=delivery_partner_id)
                address_parts = [
                    partner_street or selected_partner.address_street,
                    partner_upazila or selected_partner.address_upazila,
                    partner_district or selected_partner.address_district,
                    partner_division or selected_partner.address_division,
                ]
                address = ', '.join(p for p in address_parts if p)
                if address:
                    address += ', Bangladesh.'
            except Partner.DoesNotExist:
                pass

        errors = {}
        if not name:
            errors['name'] = 'Name is required.'
        if not phone:
            errors['phone'] = 'Phone number is required.'
        if not address:
            errors['address'] = 'Address is required.'
        if errors:
            return render(request, 'store/checkout.html', {
                'cart': cart,
                'errors': errors,
                'name': name,
                'phone': phone,
                'address': address,
                'subtotal': subtotal,
                'active_fees': active_fees,
                'applicable_fees': applicable_fees,
                'fees_total': fees_total,
                'coupon': coupon,
                'discount_amount': discount_amount,
                'grand_total': grand_total,
                'partners': partners,
                'delivery_option': delivery_option,
                'selected_partner': selected_partner,
                'partner_street': partner_street,
                'partner_division': partner_division,
                'partner_district': partner_district,
                'partner_upazila': partner_upazila,
                'partner_delivery_enabled': partner_delivery_enabled,
                'saved_addresses': saved_addresses,
            })

        user = request.user if request.user.is_authenticated else None
        if not user:
            username = phone.strip()
            random_password = secrets.token_urlsafe(12)
            try:
                user = User.objects.get(username=username)
            except User.DoesNotExist:
                user = User.objects.create_user(
                    username=username,
                    password=random_password,
                    first_name=name.split()[0] if name.split() else name,
                )
                login(request, user, backend=BACKEND)
                request.session['auto_account_phone'] = phone.strip()
                request.session.pop('auto_account_password', None)
            else:
                user.first_name = name
                user.save(update_fields=['first_name'])

        # Stock check before placing order
        for item in cart.items.all():
            available = item.product.stock
            if item.color_variant and item.color_variant.stock > 0:
                available = item.color_variant.stock
            if item.size_variant and item.size_variant.stock > 0:
                available = item.size_variant.stock
            if available < item.quantity:
                return render(request, 'store/checkout.html', {
                    'cart': cart,
                    'errors': {'stock': f'Insufficient stock for {item.product.name}. Only {available} available.'},
                    'name': name,
                    'phone': phone,
                    'address': address,
                    'subtotal': subtotal,
                    'active_fees': active_fees,
                    'applicable_fees': applicable_fees,
                    'fees_total': fees_total,
                    'coupon': coupon,
                    'discount_amount': discount_amount,
                    'grand_total': grand_total,
                    'partners': partners,
                    'delivery_option': delivery_option,
                    'selected_partner': selected_partner,
                    'partner_street': partner_street,
                    'partner_division': partner_division,
                    'partner_district': partner_district,
                    'partner_upazila': partner_upazila,
                    'partner_delivery_enabled': partner_delivery_enabled,
                    'saved_addresses': saved_addresses,
                })

        order = Order.objects.create(
            user=user,
            name=name,
            phone=phone,
            address=address,
            subtotal=subtotal,
            discount_amount=discount_amount,
            fees_total=fees_total,
            total=grand_total,
            coupon=coupon,
            applied_fees=json.dumps([{**f, 'value': float(f['value'])} for f in applicable_fees]) if applicable_fees else '',
            payment_method='Cash on Delivery',
            delivery_partner=selected_partner,
        )
        for item in cart.items.all():
            OrderItem.objects.create(
                order=order,
                product=item.product,
                product_name=item.product.name,
                price=item.product.price,
                quantity=item.quantity,
                cost_price=item.product.cost_price,
                color_name=item.color_variant.color_name if item.color_variant else '',
                color_code=item.color_variant.color_code if item.color_variant else '',
                size_name=item.size_variant.size_name if item.size_variant else '',
            )

        # Record coupon usage
        if coupon and discount_amount > 0:
            CouponUsage.objects.create(
                coupon=coupon,
                order=order,
                user=user if user.is_authenticated else None,
                discount_amount=discount_amount,
            )
            Coupon.objects.filter(id=coupon.id).update(used_count=coupon.used_count + 1)

        # Clear coupon from session
        request.session.pop('applied_coupon', None)

        # Decrement stock (variant stock takes priority if > 0)
        for item in cart.items.all():
            if item.color_variant and item.color_variant.stock > 0:
                ProductColorVariant.objects.filter(pk=item.color_variant.pk).update(
                    stock=F('stock') - item.quantity
                )
            elif item.size_variant and item.size_variant.stock > 0:
                ProductSizeVariant.objects.filter(pk=item.size_variant.pk).update(
                    stock=F('stock') - item.quantity
                )
            else:
                Product.objects.filter(pk=item.product.pk).update(
                    stock=F('stock') - item.quantity
                )

        cart.is_active = False
        cart.save()
        return redirect('order_confirmation', order_id=order.id)

    return render(request, 'store/checkout.html', {
        'cart': cart,
        'name': name,
        'phone': phone,
        'address': address,
        'subtotal': subtotal,
        'active_fees': active_fees,
        'applicable_fees': applicable_fees,
        'fees_total': fees_total,
        'coupon': coupon,
        'discount_amount': discount_amount,
        'grand_total': grand_total,
        'partners': partners,
        'delivery_option': 'manual' if partner_delivery_enabled else 'manual',
        'selected_partner': None,
        'partner_street': '',
        'partner_division': '',
        'partner_district': '',
        'partner_upazila': '',
        'partner_delivery_enabled': partner_delivery_enabled,
        'saved_addresses': saved_addresses,
    })


def order_confirmation(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    auto_phone = request.session.pop('auto_account_phone', None)
    auto_password = request.session.pop('auto_account_password', None)
    return render(request, 'store/order_confirmation.html', {
        'order': order,
        'auto_phone': auto_phone,
        'auto_password': auto_password,
    })


@login_required
def public_print_invoice(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return render(request, 'admin/store/print_invoice.html', {'order': order})


@login_required
def invoice_pdf(request, order_id):
    from weasyprint import HTML
    order = get_object_or_404(Order, id=order_id, user=request.user)
    html = render_to_string('admin/store/print_invoice.html', {'order': order}, request)
    pdf = HTML(string=html).write_pdf()
    from django.http import HttpResponse
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="invoice-{order.id}.pdf"'
    return response


def apply_coupon(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})
    code = request.POST.get('code', '').strip()
    if not code:
        return JsonResponse({'success': False, 'error': 'Please enter a coupon code.'})
    try:
        cart = get_or_create_active_cart(request)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Your cart is empty.'})
    if not cart.items.exists():
        return JsonResponse({'success': False, 'error': 'Your cart is empty.'})

    try:
        coupon = Coupon.objects.get(code__iexact=code, is_active=True)
    except Coupon.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid coupon code.'})

    subtotal = cart.total_amount()
    cart_products = [item.product for item in cart.items.all()]
    user = request.user if request.user.is_authenticated else None
    valid, msg = coupon.is_valid(user, subtotal, cart_products)
    if not valid:
        return JsonResponse({'success': False, 'error': msg})

    discount = coupon.calculate_discount(subtotal, cart_products)

    # Store coupon in session
    request.session['applied_coupon'] = coupon.code

    # Calculate fees for display
    active_fees = ServerFee.objects.filter(is_active=True).order_by('order')
    fees_total = 0
    for fee in active_fees:
        if fee.min_order_amount and subtotal < fee.min_order_amount:
            continue
        fees_total += fee.calculate(subtotal)
    grand_total = subtotal - discount + fees_total

    return JsonResponse({
        'success': True,
        'message': f'Coupon "{coupon.code}" applied! You saved ৳{discount}',
        'discount': float(discount),
        'grand_total': float(grand_total),
        'fees_total': float(fees_total),
    })


@login_required
def dashboard(request):
    try:
        partner = request.user.partner
        products_count = partner.products.count()
        orders = Order.objects.filter(items__product__partner=partner).distinct().count()
        custom_orders_pending = CustomOrder.objects.filter(partner=partner, status='pending').count()

        # Profit calculations
        partner_items = OrderItem.objects.filter(product__partner=partner, cost_price__isnull=False)
        revenue = partner_items.annotate(line_total=F('price') * F('quantity')).aggregate(t=Sum('line_total'))['t'] or 0
        cost = partner_items.annotate(line_cost=F('cost_price') * F('quantity')).aggregate(t=Sum('line_cost'))['t'] or 0
        gross_profit = revenue - cost

        commission = None
        net_profit = gross_profit
        commission_rate = 0
        try:
            ac = AdminCommission.objects.first()
            if ac and ac.is_active:
                commission_rate = ac.commission_percentage
                commission = ac.calculate_commission(gross_profit, revenue)
                net_profit = gross_profit - commission
        except Exception:
            pass

        wallet = PartnerWallet.objects.filter(partner=partner).first()
        pending_withdrawals = WithdrawalRequest.objects.filter(partner=partner, status='pending').count()
        unread_messages = _unread_message_count(request.user)
        open_tickets = SupportTicket.objects.filter(partner=partner, status__in=['open', 'in_progress']).count()
        low_stock_products = partner.products.filter(
            Q(stock__lte=F('low_stock_threshold'), available=True, trashed=False) |
            Q(color_variants__stock__gt=0, color_variants__stock__lte=F('low_stock_threshold'), available=True, trashed=False) |
            Q(size_variants__stock__gt=0, size_variants__stock__lte=F('low_stock_threshold'), available=True, trashed=False)
        ).distinct().order_by('stock')

        return render(request, 'store/dashboard.html', {
            'partner': partner,
            'products_count': products_count,
            'orders_count': orders,
            'custom_orders_pending': custom_orders_pending,
            'total_revenue': revenue,
            'total_cost': cost,
            'gross_profit': gross_profit,
            'net_profit': net_profit,
            'commission': commission,
            'commission_rate': commission_rate,
            'profit_margin': round(net_profit / revenue * 100, 2) if revenue > 0 else 0,
            'wallet': wallet,
            'pending_withdrawals': pending_withdrawals,
            'unread_messages': unread_messages,
            'open_tickets': open_tickets,
            'low_stock_products': low_stock_products,
        })
    except Partner.DoesNotExist:
        orders = Order.objects.filter(user=request.user).prefetch_related('items__product').order_by('-created')
        unread_messages = _unread_message_count(request.user)
        return render(request, 'store/dashboard.html', {'partner': None, 'orders': orders, 'unread_messages': unread_messages})


@login_required
def dashboard_toggle_products(request):
    try:
        partner = request.user.partner
        partner.show_products = not partner.show_products
        partner.save()
        if partner.show_products:
            messages.success(request, 'Your products are now visible on the website.')
        else:
            messages.warning(request, 'Your products are now hidden from the website.')
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
    return redirect('dashboard')


@login_required
def dashboard_toggle_all_products_section(request):
    try:
        partner = request.user.partner
        partner.show_all_products_section = not partner.show_all_products_section
        partner.save()
        if partner.show_all_products_section:
            messages.success(request, '"All Over Available Product" section is now visible on your store page.')
        else:
            messages.warning(request, '"All Over Available Product" section is now hidden from your store page.')
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
    return redirect('dashboard')


@login_required
def dashboard_toggle_random_products_section(request):
    try:
        partner = request.user.partner
        partner.show_random_products_section = not partner.show_random_products_section
        partner.save()
        if partner.show_random_products_section:
            messages.success(request, '"Random product list" section is now visible on your store page.')
        else:
            messages.warning(request, '"Random product list" section is now hidden from your store page.')
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
    return redirect('dashboard')


@login_required
def dashboard_toggle_slider(request):
    try:
        partner = request.user.partner
        partner.show_slider = not partner.show_slider
        partner.save()
        if partner.show_slider:
            messages.success(request, 'Slider is now visible on your store page.')
        else:
            messages.warning(request, 'Slider is now hidden from your store page.')
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
    return redirect('dashboard_slider_list')


@login_required
def cancel_order(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    if order.status == 'pending':
        order.status = 'cancelled'
        order.save(update_fields=['status'])
    return redirect('dashboard')


# ─── Custom Orders (Seller Dashboard) ───

@login_required
def dashboard_custom_orders(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    custom_orders = CustomOrder.objects.filter(partner=partner).order_by('-created')
    return render(request, 'store/dashboard_custom_orders.html', {
        'partner': partner,
        'custom_orders': custom_orders,
    })


@login_required
def dashboard_custom_order_create(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')

    if request.method == 'POST':
        product_title = request.POST.get('product_title', '').strip()
        product_description = request.POST.get('product_description', '').strip()
        price = request.POST.get('price', '').strip()
        quantity = request.POST.get('quantity', '1').strip()
        buyer_name = request.POST.get('buyer_name', '').strip()
        buyer_phone = request.POST.get('buyer_phone', '').strip()
        buyer_address = request.POST.get('buyer_address', '').strip()
        buyer_email = request.POST.get('buyer_email', '').strip()

        errors = {}
        if not product_title:
            errors['product_title'] = 'Product title is required.'
        if not price:
            errors['price'] = 'Price is required.'
        if not buyer_name:
            errors['buyer_name'] = 'Buyer name is required.'
        if not buyer_phone:
            errors['buyer_phone'] = 'Buyer phone is required.'

        if errors:
            return render(request, 'store/dashboard_custom_order_form.html', {
                'partner': partner,
                'errors': errors,
                'form_data': request.POST,
            })

        try:
            price_val = float(price)
        except ValueError:
            return render(request, 'store/dashboard_custom_order_form.html', {
                'partner': partner,
                'errors': {'price': 'Invalid price.'},
                'form_data': request.POST,
            })

        try:
            qty_val = int(quantity)
        except ValueError:
            qty_val = 1

        CustomOrder.objects.create(
            partner=partner,
            product_title=product_title,
            product_description=product_description,
            price=price_val,
            quantity=qty_val,
            buyer_name=buyer_name,
            buyer_phone=buyer_phone,
            buyer_address=buyer_address,
            buyer_email=buyer_email,
            seller_name=partner.name,
            seller_phone=partner.phone,
        )
        return redirect('dashboard_custom_orders')

    return render(request, 'store/dashboard_custom_order_form.html', {'partner': partner})


@login_required
def dashboard_custom_order_edit(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    custom_order = get_object_or_404(CustomOrder, pk=pk, partner=partner)
    if custom_order.status != 'pending':
        return redirect('dashboard_custom_orders')

    if request.method == 'POST':
        custom_order.product_title = request.POST.get('product_title', '').strip()
        custom_order.product_description = request.POST.get('product_description', '').strip()
        price_raw = request.POST.get('price', '').strip()
        if not price_raw:
            messages.error(request, 'Price is required.')
            return redirect('dashboard_custom_order_edit', pk=pk)
        try:
            price_val = Decimal(price_raw)
            if price_val <= 0:
                raise ValueError
        except Exception:
            messages.error(request, 'Invalid price.')
            return redirect('dashboard_custom_order_edit', pk=pk)
        custom_order.price = price_val
        custom_order.quantity = int(request.POST.get('quantity', '1'))
        custom_order.buyer_name = request.POST.get('buyer_name', '').strip()
        custom_order.buyer_phone = request.POST.get('buyer_phone', '').strip()
        custom_order.buyer_address = request.POST.get('buyer_address', '').strip()
        custom_order.buyer_email = request.POST.get('buyer_email', '').strip()
        custom_order.save()
        return redirect('dashboard_custom_orders')

    return render(request, 'store/dashboard_custom_order_form.html', {
        'partner': partner,
        'custom_order': custom_order,
        'form_data': {
            'product_title': custom_order.product_title,
            'product_description': custom_order.product_description,
            'price': custom_order.price,
            'quantity': custom_order.quantity,
            'buyer_name': custom_order.buyer_name,
            'buyer_phone': custom_order.buyer_phone,
            'buyer_address': custom_order.buyer_address,
            'buyer_email': custom_order.buyer_email,
        },
    })


@login_required
def dashboard_custom_order_delete(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    custom_order = get_object_or_404(CustomOrder, pk=pk, partner=partner)
    if custom_order.status == 'pending':
        custom_order.delete()
    return redirect('dashboard_custom_orders')


@login_required
def dashboard_become_seller(request):
    try:
        request.user.partner
        return redirect('dashboard')
    except Partner.DoesNotExist:
        pass
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        if not name:
            return render(request, 'store/dashboard_become_seller.html', {'error': 'Name is required.'})
        Partner.objects.create(
            user=request.user,
            name=name,
            phone=phone,
            is_seller=True,
        )
        return redirect('dashboard')
    return render(request, 'store/dashboard_become_seller.html')


@login_required
def dashboard_sellers(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if not partner.is_dealer and not partner.is_union_agent:
        return redirect('dashboard')
    sellers = partner.sellers.all()
    return render(request, 'store/dashboard_sellers.html', {
        'partner': partner,
        'sellers': sellers,
    })


@login_required
def dashboard_seller_add(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if not partner.is_dealer and not partner.is_union_agent:
        return redirect('dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        description = request.POST.get('description', '').strip()
        if not name:
            return render(request, 'store/dashboard_seller_form.html', {
                'error': 'Name is required.',
                'partner': partner,
            })
        seller = Partner.objects.create(
            parent=partner,
            name=name,
            phone=phone,
            description=description,
            is_seller=True,
        )
        if 'profile_image' in request.FILES:
            try:
                validate_upload(request.FILES['profile_image'])
                seller.profile_image = request.FILES['profile_image']
            except ValueError as e:
                return render(request, 'store/dashboard_seller_form.html', {'error': str(e), 'partner': partner})
            seller.save()
        wallet, _ = PartnerWallet.objects.get_or_create(partner=seller)
        wallet.balance += 500
        wallet.locked_balance += 500
        wallet.save()
        WalletTransaction.objects.create(
            wallet=wallet,
            amount=500,
            balance_before=wallet.balance - 500,
            balance_after=wallet.balance,
            type='signup_bonus',
            description='Signup bonus from dealer',
        )
        return redirect('dashboard_sellers')
    return render(request, 'store/dashboard_seller_form.html', {'partner': partner})


@login_required
def dashboard_union_agents(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if not partner.is_dealer:
        return redirect('dashboard')
    union_agents = Partner.objects.filter(parent=partner, is_union_agent=True)
    return render(request, 'store/dashboard_union_agents.html', {
        'partner': partner,
        'union_agents': union_agents,
    })


@login_required
def dashboard_union_agent_add(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if not partner.is_dealer:
        return redirect('dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        description = request.POST.get('description', '').strip()
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password', '')
        if not name:
            return render(request, 'store/dashboard_union_agent_form.html', {
                'error': 'Name is required.',
                'partner': partner,
            })
        if not email:
            return render(request, 'store/dashboard_union_agent_form.html', {
                'error': 'Email is required.',
                'partner': partner,
            })
        if len(password) < 8:
            return render(request, 'store/dashboard_union_agent_form.html', {
                'error': 'Password must be at least 8 characters.',
                'partner': partner,
            })
        if User.objects.filter(username=email).exists():
            return render(request, 'store/dashboard_union_agent_form.html', {
                'error': 'A user with this email already exists.',
                'partner': partner,
            })
        user = User.objects.create_user(username=email, email=email, password=password)
        agent = Partner.objects.create(
            user=user,
            parent=partner,
            name=name,
            phone=phone,
            description=description,
            is_seller=True,
            is_union_agent=True,
        )
        if 'profile_image' in request.FILES:
            try:
                validate_upload(request.FILES['profile_image'])
                agent.profile_image = request.FILES['profile_image']
            except ValueError as e:
                return render(request, 'store/dashboard_union_agent_form.html', {'error': str(e), 'partner': partner})
            agent.save()
        wallet, _ = PartnerWallet.objects.get_or_create(partner=agent)
        wallet.balance += 500
        wallet.locked_balance += 500
        wallet.save()
        WalletTransaction.objects.create(
            wallet=wallet,
            amount=500,
            balance_before=wallet.balance - 500,
            balance_after=wallet.balance,
            type='signup_bonus',
            description='Signup bonus from dealer',
        )
        return redirect('dashboard_union_agents')
    return render(request, 'store/dashboard_union_agent_form.html', {'partner': partner})


@login_required
def dashboard_products(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    q = request.GET.get('q', '').strip()
    products = Product.objects.filter(partner=partner, trashed=False).select_related('category')
    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q) | Q(medicine_generic_name__icontains=q))
    trashed_count = Product.objects.filter(partner=partner, trashed=True).count()
    paginator = Paginator(products, 999999)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'store/dashboard_products.html', {
        'products': page_obj,
        'partner': partner,
        'trashed_count': trashed_count,
        'query': q,
        'page_obj': page_obj,
    })


@login_required
def dashboard_products_export_csv(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    products = Product.objects.filter(partner=partner, trashed=False).select_related('category')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'
    writer = csv.writer(response)
    writer.writerow(['Name', 'SKU', 'Category', 'Price', 'Old Price', 'Cost Price', 'Profit', 'Label', 'Stock', 'Available', 'Created'])
    for p in products:
        writer.writerow([
            p.name,
            p.sku or '',
            p.category.name if p.category else '',
            str(p.price),
            str(p.old_price) if p.old_price else '',
            str(p.cost_price) if p.cost_price is not None else '',
            str(p.profit) if p.profit is not None else '',
            p.label or '',
            p.stock,
            'Yes' if p.available else 'No',
            p.created.strftime('%Y-%m-%d') if p.created else '',
        ])
    return response


@login_required
def dashboard_products_export_pdf(request):
    from weasyprint import HTML
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    products = Product.objects.filter(partner=partner, trashed=False).select_related('category')
    html = render_to_string('store/dashboard_products_pdf.html', {
        'products': products,
        'partner': partner,
    }, request)
    pdf = HTML(string=html).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="products.pdf"'
    return response


def _save_color_variants(request, product):
    color_ids = request.POST.getlist('color_id')
    color_names = request.POST.getlist('color_name')
    color_codes = request.POST.getlist('color_code')
    color_prices = request.POST.getlist('color_price')
    color_stocks = request.POST.getlist('color_stock')
    delete_ids = request.POST.get('color_delete_ids', '')

    # Delete marked variants
    if delete_ids:
        for cid in delete_ids.split(','):
            cid = cid.strip()
            if cid.isdigit():
                product.color_variants.filter(id=int(cid)).delete()

    for i, name in enumerate(color_names):
        name = name.strip()
        if not name:
            continue
        code = color_codes[i].strip() if i < len(color_codes) else '#000000'
        price_adj = color_prices[i].strip() if i < len(color_prices) else '0'
        stock_val = color_stocks[i].strip() if i < len(color_stocks) else '0'
        cid = color_ids[i].strip() if i < len(color_ids) else ''

        if cid.isdigit():
            variant = product.color_variants.filter(id=int(cid)).first()
            if variant:
                variant.color_name = name
                variant.color_code = code
                variant.price_adjustment = float(price_adj) if price_adj else 0
                variant.stock = int(stock_val) if stock_val.isdigit() else 0
                variant.order = i
                color_file_key = f'color_image_{i}'
                if color_file_key in request.FILES:
                    variant.image = request.FILES[color_file_key]
                variant.save()
        else:
            variant = ProductColorVariant.objects.create(
                product=product,
                color_name=name,
                color_code=code,
                price_adjustment=float(price_adj) if price_adj else 0,
                stock=int(stock_val) if stock_val.isdigit() else 0,
                order=i,
            )
            color_file_key = f'color_image_{i}'
            if color_file_key in request.FILES:
                variant.image = request.FILES[color_file_key]
                variant.save()


def _save_size_variants(request, product):
    size_ids = request.POST.getlist('size_id')
    size_names = request.POST.getlist('size_name')
    size_prices = request.POST.getlist('size_price')
    size_stocks = request.POST.getlist('size_stock')
    delete_ids = request.POST.get('size_delete_ids', '')

    if delete_ids:
        for sid in delete_ids.split(','):
            sid = sid.strip()
            if sid.isdigit():
                product.size_variants.filter(id=int(sid)).delete()

    for i, name in enumerate(size_names):
        name = name.strip()
        if not name:
            continue
        price_adj = size_prices[i].strip() if i < len(size_prices) else '0'
        stock_val = size_stocks[i].strip() if i < len(size_stocks) else '0'
        sid = size_ids[i].strip() if i < len(size_ids) else ''

        if sid.isdigit():
            variant = product.size_variants.filter(id=int(sid)).first()
            if variant:
                variant.size_name = name
                variant.price_adjustment = float(price_adj) if price_adj else 0
                variant.stock = int(stock_val) if stock_val.isdigit() else 0
                variant.order = i
                variant.save()
        else:
            ProductSizeVariant.objects.create(
                product=product,
                size_name=name,
                price_adjustment=float(price_adj) if price_adj else 0,
                stock=int(stock_val) if stock_val.isdigit() else 0,
                order=i,
            )


@login_required
def dashboard_product_create(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        sku = request.POST.get('sku', '').strip()
        barcode = request.POST.get('barcode', '').strip()
        price = request.POST.get('price', '').strip()
        stock = request.POST.get('stock', '0').strip()
        low_stock_threshold = request.POST.get('low_stock_threshold', '5').strip()
        short_description = request.POST.get('short_description', '').strip()
        description = request.POST.get('description', '').strip()
        category_id = request.POST.get('category', '').strip()
        old_price = request.POST.get('old_price', '').strip()
        label = request.POST.get('label', '').strip()
        custom_label = request.POST.get('custom_label', '').strip()
        video_url = request.POST.get('video_url', '').strip()
        available = request.POST.get('available') == 'on'

        cost_price = request.POST.get('cost_price', '').strip()
        weight = request.POST.get('weight', '').strip()
        weight_unit = request.POST.get('weight_unit', 'kg').strip()
        length = request.POST.get('length', '').strip()
        width = request.POST.get('width', '').strip()
        height = request.POST.get('height', '').strip()
        dimension_unit = request.POST.get('dimension_unit', 'cm').strip()
        sale_unit = request.POST.get('sale_unit', 'piece').strip()
        custom_unit_label = request.POST.get('custom_unit_label', '').strip()
        price_on_request = request.POST.get('price_on_request') == 'on'
        medicine_brand_name = request.POST.get('medicine_brand_name', '').strip()
        medicine_generic_name = request.POST.get('medicine_generic_name', '').strip()
        medicine_strength = request.POST.get('medicine_strength', '').strip()
        medicine_dosage_form = request.POST.get('medicine_dosage_form', '').strip()

        if not name or not price:
            return render(request, 'store/dashboard_product_form.html', {
                'error': 'Name and price are required.',
                'categories': Category.objects.all(),
                'partner': partner,
                'is_edit': False,
                'preset_colors': PRESET_COLORS,
            })
        product = Product.objects.create(
            name=name,
            sku=sku,
            barcode=barcode,
            partner=partner,
            category_id=int(category_id) if category_id.isdigit() else None,
            price=price,
            cost_price=float(cost_price) if cost_price else None,
            old_price=old_price if old_price else None,
            label=label if label else None,
            custom_label=custom_label or None,
            stock=int(stock) if stock.isdigit() else 0,
            low_stock_threshold=int(low_stock_threshold) if low_stock_threshold.isdigit() else 5,
            short_description=short_description,
            description=description,
            video_url=video_url or None,
            available=available,
            weight=float(weight) if weight else None,
            weight_unit=weight_unit if weight else 'kg',
            length=float(length) if length else None,
            width=float(width) if width else None,
            height=float(height) if height else None,
            dimension_unit=dimension_unit if (length or width or height) else 'cm',
            sale_unit=sale_unit,
            custom_unit_label=custom_unit_label if sale_unit == 'custom' else '',
            price_on_request=price_on_request,
            medicine_brand_name=medicine_brand_name,
            medicine_generic_name=medicine_generic_name,
            medicine_strength=medicine_strength,
            medicine_dosage_form=medicine_dosage_form,
            is_published=True,
        )
        if 'image' in request.FILES:
            try:
                validate_upload(request.FILES['image'])
                product.image = request.FILES['image']
            except ValueError as e:
                product.delete()
                return render(request, 'store/dashboard_product_form.html', {'error': str(e), 'partner': partner, 'categories': Category.objects.all()})
        if 'hover_image' in request.FILES:
            try:
                validate_upload(request.FILES['hover_image'])
                product.hover_image = request.FILES['hover_image']
            except ValueError as e:
                product.delete()
                return render(request, 'store/dashboard_product_form.html', {'error': str(e), 'partner': partner, 'categories': Category.objects.all()})
        product.save()
        _save_color_variants(request, product)
        _save_size_variants(request, product)
        for f in request.FILES.getlist('extra_images'):
            try:
                validate_upload(f)
                ProductImage.objects.create(product=product, image=f)
            except ValueError:
                pass
        return redirect('dashboard_products')
    return render(request, 'store/dashboard_product_form.html', {
        'categories': Category.objects.all(),
        'partner': partner,
        'product': None,
        'is_edit': False,
        'preset_colors': PRESET_COLORS,
        'preset_sizes': PRESET_SIZES,
    })


@login_required
def dashboard_product_edit(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    product = get_object_or_404(Product, pk=pk, partner=partner)
    if request.method == 'POST':
        product.name = request.POST.get('name', product.name).strip()
        product.sku = request.POST.get('sku', product.sku).strip()
        product.barcode = request.POST.get('barcode', product.barcode).strip()
        product.low_stock_threshold = int(request.POST.get('low_stock_threshold', product.low_stock_threshold))
        price = request.POST.get('price', '').strip()
        stock = request.POST.get('stock', '').strip()
        category_id = request.POST.get('category', '').strip()
        old_price = request.POST.get('old_price', '').strip()
        label = request.POST.get('label', '').strip()
        product.custom_label = request.POST.get('custom_label', '').strip() or None
        product.short_description = request.POST.get('short_description', '').strip()
        product.description = request.POST.get('description', product.description).strip()
        product.video_url = request.POST.get('video_url', '').strip() or None
        product.available = request.POST.get('available') == 'on'

        cost_price = request.POST.get('cost_price', '').strip()
        product.cost_price = float(cost_price) if cost_price else None

        weight = request.POST.get('weight', '').strip()
        product.weight = float(weight) if weight else None
        product.weight_unit = request.POST.get('weight_unit', 'kg').strip()
        length = request.POST.get('length', '').strip()
        width = request.POST.get('width', '').strip()
        height = request.POST.get('height', '').strip()
        product.length = float(length) if length else None
        product.width = float(width) if width else None
        product.height = float(height) if height else None
        product.dimension_unit = request.POST.get('dimension_unit', 'cm').strip()
        product.sale_unit = request.POST.get('sale_unit', 'piece').strip()
        product.custom_unit_label = request.POST.get('custom_unit_label', '').strip() if product.sale_unit == 'custom' else ''
        product.price_on_request = request.POST.get('price_on_request') == 'on'
        product.medicine_brand_name = request.POST.get('medicine_brand_name', '').strip()
        product.medicine_generic_name = request.POST.get('medicine_generic_name', '').strip()
        product.medicine_strength = request.POST.get('medicine_strength', '').strip()
        product.medicine_dosage_form = request.POST.get('medicine_dosage_form', '').strip()

        if price:
            product.price = price
        if stock:
            product.stock = int(stock) if stock.isdigit() else product.stock
        product.category_id = int(category_id) if category_id.isdigit() else product.category_id
        product.old_price = old_price if old_price else None
        product.label = label if label else None
        if 'image' in request.FILES:
            try:
                validate_upload(request.FILES['image'])
                product.image = request.FILES['image']
            except ValueError as e:
                messages.error(request, str(e))
                return redirect('dashboard_product_edit', pk=product.pk)
        if 'hover_image' in request.FILES:
            try:
                validate_upload(request.FILES['hover_image'])
                product.hover_image = request.FILES['hover_image']
            except ValueError as e:
                messages.error(request, str(e))
                return redirect('dashboard_product_edit', pk=product.pk)
        product.save()
        _save_color_variants(request, product)
        _save_size_variants(request, product)

        # Handle extra images deletion and upload
        extra_delete_ids = request.POST.get('extra_delete_ids', '')
        for eid in extra_delete_ids.split(','):
            eid = eid.strip()
            if eid.isdigit():
                product.extra_images.filter(id=int(eid)).delete()
        for f in request.FILES.getlist('extra_images'):
            try:
                validate_upload(f)
                ProductImage.objects.create(product=product, image=f)
            except ValueError:
                pass

        return redirect('dashboard_products')
    return render(request, 'store/dashboard_product_form.html', {
        'product': product,
        'categories': Category.objects.all(),
        'partner': partner,
        'is_edit': True,
        'color_variants': product.color_variants.all(),
        'size_variants': product.size_variants.all(),
        'preset_colors': PRESET_COLORS,
        'preset_sizes': PRESET_SIZES,
    })


@login_required
def dashboard_product_delete(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    product = get_object_or_404(Product, pk=pk, partner=partner)
    if request.method == 'POST':
        product.trashed = True
        product.trashed_at = timezone.now()
        product.save(update_fields=['trashed', 'trashed_at'])
        messages.success(request, f'"{product.name}" moved to trash.')
        return redirect('dashboard_trash_list')
    return render(request, 'store/dashboard_product_confirm_delete.html', {'product': product})


@login_required
def dashboard_trash_list(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    now = timezone.now()
    # Auto-purge expired trash (older than 7 days)
    cutoff = now - timedelta(days=7)
    expired = Product.objects.filter(partner=partner, trashed=True, trashed_at__lt=cutoff)
    purged_count = expired.count()
    if purged_count:
        expired.delete()
    products = Product.objects.filter(partner=partner, trashed=True).order_by('-trashed_at')
    TRASH_DAYS = 7
    for p in products:
        if p.trashed_at:
            elapsed = (now - p.trashed_at).days
            p.days_remaining = max(TRASH_DAYS - elapsed, 0)
            p.is_expired = elapsed >= TRASH_DAYS
        else:
            p.days_remaining = TRASH_DAYS
            p.is_expired = False
    return render(request, 'store/dashboard_product_trash.html', {
        'partner': partner,
        'products': products,
        'purged_count': purged_count,
        'trash_days': TRASH_DAYS,
    })


@login_required
def dashboard_trash_restore(request):
    if request.method != 'POST':
        return redirect('dashboard_trash_list')
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    product_ids = request.POST.getlist('product_ids')
    if not product_ids:
        messages.error(request, 'No products selected.')
    else:
        count = Product.objects.filter(partner=partner, id__in=product_ids, trashed=True).update(trashed=False, trashed_at=None)
        messages.success(request, f'{count} product(s) restored from trash.')
    return redirect('dashboard_trash_list')


@login_required
def dashboard_trash_restore_all(request):
    if request.method != 'POST':
        return redirect('dashboard_trash_list')
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    count = Product.objects.filter(partner=partner, trashed=True).update(trashed=False, trashed_at=None)
    messages.success(request, f'All {count} product(s) restored from trash.')
    return redirect('dashboard_trash_list')


@login_required
def dashboard_trash_empty(request):
    if request.method != 'POST':
        return redirect('dashboard_trash_list')
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    product_ids = request.POST.getlist('product_ids')
    if product_ids:
        count = Product.objects.filter(partner=partner, id__in=product_ids, trashed=True).delete()[0]
        messages.success(request, f'{count} product(s) permanently deleted.')
    else:
        count = Product.objects.filter(partner=partner, trashed=True).delete()[0]
        messages.success(request, f'Trash emptied — {count} product(s) permanently deleted.')
    return redirect('dashboard_trash_list')


@login_required
def dashboard_product_bulk_delete(request):
    if request.method != 'POST':
        return redirect('dashboard_products')
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    product_ids = request.POST.getlist('product_ids')
    if not product_ids:
        messages.error(request, 'No products selected.')
        return redirect('dashboard_products')
    now = timezone.now()
    count = Product.objects.filter(partner=partner, id__in=product_ids, trashed=False).update(trashed=True, trashed_at=now)
    messages.success(request, f'{count} product(s) moved to trash.')
    return redirect('dashboard_products')


@login_required
def dashboard_product_duplicate(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    original = get_object_or_404(Product, pk=pk, partner=partner)

    duplicate = Product(
        name=original.name,
        slug='',
        sku='',
        category=original.category,
        partner=original.partner,
        image=original.image,
        hover_image=original.hover_image,
        price=original.price,
        old_price=original.old_price,
        label=original.label,
        custom_label=original.custom_label,
        short_description=original.short_description,
        description=original.description,
        stock=0,
        available=True,
        video_url=original.video_url,
        weight=original.weight,
        weight_unit=original.weight_unit,
        length=original.length,
        width=original.width,
        height=original.height,
        dimension_unit=original.dimension_unit,
        sale_unit=original.sale_unit,
        custom_unit_label=original.custom_unit_label,
        price_on_request=original.price_on_request,
        cost_price=original.cost_price,
        meta_title=original.meta_title,
        meta_description=original.meta_description,
        og_image=original.og_image,
    )
    duplicate.save()

    for img in original.extra_images.all():
        ProductImage.objects.create(product=duplicate, image=img.image, order=img.order)
    for cv in original.color_variants.all():
        ProductColorVariant.objects.create(
            product=duplicate, color_name=cv.color_name, color_code=cv.color_code,
            image=cv.image, price_adjustment=cv.price_adjustment, order=cv.order,
        )
    for sv in original.size_variants.all():
        ProductSizeVariant.objects.create(
            product=duplicate, size_name=sv.size_name,
            price_adjustment=sv.price_adjustment, order=sv.order,
        )

    messages.success(request, f'"{original.name}" duplicated successfully.')
    return redirect('dashboard_product_edit', pk=duplicate.pk)


@login_required
def dashboard_profile_edit(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        description = request.POST.get('description', '').strip()
        shop_style = request.POST.get('shop_style', '').strip()
        theme = request.POST.get('theme', '').strip()
        if name:
            partner.name = name
        if phone:
            partner.phone = phone
        if description:
            partner.description = description
        if shop_style:
            partner.shop_style = shop_style
        partner.theme = theme if theme else ''
        if 'logo' in request.FILES:
            try:
                validate_upload(request.FILES['logo'])
                partner.logo = request.FILES['logo']
            except ValueError as e:
                messages.error(request, str(e))
        if request.POST.get('delete_logo'):
            partner.logo.delete()
            partner.logo = None
        if 'profile_image' in request.FILES:
            try:
                validate_upload(request.FILES['profile_image'])
                partner.profile_image = request.FILES['profile_image']
            except ValueError as e:
                messages.error(request, str(e))
        if request.POST.get('delete_profile_image'):
            partner.profile_image.delete()
            partner.profile_image = None
        if 'banner' in request.FILES:
            try:
                validate_upload(request.FILES['banner'])
                partner.banner = request.FILES['banner']
            except ValueError as e:
                messages.error(request, str(e))
        if request.POST.get('delete_banner'):
            partner.banner.delete()
            partner.banner = None
        for f in request.FILES.getlist('new_banners'):
            if f:
                try:
                    validate_upload(f)
                    PartnerBanner.objects.create(partner=partner, image=f)
                except ValueError:
                    pass
        for bid in request.POST.getlist('delete_extra_banner'):
            PartnerBanner.objects.filter(id=bid, partner=partner).delete()
        partner.save()
        return redirect('dashboard')
    extra_banners = partner.extra_banners.all()
    return render(request, 'store/dashboard_profile_edit.html', {
        'partner': partner,
        'extra_banners': extra_banners,
        'THEME_CHOICES': THEME_CHOICES,
    })


@login_required
def dashboard_orders(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    orders = Order.objects.filter(items__product__partner=partner).distinct().prefetch_related('items').order_by('-created')

    # Annotate each order with the partner's revenue, cost, and profit
    for order in orders:
        partner_items = [item for item in order.items.all() if item.product and item.product.partner == partner]
        order_revenue = sum(item.price * item.quantity for item in partner_items)
        order_cost = sum((item.cost_price or 0) * item.quantity for item in partner_items)
        order.partner_revenue = order_revenue
        order.partner_cost = order_cost
        order.partner_profit = order_revenue - order_cost

    return render(request, 'store/dashboard_orders.html', {
        'orders': orders,
        'partner': partner,
    })


def add_review(request, product_id):
    product = get_object_or_404(Product.objects.filter(trashed=False), id=product_id)
    if request.method == 'POST':
        rating = request.POST.get('rating')
        comment = request.POST.get('comment', '').strip()
        if not rating or not comment:
            return redirect('product_detail', slug=product.slug)
        if request.user.is_authenticated:
            user = request.user
        else:
            return redirect('login')
        has_purchased = Order.objects.filter(user=user, items__product=product, status__in=['completed', 'processing', 'pending']).exists()
        if not has_purchased and not user.is_superuser:
            return redirect('product_detail', slug=product.slug)
        image = request.FILES.get('image')
        ProductReview.objects.create(
            product=product,
            user=user,
            rating=int(rating),
            comment=comment,
            image=image,
        )
    return redirect('product_detail', slug=product.slug)


@login_required
def dashboard_addresses(request):
    addresses = request.user.addresses.all()
    return render(request, 'store/dashboard_addresses.html', {'addresses': addresses})


@login_required
def dashboard_address_add(request):
    if request.method == 'POST':
        label = request.POST.get('label', '').strip()
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        street = request.POST.get('street_address', '').strip()
        area = request.POST.get('area', '').strip()
        district = request.POST.get('district', '').strip()
        division = request.POST.get('division', '').strip()
        zip_code = request.POST.get('zip_code', '').strip()
        is_default = request.POST.get('is_default') == 'on'

        if is_default:
            request.user.addresses.update(is_default=False)

        Address.objects.create(
            user=request.user,
            label=label or None,
            name=name,
            phone=phone,
            street_address=street,
            area=area,
            district=district,
            division=division,
            zip_code=zip_code,
            is_default=is_default,
        )
        messages.success(request, 'Address added successfully.')
        return redirect('dashboard_addresses')
    return render(request, 'store/dashboard_address_form.html', {'address': None})


@login_required
def dashboard_address_edit(request, pk):
    address = get_object_or_404(Address, pk=pk, user=request.user)
    if request.method == 'POST':
        address.label = request.POST.get('label', '').strip() or None
        address.name = request.POST.get('name', '').strip()
        address.phone = request.POST.get('phone', '').strip()
        address.street_address = request.POST.get('street_address', '').strip()
        address.area = request.POST.get('area', '').strip()
        address.district = request.POST.get('district', '').strip()
        address.division = request.POST.get('division', '').strip()
        address.zip_code = request.POST.get('zip_code', '').strip()
        is_default = request.POST.get('is_default') == 'on'
        if is_default:
            request.user.addresses.exclude(pk=address.pk).update(is_default=False)
        address.is_default = is_default
        address.save()
        messages.success(request, 'Address updated successfully.')
        return redirect('dashboard_addresses')
    return render(request, 'store/dashboard_address_form.html', {'address': address})


@login_required
def dashboard_address_delete(request, pk):
    address = get_object_or_404(Address, pk=pk, user=request.user)
    address.delete()
    messages.success(request, 'Address deleted.')
    return redirect('dashboard_addresses')


@login_required
def dashboard_address_set_default(request, pk):
    address = get_object_or_404(Address, pk=pk, user=request.user)
    request.user.addresses.update(is_default=False)
    address.is_default = True
    address.save(update_fields=['is_default'])
    messages.success(request, f'"{address}" is now your default address.')
    return redirect('dashboard_addresses')


@login_required
def edit_review(request, review_id):
    review = get_object_or_404(ProductReview, id=review_id)
    user = request.user
    is_partner = hasattr(user, 'partner') and review.product.partner == user.partner
    if not (user.is_superuser or is_partner):
        return redirect('product_detail', slug=review.product.slug)
    if request.method == 'POST':
        rating = request.POST.get('rating')
        comment = request.POST.get('comment', '').strip()
        if rating and comment:
            review.rating = int(rating)
            review.comment = comment
            review.save()
            return redirect('product_detail', slug=review.product.slug)
    return render(request, 'store/edit_review.html', {'review': review})


@login_required
def delete_review(request, review_id):
    review = get_object_or_404(ProductReview, id=review_id)
    user = request.user
    is_partner = hasattr(user, 'partner') and review.product.partner == user.partner
    if not (user.is_superuser or is_partner):
        return redirect('product_detail', slug=review.product.slug)
    if request.method == 'POST':
        review.delete()
    return redirect('product_detail', slug=review.product.slug)


@login_required
def dashboard_reviews(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    product_ids = Product.objects.filter(partner=partner, trashed=False).values_list('id', flat=True)
    reviews = ProductReview.objects.filter(product_id__in=product_ids).select_related('product', 'user').order_by('-created')
    return render(request, 'store/dashboard_reviews.html', {'reviews': reviews, 'partner': partner})


@login_required
def dashboard_menu_list(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    items = PartnerNavMenu.objects.filter(partner=partner, parent=None).prefetch_related('children').order_by('order')
    return render(request, 'store/dashboard_partner_menu.html', {'items': items, 'partner': partner})


@login_required
def dashboard_menu_add(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    parents = PartnerNavMenu.objects.filter(partner=partner, parent=None)
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        url = request.POST.get('url', '').strip()
        url_type = request.POST.get('url_type', 'named_url')
        parent_id = request.POST.get('parent')
        order = request.POST.get('order', 0)
        is_active = request.POST.get('is_active') == 'on'
        parent = PartnerNavMenu.objects.filter(id=parent_id, partner=partner).first() if parent_id else None
        if not title:
            return render(request, 'store/dashboard_partner_menu_form.html', {
                'partner': partner, 'parents': parents, 'error': 'Title is required.'
            })
        PartnerNavMenu.objects.create(
            partner=partner, title=title, url=url, url_type=url_type,
            parent=parent, order=order, is_active=is_active,
        )
        return redirect('dashboard_menu_list')
    return render(request, 'store/dashboard_partner_menu_form.html', {'partner': partner, 'parents': parents})


@login_required
def dashboard_menu_edit(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    item = get_object_or_404(PartnerNavMenu, id=pk, partner=partner)
    parents = PartnerNavMenu.objects.filter(partner=partner, parent=None).exclude(id=pk)
    if request.method == 'POST':
        item.title = request.POST.get('title', '').strip()
        item.url = request.POST.get('url', '').strip()
        item.url_type = request.POST.get('url_type', 'named_url')
        parent_id = request.POST.get('parent')
        item.parent = PartnerNavMenu.objects.filter(id=parent_id, partner=partner).first() if parent_id else None
        item.order = request.POST.get('order', 0)
        item.is_active = request.POST.get('is_active') == 'on'
        if not item.title:
            return render(request, 'store/dashboard_partner_menu_form.html', {
                'partner': partner, 'item': item, 'parents': parents, 'error': 'Title is required.'
            })
        item.save()
        return redirect('dashboard_menu_list')
    return render(request, 'store/dashboard_partner_menu_form.html', {'partner': partner, 'item': item, 'parents': parents})


@login_required
def dashboard_menu_delete(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    item = get_object_or_404(PartnerNavMenu, id=pk, partner=partner)
    if request.method == 'POST':
        item.delete()
        return redirect('dashboard_menu_list')
    return render(request, 'store/dashboard_partner_menu_confirm_delete.html', {'partner': partner, 'item': item})


@login_required
def dashboard_side_banners(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    banners = SideBanner.objects.filter(partner=partner).order_by('order')
    return render(request, 'store/dashboard_side_banners.html', {'banners': banners, 'partner': partner})


@login_required
def dashboard_side_banner_add(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        link_url = request.POST.get('link_url', '').strip()
        image = request.FILES.get('image')
        if image:
            SideBanner.objects.create(partner=partner, title=title, image=image, link_url=link_url)
        return redirect('dashboard_side_banners')
    return render(request, 'store/dashboard_side_banner_form.html', {'partner': partner})


@login_required
def dashboard_side_banner_edit(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    banner = get_object_or_404(SideBanner, id=pk, partner=partner)
    if request.method == 'POST':
        banner.title = request.POST.get('title', '').strip()
        banner.link_url = request.POST.get('link_url', '').strip()
        if 'image' in request.FILES:
            banner.image = request.FILES['image']
        if request.POST.get('delete_image'):
            banner.image.delete()
            banner.image = None
        banner.save()
        return redirect('dashboard_side_banners')
    return render(request, 'store/dashboard_side_banner_form.html', {'banner': banner, 'partner': partner})


@login_required
def dashboard_slider_list(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    sliders = PartnerSlider.objects.filter(partner=partner).order_by('order')
    return render(request, 'store/dashboard_sliders.html', {'sliders': sliders, 'partner': partner})


@login_required
def dashboard_slider_add(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        subtitle = request.POST.get('subtitle', '').strip()
        link_url = request.POST.get('link_url', '').strip()
        button_text = request.POST.get('button_text', '').strip()
        max_height = request.POST.get('max_height', '400').strip()
        order = request.POST.get('order', '0').strip()
        image = request.FILES.get('image')
        if image:
            PartnerSlider.objects.create(
                partner=partner, title=title, subtitle=subtitle,
                image=image, link_url=link_url,
                button_text=button_text or 'Click for Partner',
                max_height=min(int(max_height) if max_height.isdigit() else 400, 500),
                order=int(order) if order.isdigit() else 0,
            )
        return redirect('dashboard_slider_list')
    return render(request, 'store/dashboard_slider_form.html', {'partner': partner})


@login_required
def dashboard_slider_edit(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    slider = get_object_or_404(PartnerSlider, id=pk, partner=partner)
    if request.method == 'POST':
        slider.title = request.POST.get('title', '').strip()
        slider.subtitle = request.POST.get('subtitle', '').strip()
        slider.link_url = request.POST.get('link_url', '').strip()
        slider.button_text = request.POST.get('button_text', '').strip() or 'Click for Partner'
        max_height = request.POST.get('max_height', '400').strip()
        slider.max_height = min(int(max_height) if max_height.isdigit() else 400, 500)
        order = request.POST.get('order', '0').strip()
        slider.order = int(order) if order.isdigit() else slider.order
        if 'image' in request.FILES:
            slider.image = request.FILES['image']
        if request.POST.get('delete_image'):
            slider.image.delete()
        slider.save()
        return redirect('dashboard_slider_list')
    return render(request, 'store/dashboard_slider_form.html', {'slider': slider, 'partner': partner})


@login_required
def dashboard_slider_delete(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    slider = get_object_or_404(PartnerSlider, id=pk, partner=partner)
    if request.method == 'POST':
        slider.delete()
        return redirect('dashboard_slider_list')
    return render(request, 'store/dashboard_slider_confirm_delete.html', {'slider': slider, 'partner': partner})


@login_required
def dashboard_side_banner_delete(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    banner = get_object_or_404(SideBanner, id=pk, partner=partner)
    if request.method == 'POST':
        banner.delete()
        return redirect('dashboard_side_banners')
    return render(request, 'store/dashboard_side_banner_confirm_delete.html', {'banner': banner, 'partner': partner})


def ratelimit_error(request):
    return render(request, '404.html', {'message': 'Too many requests. Please slow down.'}, status=429)


def payment_init(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    manual_methods = ManualPaymentMethod.objects.filter(is_active=True)
    return render(request, 'store/payment_init.html', {'order': order, 'manual_methods': manual_methods})


@login_required
def payment_process(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    payment_method = request.POST.get('payment_method', 'cod')

    if payment_method == 'cod':
        order.payment_method = 'Cash on Delivery'
        order.payment_status = 'unpaid'
        order.save(update_fields=['payment_method', 'payment_status'])
        return redirect('order_confirmation', order_id=order.id)

    if payment_method == 'sslcommerz':
        if not settings.SSLCOMMERZ_STORE_ID:
            order.payment_method = 'Cash on Delivery'
            order.save(update_fields=['payment_method'])
            return redirect('order_confirmation', order_id=order.id)

        from .payment import sslcommerz_init
        gateway_url = sslcommerz_init(order, request)
        if gateway_url:
            order.payment_method = 'Online'
            order.save(update_fields=['payment_method'])
            return redirect(gateway_url)

        return render(request, 'store/payment_failed.html', {'order': order, 'error': 'Payment gateway unavailable. Please try Cash on Delivery.'})

    # Manual payment methods (bKash, Nagad, Rocket, etc.)
    manual_method = ManualPaymentMethod.objects.filter(name__iexact=payment_method, is_active=True).first()
    if manual_method:
        transaction_id = request.POST.get('transaction_id', '').strip()
        if not transaction_id:
            manual_methods = ManualPaymentMethod.objects.filter(is_active=True)
            return render(request, 'store/payment_init.html', {
                'order': order,
                'manual_methods': manual_methods,
                'error': 'Please enter your Transaction ID.',
                'selected_method': payment_method,
                'transaction_id': transaction_id,
            })
        if not re.match(r'^[A-Za-z0-9]+$', transaction_id):
            manual_methods = ManualPaymentMethod.objects.filter(is_active=True)
            return render(request, 'store/payment_init.html', {
                'order': order,
                'manual_methods': manual_methods,
                'error': 'Transaction ID must contain only letters and numbers.',
                'selected_method': payment_method,
                'transaction_id': transaction_id,
            })
        order.payment_method = manual_method.name
        order.tracking_id = transaction_id[:100]
        order.payment_status = 'unpaid'
        order.save(update_fields=['payment_method', 'tracking_id', 'payment_status'])
        return redirect('payment_manual', order_id=order.id)

    return redirect('order_confirmation', order_id=order.id)


@login_required
def payment_success(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)

    if order.payment_status != 'paid':
        from .payment import sslcommerz_validate
        if sslcommerz_validate(order):
            order.payment_status = 'paid'
            order.status = 'processing'
            order.save(update_fields=['payment_status', 'status'])

    return render(request, 'store/payment_success.html', {'order': order})


def payment_fail(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'store/payment_failed.html', {'order': order})


def payment_cancel(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'store/payment_failed.html', {'order': order})


def payment_manual(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    method = ManualPaymentMethod.objects.filter(name__iexact=order.payment_method).first()
    return render(request, 'store/payment_manual.html', {
        'order': order,
        'method': method,
    })


@login_required
def dashboard_wallet(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    wallet, _ = PartnerWallet.objects.get_or_create(partner=partner)
    transactions = wallet.transactions.all()[:20]
    pending_withdrawals = WithdrawalRequest.objects.filter(partner=partner, status='pending').count()
    ws = WalletSettings.objects.first()
    return render(request, 'store/dashboard_wallet.html', {
        'partner': partner,
        'wallet': wallet,
        'transactions': transactions,
        'pending_withdrawals': pending_withdrawals,
        'wallet_settings': ws,
    })


@login_required
def dashboard_wallet_transactions(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    wallet, _ = PartnerWallet.objects.get_or_create(partner=partner)
    txn_type = request.GET.get('type', '')
    txns = wallet.transactions.all()
    if txn_type:
        txns = txns.filter(type=txn_type)
    paginator = Paginator(txns, 30)
    page = request.GET.get('page', 1)
    transactions = paginator.get_page(page)
    return render(request, 'store/dashboard_wallet_transactions.html', {
        'partner': partner,
        'wallet': wallet,
        'transactions': transactions,
        'current_type': txn_type,
    })


@login_required
def dashboard_withdraw(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    wallet, _ = PartnerWallet.objects.get_or_create(partner=partner)
    ws = WalletSettings.objects.first()
    payout_methods = PayoutMethod.objects.filter(partner=partner, is_active=True)
    error = ''
    success = ''

    if wallet.is_frozen:
        error = 'Your wallet is currently frozen. Withdrawals are disabled. Please contact support.'

    if request.method == 'POST' and not error:
        amount = request.POST.get('amount', '').strip()
        method = request.POST.get('method', '').strip()
        account_number = request.POST.get('account_number', '').strip()
        account_holder = request.POST.get('account_holder', '').strip()
        payout_id = request.POST.get('payout_method', '').strip()

        if not amount:
            error = 'Amount is required.'
        elif not amount.replace('.', '', 1).isdigit():
            error = 'Invalid amount.'
        else:
            amount = Decimal(amount)
            if amount <= 0:
                error = 'Amount must be greater than zero.'
            elif ws and ws.min_withdrawal and amount < ws.min_withdrawal:
                error = f'Minimum withdrawal amount is ৳{ws.min_withdrawal}.'
            elif ws and ws.max_withdrawal and amount > ws.max_withdrawal:
                error = f'Maximum withdrawal amount is ৳{ws.max_withdrawal}.'
            elif wallet.available_balance < amount:
                error = f'Insufficient wallet balance. Available: ৳{wallet.available_balance} (৳{wallet.locked_balance} is locked as bonus).'
            else:
                pending_count = WithdrawalRequest.objects.filter(partner=partner, status='pending').count()
                if ws and ws.max_pending_withdrawals and pending_count >= ws.max_pending_withdrawals:
                    error = f'You already have {pending_count} pending withdrawal request(s). Maximum {ws.max_pending_withdrawals} allowed at a time.'
                else:
                    if payout_id and payout_id.isdigit():
                        pm = PayoutMethod.objects.filter(id=int(payout_id), partner=partner).first()
                        if pm:
                            method = pm.method_type
                            account_number = pm.account_number
                            account_holder = pm.account_holder

                    if not method:
                        error = 'Payment method is required.'
                    elif not account_number:
                        error = 'Account number is required.'
                    else:
                        fee = ws.calculate_fee(amount) if ws else 0
                        net_amount = amount - fee
                        withdrawal = WithdrawalRequest.objects.create(
                            partner=partner,
                            amount=amount,
                            fee=fee,
                            net_amount=net_amount,
                            method=method,
                            account_number=account_number,
                            account_holder=account_holder,
                        )
                        success = f'Withdrawal request of ৳{amount} submitted successfully. Reference: #{withdrawal.id}'

    return render(request, 'store/dashboard_withdraw.html', {
        'partner': partner,
        'wallet': wallet,
        'wallet_settings': ws,
        'payout_methods': payout_methods,
        'error': error,
        'success': success,
    })


@login_required
def dashboard_withdrawal_list(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    withdrawals = WithdrawalRequest.objects.filter(partner=partner).order_by('-requested_at')
    return render(request, 'store/dashboard_withdrawal_list.html', {
        'partner': partner,
        'withdrawals': withdrawals,
    })


@login_required
def dashboard_payout_methods(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    methods = PayoutMethod.objects.filter(partner=partner)
    error = ''
    success = ''

    if request.method == 'POST':
        method_type = request.POST.get('method_type', '').strip()
        account_number = request.POST.get('account_number', '').strip()
        account_holder = request.POST.get('account_holder', '').strip()
        bank_name = request.POST.get('bank_name', '').strip()
        branch = request.POST.get('branch', '').strip()
        routing_number = request.POST.get('routing_number', '').strip()
        is_default = request.POST.get('is_default') == 'on'

        if not method_type:
            error = 'Method type is required.'
        elif not account_number:
            error = 'Account number is required.'
        else:
            PayoutMethod.objects.create(
                partner=partner,
                method_type=method_type,
                account_number=account_number,
                account_holder=account_holder,
                bank_name=bank_name if method_type == 'bank' else '',
                branch=branch if method_type == 'bank' else '',
                routing_number=routing_number if method_type == 'bank' else '',
                is_default=is_default,
            )
            success = 'Payout method added successfully.'
            methods = PayoutMethod.objects.filter(partner=partner)

    return render(request, 'store/dashboard_payout_methods.html', {
        'partner': partner,
        'methods': methods,
        'error': error,
        'success': success,
    })


@login_required
def dashboard_payout_method_delete(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    method = get_object_or_404(PayoutMethod, id=pk, partner=partner)
    method.delete()
    return redirect('dashboard_payout_methods')


# ════════════════════════════════════════════
# Feature 1: Wishlist
# ════════════════════════════════════════════

def wishlist_view(request):
    wl = Wishlist.get_or_create_for(request)
    return render(request, 'store/wishlist.html', {'wishlist': wl, 'items': wl.items.select_related('product')})


def wishlist_add(request, product_id):
    product = get_object_or_404(Product.objects.filter(trashed=False), id=product_id)
    wl = Wishlist.get_or_create_for(request)
    _, created = WishlistItem.objects.get_or_create(wishlist=wl, product=product)
    if created:
        messages.success(request, f'"{product.name}" added to your wishlist.')
    else:
        messages.info(request, f'"{product.name}" is already in your wishlist.')
    return redirect(request.META.get('HTTP_REFERER', 'wishlist'))


def wishlist_remove(request, item_id):
    if request.user.is_authenticated:
        wl = Wishlist.objects.filter(user=request.user).first()
    else:
        wl = Wishlist.objects.filter(session_id=request.session.session_key, user__isnull=True).first()
    if wl:
        wl.items.filter(id=item_id).delete()
    return redirect('wishlist')


# ════════════════════════════════════════════
# Feature 3: Compare Products
# ════════════════════════════════════════════

def compare_view(request):
    ids = request.session.get('compare', [])
    products = Product.objects.filter(id__in=ids, available=True, trashed=False, is_published=True).filter(Q(partner__isnull=True) | Q(partner__show_products=True)) if ids else []
    return render(request, 'store/compare.html', {'products': products})


def compare_add(request, product_id):
    compare = request.session.get('compare', [])
    if product_id in compare:
        messages.info(request, 'This product is already in your comparison list.')
    elif len(compare) >= 4:
        messages.warning(request, 'Comparison list is full (max 4 products).')
    else:
        compare.append(product_id)
        request.session['compare'] = compare
        messages.success(request, 'Product added to comparison list.')
    return redirect(request.META.get('HTTP_REFERER', 'compare'))


def compare_remove(request, product_id):
    compare = request.session.get('compare', [])
    if product_id in compare:
        compare.remove(product_id)
        request.session['compare'] = compare
    return redirect('compare')


def compare_clear(request):
    request.session['compare'] = []
    return redirect('compare')


# ════════════════════════════════════════════
# Feature 4: Notifications
# ════════════════════════════════════════════

@login_required
def dashboard_notifications(request):
    notifs = Notification.objects.filter(
        Q(user=request.user) | Q(partner__user=request.user)
    )[:50]
    return render(request, 'store/dashboard_notifications.html', {'notifications': notifs})


@login_required
def notification_read(request, pk):
    notif = get_object_or_404(Notification, pk=pk)
    if notif.user == request.user or (notif.partner and notif.partner.user == request.user):
        notif.is_read = True
        notif.save(update_fields=['is_read'])
    return redirect(notif.link or 'dashboard_notifications')


@login_required
def notification_read_all(request):
    Notification.objects.filter(
        Q(user=request.user) | Q(partner__user=request.user),
        is_read=False,
    ).update(is_read=True)
    return redirect('dashboard_notifications')


# ════════════════════════════════════════════
# Feature 6: Refunds
# ════════════════════════════════════════════

@login_required
def request_refund(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    if order.status not in ('completed',):
        return redirect('dashboard')
    existing = RefundRequest.objects.filter(order=order, user=request.user, status='pending').first()
    if existing:
        return redirect('dashboard')
    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        amount = request.POST.get('amount', '').strip()
        if not reason or not amount:
            return render(request, 'store/request_refund.html', {'order': order, 'error': 'All fields required.'})
        amount = Decimal(amount)
        if amount <= 0 or amount > order.total:
            return render(request, 'store/request_refund.html', {'order': order, 'error': 'Invalid amount.'})
        partner = None
        first_item = order.items.filter(product__partner__isnull=False).first()
        if first_item and first_item.product:
            partner = first_item.product.partner
        RefundRequest.objects.create(order=order, user=request.user, partner=partner, amount=amount, reason=reason)
        return render(request, 'store/request_refund.html', {'order': order, 'success': True})
    return render(request, 'store/request_refund.html', {'order': order})


@login_required
def dashboard_refunds(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    refunds = RefundRequest.objects.filter(partner=partner).order_by('-created')
    return render(request, 'store/dashboard_refunds.html', {'partner': partner, 'refunds': refunds})


# ════════════════════════════════════════════
# Feature 7: Seller Analytics
# ════════════════════════════════════════════

@login_required
def dashboard_analytics(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    return render(request, 'store/dashboard_analytics.html', {'partner': partner})


@login_required
def dashboard_analytics_data(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return JsonResponse({'error': 'Not a partner'}, status=403)

    from django.db.models.functions import TruncDate

    days = int(request.GET.get('days', 30))
    since = timezone.now() - timezone.timedelta(days=days)

    # Sales over time
    items = OrderItem.objects.filter(
        product__partner=partner,
        order__created__gte=since,
        order__status__in=('completed', 'processing'),
    )
    daily_sales = list(items.annotate(
        date=TruncDate('order__created')
    ).values('date').annotate(
        revenue=Sum(F('price') * F('quantity')),
        cost=Sum(F('cost_price') * F('quantity')),
    ).order_by('date'))

    sales_labels = []
    sales_revenue = []
    sales_profit = []
    for entry in daily_sales:
        sales_labels.append(entry['date'].strftime('%d %b'))
        rev = float(entry['revenue'] or 0)
        cst = float(entry['cost'] or 0)
        sales_revenue.append(rev)
        sales_profit.append(rev - cst)

    # Top products
    top = items.values('product__name', 'product__id').annotate(
        total_qty=Sum('quantity'),
        total_rev=Sum(F('price') * F('quantity')),
    ).order_by('-total_rev')[:10]

    # Status counts
    orders = Order.objects.filter(items__product__partner=partner).distinct()
    status_counts = {}
    for s, label in Order.STATUS_CHOICES:
        cnt = orders.filter(status=s).count()
        if cnt:
            status_counts[label] = cnt

    return JsonResponse({
        'sales_labels': sales_labels,
        'sales_revenue': sales_revenue,
        'sales_profit': sales_profit,
        'top_products': [
            {'name': p['product__name'], 'revenue': float(p['total_rev'] or 0), 'qty': p['total_qty']}
            for p in top
        ],
        'status_counts': status_counts,
        'total_revenue': sum(sales_revenue),
        'total_cost': sum(float(e['cost'] or 0) for e in daily_sales),
    })


# ════════════════════════════════════════════
# Feature 8: Bulk Product Upload
# ════════════════════════════════════════════

@login_required
def dashboard_product_bulk_upload(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')

    import io
    import openpyxl

    MEDICINE_COLUMNS = {'brand name', 'brand_name', 'dosage form', 'dosage_form', 'generic', 'strength', 'package container', 'package_container'}

    results = {'created': 0, 'errors': [], 'total': 0}
    is_medicine = False

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        ext = file.name.rsplit('.', 1)[-1].lower()
        try:
            if ext == 'csv':
                data = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(data))
                rows = list(reader)
            elif ext in ('xlsx', 'xls'):
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                results['errors'].append('Unsupported file format. Use CSV or XLSX.')

            # Auto-detect medicine dataset
            if rows:
                headers_lower = {k.lower() for k in rows[0].keys()}
                is_medicine = bool(headers_lower & MEDICINE_COLUMNS)

            if is_medicine:
                from .tasks import process_medicine_csv_chunk
                batch_id = str(uuid.uuid4())
                chunk_size = 500
                chunks = [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)]
                cache.set(f'med_import_{batch_id}', {'processed': 0, 'errors': [], 'total': len(rows)}, 3600)
                for idx, chunk in enumerate(chunks):
                    process_medicine_csv_chunk.delay(json.dumps(chunk), partner.id, batch_id, idx)
                return render(request, 'store/dashboard_product_bulk_upload.html', {
                    'partner': partner,
                    'batch_id': batch_id,
                    'total_rows': len(rows),
                    'chunks': len(chunks),
                })
            else:
                for i, row in enumerate(rows, 1):
                    name = (row.get('name') or '').strip()
                    price = (row.get('price') or '').strip()
                    if not name or not price:
                        results['errors'].append(f'Row {i}: name and price are required.')
                        continue
                    try:
                        category_slug = (row.get('category_slug') or '').strip()
                        category = Category.objects.filter(slug=category_slug).first() if category_slug else None
                        Product.objects.create(
                            partner=partner,
                            name=name,
                            price=Decimal(price),
                            category=category,
                            old_price=Decimal(row['old_price']) if row.get('old_price') else None,
                            stock=int(row.get('stock', 0)),
                            description=(row.get('description') or ''),
                            short_description=(row.get('short_description') or ''),
                            cost_price=Decimal(row['cost_price']) if row.get('cost_price') else None,
                            available=(row.get('available', '1').strip() in ('1', 'true', 'True', 'yes')),
                        )
                        results['created'] += 1
                    except Exception as e:
                        results['errors'].append(f'Row {i}: {e}')
                results['total'] = len(rows)
        except Exception as e:
            results['errors'].append(str(e))

        if not is_medicine:
            return render(request, 'store/dashboard_product_bulk_upload.html', {
                'partner': partner, 'results': results,
            })

    return render(request, 'store/dashboard_product_bulk_upload.html', {
        'partner': partner, 'results': None,
    })


# ════════════════════════════════════════════
# Feature: Customer Order Tracking
# ════════════════════════════════════════════

@login_required
def customer_orders(request):
    orders = Order.objects.filter(user=request.user).prefetch_related('items__product', 'delivery_logs').order_by('-created')
    return render(request, 'store/customer_orders.html', {'orders': orders})


@login_required
def customer_order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    delivery_logs = order.delivery_logs.all().order_by('-created')
    return render(request, 'store/customer_order_detail.html', {
        'order': order,
        'delivery_logs': delivery_logs,
    })


# ════════════════════════════════════════════
# POS (Point of Sale) System
# ════════════════════════════════════════════

@login_required
def dashboard_bulk_import_progress(request, batch_id):
    data = cache.get(f'med_import_{batch_id}', None)
    if data is None:
        return JsonResponse({'error': 'Batch not found'}, status=404)
    return JsonResponse(data)


@login_required
def pos_dashboard(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    if partner.shop_style == 'medicine' and partner.has_medicine_access:
        products = Product.objects.filter(
            Q(partner=partner) | Q(partner__isnull=True),
            available=True, trashed=False, stock__gt=0,
        ).order_by('name')
    else:
        products = Product.objects.filter(partner=partner, available=True, trashed=False, stock__gt=0).order_by('name')
    today_orders = PosOrder.objects.filter(partner=partner, created__date=timezone.now().date())
    today_count = today_orders.count()
    today_total = today_orders.aggregate(t=Sum('total'))['t'] or 0
    return render(request, 'store/pos_dashboard.html', {
        'partner': partner,
        'products': products,
        'today_count': today_count,
        'today_total': today_total,
    })


@login_required
def pos_product_search(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return JsonResponse({'error': 'No seller account'}, status=403)
    q = request.GET.get('q', '').strip()
    if partner.shop_style == 'medicine' and partner.has_medicine_access:
        products = Product.objects.filter(
            Q(partner=partner) | Q(partner__isnull=True),
            available=True, trashed=False, stock__gt=0,
        )
    else:
        products = Product.objects.filter(partner=partner, available=True, trashed=False, stock__gt=0)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q) | Q(barcode__icontains=q))
    products = products.order_by('name')[:30]
    data = [{
        'id': p.id,
        'name': p.name,
        'price': str(p.price),
        'stock': p.stock,
        'barcode': p.barcode or '',
        'image': p.image.url if p.image else '',
    } for p in products]
    return JsonResponse({'products': data})


@login_required
def pos_order_create(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return JsonResponse({'error': 'No seller account'}, status=403)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    items_data = data.get('items', [])
    if not items_data:
        return JsonResponse({'error': 'No items in order'}, status=400)
    subtotal = Decimal('0')
    order_items = []
    for item_data in items_data:
        pid = item_data.get('product_id')
        qty = int(item_data.get('quantity', 1))
        item_discount = Decimal(str(item_data.get('discount', 0)))
        if qty < 1:
            continue
        if partner.shop_style == 'medicine' and partner.has_medicine_access:
            product = get_object_or_404(Product.objects.filter(trashed=False), id=pid)
        else:
            product = get_object_or_404(Product.objects.filter(trashed=False), id=pid, partner=partner)
        if product.stock < qty:
            return JsonResponse({'error': f'Not enough stock for {product.name}'}, status=400)
        line_total = product.price * qty
        subtotal += line_total
        order_items.append({
            'product': product,
            'quantity': qty,
            'item_discount': item_discount,
        })
    order_discount = min(Decimal(str(data.get('discount', 0))), subtotal * Decimal('0.1'))
    tax = Decimal(str(data.get('tax', 0)))
    total = subtotal - order_discount + tax
    customer_name = data.get('customer_name', '').strip()
    customer_phone = data.get('customer_phone', '').strip()
    customer_email = data.get('customer_email', '').strip()
    payment_method = data.get('payment_method', 'cash')
    with transaction.atomic():
        pos_order = PosOrder.objects.create(
            partner=partner,
            customer_name=customer_name,
            customer_phone=customer_phone,
            customer_email=customer_email,
            subtotal=subtotal,
            discount=order_discount,
            tax=tax,
            total=total,
            payment_method=payment_method,
            payment_status='paid',
        )
        for oi in order_items:
            product = oi['product']
            qty = oi['quantity']
            item_discount = oi['item_discount']
            PosOrderItem.objects.create(
                pos_order=pos_order,
                product=product,
                product_name=product.name,
                price=product.price,
                quantity=qty,
                cost_price=product.cost_price,
                discount=item_discount,
            )
            Product.objects.filter(id=product.id, stock__gte=qty).update(stock=F('stock') - qty)
    return JsonResponse({
        'success': True,
        'invoice': pos_order.invoice_number,
        'order_id': pos_order.id,
        'total': str(total),
        'subtotal': str(subtotal),
        'discount': str(order_discount),
        'payment_method': payment_method,
        'seller_name': partner.name,
        'customer_name': customer_name,
        'customer_phone': customer_phone,
        'created': pos_order.created.strftime('%d %b %Y %I:%M %p'),
        'items': [{
            'name': oi['product'].name,
            'price': str(oi['product'].price),
            'qty': oi['quantity'],
            'total': str(oi['product'].price * oi['quantity']),
        } for oi in order_items],
    })


@login_required
def pos_order_list(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    orders = PosOrder.objects.filter(partner=partner).prefetch_related('items').order_by('-created')
    return render(request, 'store/pos_order_list.html', {'partner': partner, 'orders': orders})


@login_required
def pos_order_detail(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    order = get_object_or_404(PosOrder, id=pk, partner=partner)
    return render(request, 'store/pos_order_detail.html', {'partner': partner, 'order': order})


@login_required
def pos_order_refund(request, pk):
    if request.method != 'POST':
        messages.error(request, 'Invalid request.')
        return redirect('pos_order_detail', pk=pk)
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    order = get_object_or_404(PosOrder, id=pk, partner=partner)
    if order.status != 'completed':
        messages.error(request, 'This order cannot be refunded.')
        return redirect('pos_order_detail', pk=pk)
    for item in order.items.all():
        if item.product:
            item.product.stock += item.quantity
            item.product.save(update_fields=['stock'])
    order.status = 'refunded'
    order.save(update_fields=['status'])
    messages.success(request, f'POS order #{order.invoice_number} has been refunded.')
    return redirect('pos_order_list')


@login_required
def pos_order_delete(request, pk):
    if request.method != 'POST':
        messages.error(request, 'Invalid request.')
        return redirect('pos_order_list')
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    order = get_object_or_404(PosOrder, id=pk, partner=partner)
    for item in order.items.all():
        if item.product:
            item.product.stock += item.quantity
            item.product.save(update_fields=['stock'])
    order.delete()
    messages.success(request, f'POS order #{order.invoice_number} has been deleted.')
    return redirect('pos_order_list')


@login_required
def pos_daily_report(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    date_str = request.GET.get('date')
    if date_str:
        try:
            date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date = timezone.now().date()
    else:
        date = timezone.now().date()
    orders = PosOrder.objects.filter(partner=partner, created__date=date).prefetch_related('items')
    total_sales = orders.aggregate(t=Sum('total'))['t'] or 0
    total_items = sum(order.items.aggregate(t=Sum('quantity'))['t'] or 0 for order in orders)
    payment_breakdown = {}
    for method, label in PosOrder.PAYMENT_METHOD_CHOICES:
        pmt = orders.filter(payment_method=method).aggregate(t=Sum('total'))['t'] or 0
        if pmt:
            payment_breakdown[label] = pmt
    return render(request, 'store/pos_report.html', {
        'partner': partner,
        'orders': orders,
        'date': date,
        'total_sales': total_sales,
        'total_items': total_items,
        'order_count': orders.count(),
        'payment_breakdown': payment_breakdown,
    })


# ════════════════════════════════════════════
# Medicine POS System
# ════════════════════════════════════════════

def _check_medicine_pos_access(partner):
    """Check if partner has valid Medicine POS access. Returns (allowed, message, redirect_url)."""
    if partner.shop_style != 'medicine':
        return False, 'Your shop style is not set to Medicine Shop.', None
    if not partner.has_medicine_access:
        return False, 'Medicine catalog access has not been granted by admin.', None
    if not partner.medicine_pos_enabled:
        return False, 'Medicine POS is not enabled. Contact admin to subscribe.', None
    try:
        sub = partner.medicine_subscription
        if sub.is_locked:
            return False, '', 'medicine_subscription_page'
    except MedicineSubscription.DoesNotExist:
        return False, 'Medicine POS subscription not found. Contact admin.', 'medicine_subscription_page'
    return True, '', None


@login_required
def medicine_pos_dashboard(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    allowed, msg, sub_url = _check_medicine_pos_access(partner)
    if not allowed:
        if sub_url:
            return redirect(sub_url)
        messages.error(request, msg)
        return redirect('dashboard')
    products = MedicineProduct.objects.none()  # Start with no products; search will populate via AJAX
    today_orders = MedicinePosOrder.objects.filter(partner=partner, created__date=timezone.now().date())
    today_count = today_orders.count()
    today_total = today_orders.aggregate(t=Sum('total'))['t'] or 0
    sub = MedicineSubscription.objects.filter(partner=partner).first()
    trial_days_remaining = None
    if sub:
        if sub.status == 'trial' and sub.trial_ends_at:
            trial_days_remaining = max(0, (sub.trial_ends_at - timezone.now()).days)
        elif sub.status == 'active' and sub.current_period_end:
            trial_days_remaining = max(0, (sub.current_period_end - timezone.now()).days)
    return render(request, 'store/medicine_pos_dashboard.html', {
        'partner': partner,
        'products': products,
        'today_count': today_count,
        'today_total': today_total,
        'sub': sub,
        'trial_days_remaining': trial_days_remaining,
    })


@login_required
def medicine_pos_product_search(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return JsonResponse({'error': 'No seller account'}, status=403)
    allowed, msg, sub_url = _check_medicine_pos_access(partner)
    if not allowed:
        if sub_url:
            return JsonResponse({'error': 'Subscription required. Please subscribe to continue.', 'redirect': sub_url}, status=403)
        return JsonResponse({'error': msg}, status=403)
    q = request.GET.get('q', '').strip()
    products = MedicineProduct.objects.filter(is_approved=True)
    if q:
        products = products.filter(
            Q(brand_name__icontains=q) | Q(generic_name__icontains=q) |
            Q(strength__icontains=q) | Q(sku__icontains=q) | Q(dosage_form__icontains=q)
        )
    products = products.order_by('brand_name')[:30]
    data = [{
        'id': p.id,
        'name': str(p),
        'brand_name': p.brand_name,
        'generic_name': p.generic_name,
        'strength': p.strength,
        'dosage_form': p.dosage_form,
        'price': str(p.price),
        'stock': p.stock,
        'image': p.image.url if p.image else '',
    } for p in products]
    return JsonResponse({'products': data})


@login_required
def medicine_pos_order_create(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return JsonResponse({'error': 'No seller account'}, status=403)
    allowed, msg, sub_url = _check_medicine_pos_access(partner)
    if not allowed:
        if sub_url:
            return JsonResponse({'error': 'Subscription required.', 'redirect': sub_url}, status=403)
        return JsonResponse({'error': msg}, status=403)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    items_data = data.get('items', [])
    if not items_data:
        return JsonResponse({'error': 'No items in order'}, status=400)
    subtotal = Decimal('0')
    order_items = []
    for item_data in items_data:
        pid = item_data.get('product_id')
        qty = int(item_data.get('quantity', 1))
        if qty < 1:
            continue
        item_discount = Decimal(str(item_data.get('discount', 0)))
        product = get_object_or_404(MedicineProduct.objects.filter(is_approved=True), id=pid)

        line_total = product.price * qty
        subtotal += line_total
        order_items.append({
            'product': product,
            'quantity': qty,
            'item_discount': item_discount,
        })
    order_discount = min(Decimal(str(data.get('discount', 0))), subtotal * Decimal('0.1'))
    total = subtotal - order_discount
    customer_name = data.get('customer_name', '').strip()
    customer_phone = data.get('customer_phone', '').strip()
    payment_method = data.get('payment_method', 'cash')
    with transaction.atomic():
        pos_order = MedicinePosOrder.objects.create(
            partner=partner,
            customer_name=customer_name,
            customer_phone=customer_phone,
            subtotal=subtotal,
            discount=order_discount,
            total=total,
            payment_method=payment_method,
            payment_status='paid',
        )
        for oi in order_items:
            product = oi['product']
            qty = oi['quantity']
            item_discount = oi['item_discount']
            MedicinePosOrderItem.objects.create(
                pos_order=pos_order,
                product=product,
                product_name=str(product),
                price=product.price,
                quantity=qty,
                discount=item_discount,
            )
    return JsonResponse({
        'success': True,
        'invoice': pos_order.invoice_number,
        'order_id': pos_order.id,
        'total': str(total),
        'subtotal': str(subtotal),
        'discount': str(order_discount),
        'payment_method': payment_method,
        'seller_name': partner.name,
        'customer_name': customer_name,
        'customer_phone': customer_phone,
        'created': pos_order.created.strftime('%d %b %Y %I:%M %p'),
        'items': [{
            'name': str(oi['product']),
            'price': str(oi['product'].price),
            'qty': oi['quantity'],
            'total': str(oi['product'].price * oi['quantity']),
        } for oi in order_items],
    })


@login_required
def medicine_pos_order_list(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    allowed, msg, sub_url = _check_medicine_pos_access(partner)
    if not allowed:
        if sub_url:
            return redirect(sub_url)
        messages.error(request, msg)
        return redirect('dashboard')
    orders = MedicinePosOrder.objects.filter(partner=partner).prefetch_related('items').order_by('-created')
    return render(request, 'store/medicine_pos_order_list.html', {'partner': partner, 'orders': orders})


@login_required
def medicine_pos_order_detail(request, pk):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    order = get_object_or_404(MedicinePosOrder, id=pk, partner=partner)
    return render(request, 'store/medicine_pos_order_detail.html', {'partner': partner, 'order': order})


@login_required
def medicine_pos_order_refund(request, pk):
    if request.method != 'POST':
        messages.error(request, 'Invalid request.')
        return redirect('medicine_pos_order_detail', pk=pk)
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    order = get_object_or_404(MedicinePosOrder, id=pk, partner=partner)
    if order.status != 'completed':
        messages.error(request, 'This order cannot be refunded.')
        return redirect('medicine_pos_order_detail', pk=pk)
    for item in order.items.all():
        if item.product:
            item.product.stock += item.quantity
            item.product.save(update_fields=['stock'])
    order.status = 'refunded'
    order.save(update_fields=['status'])
    messages.success(request, f'Medicine POS order #{order.invoice_number} has been refunded.')
    return redirect('medicine_pos_order_list')


@login_required
def medicine_pos_order_delete(request, pk):
    if request.method != 'POST':
        messages.error(request, 'Invalid request.')
        return redirect('medicine_pos_order_list')
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    order = get_object_or_404(MedicinePosOrder, id=pk, partner=partner)
    for item in order.items.all():
        if item.product:
            item.product.stock += item.quantity
            item.product.save(update_fields=['stock'])
    order.delete()
    messages.success(request, f'Medicine POS order #{order.invoice_number} has been deleted.')
    return redirect('medicine_pos_order_list')


@login_required
def medicine_pos_daily_report(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    date_str = request.GET.get('date')
    if date_str:
        try:
            date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date = timezone.now().date()
    else:
        date = timezone.now().date()
    orders = MedicinePosOrder.objects.filter(partner=partner, created__date=date).prefetch_related('items')
    total_sales = orders.aggregate(t=Sum('total'))['t'] or 0
    total_items = sum(order.items.aggregate(t=Sum('quantity'))['t'] or 0 for order in orders)
    payment_breakdown = {}
    for method, label in MedicinePosOrder.PAYMENT_METHOD_CHOICES:
        pmt = orders.filter(payment_method=method).aggregate(t=Sum('total'))['t'] or 0
        if pmt:
            payment_breakdown[label] = pmt
    return render(request, 'store/medicine_pos_report.html', {
        'partner': partner,
        'orders': orders,
        'date': date,
        'total_sales': total_sales,
        'total_items': total_items,
        'order_count': orders.count(),
        'payment_breakdown': payment_breakdown,
    })


@login_required
def medicine_request_product(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    if not partner.has_medicine_access:
        messages.error(request, 'Medicine access not granted.')
        return redirect('dashboard')
    if request.method == 'POST':
        brand_name = request.POST.get('brand_name', '').strip()
        generic_name = request.POST.get('generic_name', '').strip()
        strength = request.POST.get('strength', '').strip()
        dosage_form = request.POST.get('dosage_form', '').strip()
        price = request.POST.get('price', '').strip()
        description = request.POST.get('description', '').strip()
        if not brand_name or not price:
            messages.error(request, 'Brand name and price are required.')
            return render(request, 'store/medicine_request_product.html', {'partner': partner})
        try:
            price = Decimal(price)
        except Exception:
            messages.error(request, 'Invalid price.')
            return render(request, 'store/medicine_request_product.html', {'partner': partner})
        MedicineProduct.objects.create(
            brand_name=brand_name,
            generic_name=generic_name,
            strength=strength,
            dosage_form=dosage_form,
            price=price,
            description=description,
            is_approved=False,
            requested_by=partner,
        )
        messages.success(request, f'Medicine product "{brand_name}" submitted for admin approval.')
        return redirect('dashboard')
    return render(request, 'store/medicine_request_product.html', {'partner': partner})


@login_required
def medicine_subscription_page(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')

    sub, _ = MedicineSubscription.objects.get_or_create(partner=partner)

    wallet, _ = PartnerWallet.objects.get_or_create(partner=partner)
    wallet_balance = wallet.available_balance
    packages = SubscriptionPackage.objects.filter(is_active=True).order_by('sort_order', 'name')

    if request.method == 'POST':
        action = request.POST.get('action')
        package_id = request.POST.get('package_id')
        package = get_object_or_404(SubscriptionPackage, id=package_id, is_active=True)

        if action == 'pay_wallet':
            price = int(package.price)
            if wallet_balance < price:
                messages.error(request, f'Insufficient balance. You need ৳{price} but your available balance is ৳{wallet_balance}.')
                return redirect('medicine_subscription_page')

            with transaction.atomic():
                wallet.balance -= price
                wallet.total_fees_paid += price
                wallet.save(update_fields=['balance', 'total_fees_paid'])

                WalletTransaction.objects.create(
                    wallet=wallet,
                    amount=price,
                    balance_before=wallet_balance,
                    balance_after=wallet.balance,
                    type='subscription_payment',
                    description=f'Medicine POS subscription — {package.name}',
                )

                platform, _ = PlatformBalance.objects.get_or_create(pk=1)
                platform.balance += price
                platform.save(update_fields=['balance'])

                now = timezone.now()
                sub.status = 'active'
                sub.current_period_start = now
                sub.current_period_end = now + timedelta(days=package.duration_days)
                sub.last_payment_at = now
                sub.next_payment_due = sub.current_period_end
                sub.save()

            messages.success(request, f'Payment successful! Your Medicine POS "{package.name}" subscription is active for {package.duration_days} days.')
            return redirect('medicine_pos_dashboard')

        elif action == 'manual_pay':
            messages.info(request, 'Please send the payment to the admin via the methods below, then they will activate your subscription.')
            return redirect('wallet_recharge')

    trial_remaining = 0
    if sub.status == 'trial' and sub.trial_ends_at:
        trial_remaining = max(0, (sub.trial_ends_at - timezone.now()).days)
    elif sub.status == 'active' and sub.current_period_end:
        trial_remaining = max(0, (sub.current_period_end - timezone.now()).days)

    instructions = WalletRechargeInstruction.objects.filter(is_active=True).order_by('order', 'title')

    return render(request, 'store/medicine_subscription.html', {
        'partner': partner,
        'sub': sub,
        'wallet_balance': wallet_balance,
        'trial_remaining': trial_remaining,
        'packages': packages,
        'instructions': instructions,
    })


@login_required
def wallet_recharge(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        messages.error(request, 'No seller account found.')
        return redirect('dashboard')
    instructions = WalletRechargeInstruction.objects.filter(is_active=True).order_by('order', 'title')
    return render(request, 'store/wallet_recharge.html', {
        'partner': partner,
        'instructions': instructions,
    })


# ════════════════════════════════════════════
# Messaging System
# ════════════════════════════════════════════

def _unread_message_count(user):
    """Count unread messages for a user across all their conversations."""
    convs = Conversation.objects.filter(
        Q(participants=user) | Q(is_broadcast=True)
    )
    try:
        partner = user.partner
        convs |= Conversation.objects.filter(partner=partner)
    except Partner.DoesNotExist:
        pass
    convs = convs.distinct()
    total = 0
    for conv in convs:
        try:
            status = ConversationReadStatus.objects.get(conversation=conv, user=user)
            last_read = status.last_read
        except ConversationReadStatus.DoesNotExist:
            last_read = timezone.now() - timezone.timedelta(days=365)
        total += conv.messages.filter(created__gt=last_read).exclude(sender=user).count()
    return total


# ─── Customer Messaging ───

@login_required
def customer_message_list(request):
    conversations = Conversation.objects.filter(
        Q(participants=request.user) | Q(is_broadcast=True)
    ).distinct().order_by('-updated')
    conv_data = []
    for conv in conversations:
        try:
            status = ConversationReadStatus.objects.get(conversation=conv, user=request.user)
            last_read = status.last_read
        except ConversationReadStatus.DoesNotExist:
            last_read = timezone.now() - timezone.timedelta(days=365)
        unread = conv.messages.filter(created__gt=last_read).exclude(sender=request.user).count()
        conv_data.append({'conversation': conv, 'unread_count': unread})
    return render(request, 'store/message_list.html', {
        'conversations': conv_data,
    })


@login_required
def customer_message_create(request, partner_id):
    partner = get_object_or_404(Partner, id=partner_id)
    product = None
    product_id = request.GET.get('product') or request.POST.get('product_id')
    if product_id:
        try:
            product = Product.objects.get(id=int(product_id), partner=partner, trashed=False)
        except (ValueError, Product.DoesNotExist):
            pass
    # Check for existing conversation about this product
    existing = Conversation.objects.filter(participants=request.user, partner=partner, product=product).first()
    if existing:
        return redirect('customer_message_detail', conv_id=existing.id)
    initial_subject = f'Question about: {product.name}' if product else ''
    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        content = request.POST.get('content', '').strip()
        if not subject or not content:
            return render(request, 'store/message_create.html', {
                'partner': partner, 'product': product,
                'error': 'Subject and message are required.'
            })
        conv = Conversation.objects.create(
            partner=partner,
            product=product,
            subject=subject or initial_subject,
        )
        conv.participants.add(request.user)
        if partner.user:
            conv.participants.add(partner.user)
        Message.objects.create(conversation=conv, sender=request.user, content=content)
        ConversationReadStatus.objects.create(conversation=conv, user=request.user)
        if partner.user:
            ConversationReadStatus.objects.create(conversation=conv, user=partner.user)
        Notification.objects.create(
            user=partner.user,
            partner=partner,
            type='admin_message',
            title=f'New message from {request.user.username}',
            message=content[:100],
            link=f'/dashboard/messages/{conv.id}/',
        )
        messages.success(request, 'Your message has been sent.')
        return redirect('customer_message_detail', conv_id=conv.id)
    return render(request, 'store/message_create.html', {
        'partner': partner, 'product': product,
        'initial_subject': initial_subject,
    })


@login_required
def customer_message_detail(request, conv_id):
    conv = get_object_or_404(Conversation, id=conv_id, participants=request.user)
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        if content:
            Message.objects.create(conversation=conv, sender=request.user, content=content)
            conv.save(update_fields=['updated'])  # touch updated timestamp
            # notify partner
            other_users = conv.participants.exclude(id=request.user.id)
            for u in other_users:
                Notification.objects.create(
                    user=u,
                    type='admin_message',
                    title=f'New reply from {request.user.username}',
                    message=content[:100],
                    link=f'/messages/{conv.id}/',
                )
        return redirect('customer_message_detail', conv_id=conv.id)
    # Mark as read
    status, _ = ConversationReadStatus.objects.get_or_create(
        conversation=conv, user=request.user,
        defaults={'last_read': timezone.now()}
    )
    status.last_read = timezone.now()
    status.save(update_fields=['last_read'])
    messages_list = conv.messages.all()
    return render(request, 'store/message_detail.html', {
        'conversation': conv,
        'messages': messages_list,
    })


# ─── Partner Messaging ───

@login_required
def partner_message_list(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    conversations = Conversation.objects.filter(
        Q(participants=request.user) | Q(partner=partner) | Q(is_broadcast=True)
    ).distinct().order_by('-updated')
    conv_data = []
    for conv in conversations:
        try:
            status = ConversationReadStatus.objects.get(conversation=conv, user=request.user)
            last_read = status.last_read
        except ConversationReadStatus.DoesNotExist:
            last_read = timezone.now() - timezone.timedelta(days=365)
        unread = conv.messages.filter(created__gt=last_read).exclude(sender=request.user).count()
        conv_data.append({'conversation': conv, 'unread_count': unread})
    return render(request, 'store/dashboard_messages.html', {
        'partner': partner,
        'conversations': conv_data,
    })


@login_required
def partner_message_detail(request, conv_id):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    conv = get_object_or_404(
        Conversation.objects.filter(
            Q(participants=request.user) | Q(partner=partner) | Q(is_broadcast=True)
        ).distinct(),
        id=conv_id
    )
    # Ensure partner user is a participant
    if request.user not in conv.participants.all():
        conv.participants.add(request.user)
    # Broadcasts are read-only for partners (no reply)
    is_readonly = conv.is_broadcast and not conv.messages.filter(sender=request.user).exists()
    if request.method == 'POST' and not is_readonly:
        content = request.POST.get('content', '').strip()
        if content:
            Message.objects.create(conversation=conv, sender=request.user, content=content)
            conv.save(update_fields=['updated'])
            # notify other participants
            for u in conv.participants.exclude(id=request.user.id):
                Notification.objects.create(
                    user=u,
                    type='admin_message',
                    title=f'New reply from {partner.name}',
                    message=content[:100],
                    link=f'/dashboard/messages/{conv.id}/',
                )
        return redirect('partner_message_detail', conv_id=conv.id)
    # Mark as read
    status, _ = ConversationReadStatus.objects.get_or_create(
        conversation=conv, user=request.user,
        defaults={'last_read': timezone.now()}
    )
    status.last_read = timezone.now()
    status.save(update_fields=['last_read'])
    return render(request, 'store/dashboard_message_detail.html', {
        'partner': partner,
        'conversation': conv,
        'messages': conv.messages.all(),
        'is_readonly': is_readonly,
    })


# ─── Admin Messaging ───

@user_passes_test(lambda u: u.is_staff)
def admin_message_list(request):
    conversations = Conversation.objects.all().order_by('-updated')
    conv_data = []
    for conv in conversations:
        try:
            status = ConversationReadStatus.objects.get(conversation=conv, user=request.user)
            last_read = status.last_read
        except ConversationReadStatus.DoesNotExist:
            last_read = timezone.now() - timezone.timedelta(days=365)
        unread = conv.messages.filter(created__gt=last_read).exclude(sender=request.user).count()
        conv_data.append({'conversation': conv, 'unread_count': unread})
    return render(request, 'store/admin_message_list.html', {
        'conversations': conv_data,
    })


@user_passes_test(lambda u: u.is_staff)
def admin_message_detail(request, conv_id):
    conv = get_object_or_404(Conversation, id=conv_id)
    # Ensure admin is participant
    if request.user not in conv.participants.all():
        conv.participants.add(request.user)
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        if content:
            Message.objects.create(conversation=conv, sender=request.user, content=content)
            conv.save(update_fields=['updated'])
            for u in conv.participants.exclude(id=request.user.id):
                Notification.objects.create(
                    user=u,
                    type='admin_message',
                    title=f'Admin reply: {conv.subject}',
                    message=content[:100],
                    link=f'/dashboard/messages/{conv.id}/',
                )
        return redirect('admin_message_detail', conv_id=conv.id)
    status, _ = ConversationReadStatus.objects.get_or_create(
        conversation=conv, user=request.user,
        defaults={'last_read': timezone.now()}
    )
    status.last_read = timezone.now()
    status.save(update_fields=['last_read'])
    return render(request, 'store/admin_message_detail.html', {
        'conversation': conv,
        'messages': conv.messages.all(),
    })


@user_passes_test(lambda u: u.is_staff)
def admin_broadcast_create(request):
    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        content = request.POST.get('content', '').strip()
        if not subject or not content:
            return render(request, 'store/admin_broadcast_form.html', {
                'error': 'Subject and message are required.'
            })
        conv = Conversation.objects.create(
            subject=subject,
            is_broadcast=True,
        )
        conv.participants.add(request.user)
        Message.objects.create(conversation=conv, sender=request.user, content=content)
        ConversationReadStatus.objects.create(conversation=conv, user=request.user)
        # Notify all partners
        for partner in Partner.objects.filter(user__isnull=False, is_active=True).select_related('user'):
            Notification.objects.create(
                user=partner.user,
                partner=partner,
                type='admin_message',
                title=f'Admin Broadcast: {subject}',
                message=content[:100],
                link='/dashboard/messages/',
            )
        messages.success(request, f'Broadcast sent to all partners.')
        return redirect('admin_message_detail', conv_id=conv.id)
    return render(request, 'store/admin_broadcast_form.html')


# ════════════════════════════════════════════
# Product Q&A (Public)
# ════════════════════════════════════════════

@login_required
def product_ask_question(request, product_id):
    product = get_object_or_404(Product.objects.filter(trashed=False), id=product_id)
    if request.method == 'POST':
        question = request.POST.get('question', '').strip()
        if not question:
            messages.error(request, 'Please enter your question.')
        else:
            ProductQA.objects.create(
                product=product,
                user=request.user,
                question=question,
            )
            # Notify the product's partner
            if product.partner and product.partner.user:
                Notification.objects.create(
                    user=product.partner.user,
                    partner=product.partner,
                    type='admin_message',
                    title=f'New question about {product.name}',
                    message=question[:100],
                    link=f'/dashboard/qa/',
                )
            messages.success(request, 'Your question has been submitted.')
        return redirect('product_detail', slug=product.slug)


@login_required
def product_answer_question(request, qa_id):
    qa = get_object_or_404(ProductQA, id=qa_id)
    product = qa.product
    # Only the product's partner or admin can answer
    is_owner = product.partner and request.user == product.partner.user
    if not (is_owner or request.user.is_staff):
        return redirect('product_detail', slug=product.slug)
    if request.method == 'POST':
        answer = request.POST.get('answer', '').strip()
        if answer:
            qa.answer = answer
            qa.answered_by = request.user
            qa.answered_at = timezone.now()
            qa.save(update_fields=['answer', 'answered_by', 'answered_at'])
            messages.success(request, 'Your answer has been posted.')
        return redirect('product_detail', slug=product.slug)
    return redirect('product_detail', slug=product.slug)


@login_required
def partner_qa_list(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    unanswered = ProductQA.objects.filter(product__partner=partner, answer='').order_by('-created')
    answered = ProductQA.objects.filter(product__partner=partner).exclude(answer='').order_by('-answered_at')
    return render(request, 'store/dashboard_qa.html', {
        'partner': partner,
        'unanswered': unanswered,
        'answered': answered,
    })


# ─── Support Ticket Views ───

@login_required
def dashboard_tickets(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    tickets = SupportTicket.objects.filter(partner=partner).annotate(
        reply_count=Count('replies'),
        last_reply_time=Max('replies__created'),
    ).order_by('-created')
    return render(request, 'store/dashboard_tickets.html', {
        'partner': partner,
        'tickets': tickets,
        'open_count': tickets.filter(status='open').count(),
        'resolved_count': tickets.filter(status='resolved').count(),
    })


@login_required
def dashboard_ticket_create(request):
    try:
        partner = request.user.partner
    except Partner.DoesNotExist:
        return redirect('dashboard')
    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        description = request.POST.get('description', '').strip()
        priority = request.POST.get('priority', 'medium')
        image = request.FILES.get('image')
        if not subject or not description:
            messages.error(request, 'Subject and description are required.')
        else:
            ticket = SupportTicket.objects.create(
                partner=partner,
                subject=subject,
                description=description,
                priority=priority,
                image=image,
            )
            # Notify all staff users
            staff_users = User.objects.filter(is_staff=True)
            for staff in staff_users:
                Notification.objects.create(
                    user=staff,
                    type='ticket_update',
                    title=f'New support ticket: {subject}',
                    message=description[:100],
                    link='/dashboard/admin/tickets/',
                )
            messages.success(request, 'Your ticket has been submitted. We will get back to you soon.')
            return redirect('dashboard_tickets')
    return render(request, 'store/dashboard_ticket_form.html', {
        'partner': partner,
    })


@login_required
def dashboard_ticket_detail(request, pk):
    ticket = get_object_or_404(SupportTicket, id=pk, partner__user=request.user)
    if request.method == 'POST':
        message = request.POST.get('message', '').strip()
        attachment = request.FILES.get('attachment')
        if message:
            TicketReply.objects.create(
                ticket=ticket,
                user=request.user,
                message=message,
                attachment=attachment,
            )
            if ticket.status == 'resolved' or ticket.status == 'closed':
                ticket.status = 'open'
                ticket.save(update_fields=['status'])
            # Notify staff
            staff_users = User.objects.filter(is_staff=True)
            for staff in staff_users:
                Notification.objects.create(
                    user=staff,
                    type='ticket_update',
                    title=f'New reply on ticket #{ticket.id}: {ticket.subject}',
                    message=message[:100],
                    link=f'/dashboard/admin/tickets/{ticket.id}/',
                )
            messages.success(request, 'Your reply has been posted.')
        return redirect('dashboard_ticket_detail', pk=pk)
    replies = ticket.replies.select_related('user').order_by('created')
    return render(request, 'store/dashboard_ticket_detail.html', {
        'partner': ticket.partner,
        'ticket': ticket,
        'replies': replies,
    })


@login_required
def admin_ticket_list(request):
    if not request.user.is_staff:
        return redirect('dashboard')
    status_filter = request.GET.get('status', '')
    tickets = SupportTicket.objects.all().select_related('partner').annotate(
        reply_count=Count('replies'),
        last_reply_time=Max('replies__created'),
    ).order_by('-priority', '-created')
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    return render(request, 'store/admin_ticket_list.html', {
        'tickets': tickets,
        'current_status': status_filter,
        'open_count': SupportTicket.objects.filter(status='open').count(),
        'in_progress_count': SupportTicket.objects.filter(status='in_progress').count(),
        'resolved_count': SupportTicket.objects.filter(status='resolved').count(),
        'closed_count': SupportTicket.objects.filter(status='closed').count(),
    })


@login_required
def admin_ticket_detail(request, pk):
    if not request.user.is_staff:
        return redirect('dashboard')
    ticket = get_object_or_404(SupportTicket, id=pk)
    if request.method == 'POST':
        action = request.POST.get('action', '')
        message = request.POST.get('message', '').strip()
        attachment = request.FILES.get('attachment')
        new_status = request.POST.get('status', '')
        if new_status and new_status in dict(SupportTicket.STATUS_CHOICES):
            ticket.status = new_status
            ticket.save(update_fields=['status'])
            # Notify partner
            if ticket.partner and ticket.partner.user:
                Notification.objects.create(
                    user=ticket.partner.user,
                    partner=ticket.partner,
                    type='ticket_update',
                    title=f'Ticket #{ticket.id} status updated to {ticket.get_status_display()}',
                    message=f'Your ticket "{ticket.subject}" has been updated.',
                    link=f'/dashboard/tickets/{ticket.id}/',
                )
            messages.success(request, f'Ticket status updated to {ticket.get_status_display()}.')
        if message:
            TicketReply.objects.create(
                ticket=ticket,
                user=request.user,
                message=message,
                attachment=attachment,
            )
            if ticket.partner and ticket.partner.user:
                Notification.objects.create(
                    user=ticket.partner.user,
                    partner=ticket.partner,
                    type='ticket_update',
                    title=f'Admin replied to your ticket #{ticket.id}',
                    message=message[:100],
                    link=f'/dashboard/tickets/{ticket.id}/',
                )
            messages.success(request, 'Your reply has been posted.')
        return redirect('admin_ticket_detail', pk=pk)
    replies = ticket.replies.select_related('user').order_by('created')
    return render(request, 'store/admin_ticket_detail.html', {
        'ticket': ticket,
        'replies': replies,
    })
